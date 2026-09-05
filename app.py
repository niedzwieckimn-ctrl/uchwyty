# -*- coding: utf-8 -*-
import os
import io
import html
import csv
import base64
import re
import json
import hashlib
import glob
import sqlite3
import socket
import time
import threading
import logging
import uuid
import secrets
import hmac
import urllib.parse
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

from flask import (
    Flask, request, redirect, url_for, jsonify, session, g,
    send_file, abort
)
from flask import render_template, render_template_string
from jinja2 import ChoiceLoader, DictLoader, FileSystemLoader
from werkzeug.security import check_password_hash

import qrcode
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from pypdf import PdfReader, PdfWriter
from ksef_module import (
    build_ksef_draft_xml as _legacy_build_ksef_draft_xml,
    validate_fa3_xml, validate_ksef_invoice as _legacy_validate_ksef_invoice,
    xml_filename,
)
from invoice_types import normalize_invoice_type, resolve_invoice_type
import invoice_domestic
import invoice_foreign
import ksef_domestic
import ksef_foreign
from cash_flow_module import register_cash_flow, cash_flow_overdue_invoices
from inventory_analytics import build_replenishment_analysis, recommended_replenishments
from seventeentrack_module import SeventeenTrackClient, enabled as seventeentrack_is_enabled, map_package_status, monotonic_status, parse_tracking_payload, verify_webhook_signature
from proforma_module import generate_proforma_pdf
from inpost_module import (
    InPostError,
    config_summary as inpost_config_summary,
    create_courier_shipment,
    create_dispatch_order as inpost_create_dispatch_order,
    get_label as inpost_get_label,
    get_shipment as inpost_get_shipment,
)
try:
    from ksef_api import ksef_config_summary, send_invoice_to_ksef
except Exception:
    send_invoice_to_ksef = None

    def ksef_config_summary():
        return {"configured": False, "missing": ["ksef_api.py"], "env": "", "base_url": ""}

_EMAIL_IMPORT_ERROR = ""
try:
    from email_module import (
        email_config_summary,
        send_email,
        send_order_confirmation,
        send_invoice_available,
        send_payment_reminder,
    )
except Exception as exc:
    _EMAIL_IMPORT_ERROR = str(exc)
    send_email = None
    send_order_confirmation = None
    send_invoice_available = None
    send_payment_reminder = None

    def email_config_summary():
        return {
            "configured": False,
            "missing": ["email_module.py"],
            "enabled": False,
            "import_error": _EMAIL_IMPORT_ERROR,
        }


# =========================
# KONFIG
# =========================

# TWOJE IP (z ipconfig -> IPv4)
BASE_URL = "http://192.168.68.103:5000"

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(APP_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "app.db")

os.makedirs(DATA_DIR, exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20 MB
app.config["JSON_AS_ASCII"] = False

PERF_LOG_ENABLED = (os.environ.get("PERF_LOG_ENABLED") or "1").strip().lower() in ("1", "true", "yes", "on")
if PERF_LOG_ENABLED:
    app.logger.setLevel(logging.INFO)
_raw_render_template_string = render_template_string


def _perf_add(stage: str, elapsed_seconds: float):
    try:
        if hasattr(g, "perf_stages"):
            g.perf_stages[stage] = g.perf_stages.get(stage, 0.0) + elapsed_seconds
    except RuntimeError:
        pass


def render_template_string(*args, **kwargs):
    started = time.perf_counter()
    try:
        return _raw_render_template_string(*args, **kwargs)
    finally:
        _perf_add("render_html", time.perf_counter() - started)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)


_MOJIBAKE_REPLACEMENTS = {
    "Ä…": "ą", "Ä‡": "ć", "Ä™": "ę", "Ĺ‚": "ł", "Ĺ„": "ń",
    "Ăł": "ó", "Ĺ›": "ś", "Ĺş": "ź", "ĹĽ": "ż",
    "Ä„": "Ą", "Ä†": "Ć", "Ä": "Ę", "Ĺ": "Ł", "Ĺƒ": "Ń",
    "Ă“": "Ó", "Ĺš": "Ś", "Ĺą": "Ź", "Ĺ»": "Ż",
    "Ã³": "ó", "Å‚": "ł", "Å„": "ń", "Å›": "ś", "Åº": "ź",
    "Å¼": "ż", "Å": "Ł", "Åƒ": "Ń", "Åš": "Ś", "Å¹": "Ź",
    "Å»": "Ż", "Ä": "ą", "Ä": "ć", "Ä": "ę",
    "â€˘": "•", "â€¢": "•", "â€“": "–", "â€”": "—",
    "â€ž": "„", "â€ť": "”", "â€ś": "“", "â€™": "’",
    "â†": "←", "â†’": "→",
}


def fix_polish_mojibake(text: str) -> str:
    if not text or not any(marker in text for marker in ("Ä", "Ĺ", "Ă", "Å", "Ã", "â")):
        return text
    for bad, good in _MOJIBAKE_REPLACEMENTS.items():
        text = text.replace(bad, good)
    return text


@app.after_request
def force_utf8_html(response):
    if response.direct_passthrough:
        return response
    if response.mimetype in {"text/html", "text/plain", "application/json"}:
        response.headers["Content-Type"] = f"{response.mimetype}; charset=utf-8"
    if response.mimetype == "text/html":
        body = response.get_data(as_text=True)
        fixed = fix_polish_mojibake(body)
        if fixed != body:
            response.set_data(fixed)
    return response


def _detect_lan_base_url(port: int) -> str:
    try:
        sck = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sck.connect(("8.8.8.8", 80))
        ip = sck.getsockname()[0]
        sck.close()
        if ip and not ip.startswith("127."):
            return f"http://{ip}:{port}"
    except Exception:
        pass
    return ""


def build_public_url(path: str) -> str:
    # Dla QR preferuj adres LAN; jeĹ›li aplikacja jest otwarta lokalnie,
    # sprĂłbuj wykryÄ‡ LAN IP automatycznie (bardziej niezawodne niĹĽ staĹ‚y BASE_URL).
    base_cfg = (BASE_URL or "").rstrip("/")
    try:
        host = (request.host or "").split(":")[0].lower()
        req_base = (request.host_url or "").rstrip("/")
        req_port = to_int((request.host or "").split(":")[1] if ":" in (request.host or "") else 5000, 5000)
    except RuntimeError:
        host = ""
        req_base = ""
        req_port = 5000

    if host in {"localhost", "127.0.0.1", "::1", "0.0.0.0"}:
        base = _detect_lan_base_url(req_port) or base_cfg or req_base
    else:
        base = req_base or base_cfg or _detect_lan_base_url(req_port)

    return f"{base}{path}"


# =========================
# DB
# =========================

def conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c

def init_db():
    c = conn()
    cur = c.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS products(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sku TEXT UNIQUE NOT NULL,
        model TEXT,
        ean TEXT,
        name TEXT,
        archived INTEGER NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS stock(
        product_id INTEGER PRIMARY KEY,
        qty INTEGER NOT NULL DEFAULT 0,
        FOREIGN KEY(product_id) REFERENCES products(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS customers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        address TEXT,
        phone TEXT,
        email TEXT,
        nip TEXT,
        language TEXT NOT NULL DEFAULT 'pl',
        price_list TEXT NOT NULL DEFAULT 'pln',
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS orders(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_no TEXT UNIQUE NOT NULL,
        customer_id INTEGER,
        customer_name TEXT NOT NULL,
        customer_address TEXT,
        customer_phone TEXT,
        customer_email TEXT,
        status TEXT NOT NULL DEFAULT 'new', -- new/packed/shipped/cancelled
        note TEXT,
        created_at TEXT NOT NULL,
        warehouse_issued INTEGER NOT NULL DEFAULT 0,
        currency TEXT NOT NULL DEFAULT 'PLN',
        price_list TEXT NOT NULL DEFAULT 'pln',
        tracking_no TEXT,
        carrier TEXT,
        packed_at TEXT,
        shipped_at TEXT,
        FOREIGN KEY(customer_id) REFERENCES customers(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS order_items(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        product_id INTEGER NOT NULL,
        sku TEXT NOT NULL,
        qty INTEGER NOT NULL,
        unit_net_price REAL,
        unit_gross_price REAL,
        unit_retail_price REAL,
        currency TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY(order_id) REFERENCES orders(id),
        FOREIGN KEY(product_id) REFERENCES products(id)
    )
    """)

    # Paczki z Chin (prosty moduĹ‚ na start)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS china_packages(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        package_no TEXT UNIQUE NOT NULL,
        status TEXT NOT NULL DEFAULT 'planned', -- planned/ordered/shipped/arrived
        tracking TEXT,
        note TEXT,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS china_items(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        package_id INTEGER NOT NULL,
        product_id INTEGER NOT NULL,
        sku TEXT NOT NULL,
        qty INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(package_id) REFERENCES china_packages(id),
        FOREIGN KEY(product_id) REFERENCES products(id)
    )
    """)

    # Metadane interfejsu magazynu. Te tabele nie uczestniczą w obliczaniu
    # stanów, rezerwacji ani dostaw.
    cur.execute("""
    CREATE TABLE IF NOT EXISTS product_images(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stored_path TEXT UNIQUE NOT NULL,
        filename TEXT NOT NULL,
        created_at TEXT NOT NULL
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS product_image_assignments(
        product_id INTEGER PRIMARY KEY,
        image_id INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(product_id) REFERENCES products(id),
        FOREIGN KEY(image_id) REFERENCES product_images(id)
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_product_image_assignments_image ON product_image_assignments(image_id)")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS stock_adjustments(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id INTEGER NOT NULL,
        old_qty INTEGER NOT NULL,
        new_qty INTEGER NOT NULL,
        delta INTEGER NOT NULL,
        mode TEXT NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(product_id) REFERENCES products(id)
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_stock_adjustments_product ON stock_adjustments(product_id, id DESC)")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS china_documents(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        package_id INTEGER NOT NULL,
        original_name TEXT NOT NULL,
        document_type TEXT NOT NULL DEFAULT 'order',
        stored_path TEXT NOT NULL,
        size_bytes INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(package_id) REFERENCES china_packages(id)
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_china_documents_package ON china_documents(package_id, id)")

    # Niezależny, trwały bezpiecznik idempotencji. Rekord P/O może być
    # synchronizowany z chmurą, ale tej samej dostawy nie wolno przyjąć drugi raz.
    cur.execute("""
    CREATE TABLE IF NOT EXISTS china_stock_receipts(
        package_id INTEGER PRIMARY KEY,
        received_at TEXT NOT NULL,
        quantities_json TEXT NOT NULL,
        FOREIGN KEY(package_id) REFERENCES china_packages(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS pricing(
        model TEXT PRIMARY KEY,
        net_price REAL NOT NULL DEFAULT 0,
        gross_price REAL NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS company_profile(
        id INTEGER PRIMARY KEY CHECK(id=1),
        company_name TEXT,
        address TEXT,
        nip TEXT,
        phone TEXT,
        email TEXT,
        bank_account TEXT,
        bank_swift TEXT,
        updated_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS invoices(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        invoice_no TEXT NOT NULL,
        issue_date TEXT NOT NULL,
        sell_date TEXT NOT NULL,
        payment_type TEXT NOT NULL,
        payment_to TEXT,
        buyer_name TEXT,
        buyer_tax_no TEXT,
        buyer_street TEXT,
        buyer_post_code TEXT,
        buyer_city TEXT,
        buyer_country TEXT,
        buyer_email TEXT,
        buyer_phone TEXT,
        total_net REAL NOT NULL DEFAULT 0,
        total_gross REAL NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL,
        UNIQUE(invoice_no),
        FOREIGN KEY(order_id) REFERENCES orders(id)
    )
    """)
    # Nowe dokumenty zapisują jawny typ podatkowy. Kolumna jest nullable,
    # dzięki czemu historycznych faktur nie backfillujemy ani nie zmieniamy.
    cur.execute("PRAGMA table_info(invoices)")
    invoice_cols = {r[1] for r in cur.fetchall()}
    if "invoice_type" not in invoice_cols:
        cur.execute("ALTER TABLE invoices ADD COLUMN invoice_type TEXT")
    if "currency" not in invoice_cols:
        cur.execute("ALTER TABLE invoices ADD COLUMN currency TEXT")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS invoice_meta(
        invoice_id INTEGER PRIMARY KEY,
        pdf_path TEXT,
        invoice_items_json TEXT,
        sent_to_client INTEGER NOT NULL DEFAULT 0,
        seen_by_client INTEGER NOT NULL DEFAULT 0,
        payment_reminder INTEGER NOT NULL DEFAULT 0,
        paid INTEGER NOT NULL DEFAULT 0,
        paid_at TEXT,
        seen_at TEXT,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(invoice_id) REFERENCES invoices(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS ksef_documents(
        invoice_id INTEGER PRIMARY KEY,
        status TEXT NOT NULL DEFAULT 'draft',
        ksef_number TEXT,
        xml_path TEXT,
        last_error TEXT,
        validated_at TEXT,
        sent_at TEXT,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(invoice_id) REFERENCES invoices(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS cash_flow_settings(
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL DEFAULT '',
        updated_at TEXT NOT NULL
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS cash_flow_expenses(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        expense_date TEXT NOT NULL,
        category TEXT NOT NULL,
        description TEXT,
        document_no TEXT NOT NULL,
        amount REAL NOT NULL CHECK(amount > 0),
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS invoice_allocations(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_id INTEGER NOT NULL,
        order_id INTEGER NOT NULL,
        order_item_id INTEGER NOT NULL,
        product_id INTEGER,
        sku TEXT,
        qty INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(invoice_id) REFERENCES invoices(id),
        FOREIGN KEY(order_id) REFERENCES orders(id),
        FOREIGN KEY(order_item_id) REFERENCES order_items(id)
    )
    """)

    # Trwałe źródło danych dla faktury tworzonej po liście pakowej.
    # Sesja przeglądarki może zniknąć po wdrożeniu lub restarcie Rendera.
    cur.execute("""
    CREATE TABLE IF NOT EXISTS packing_batches(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        root_order_id INTEGER NOT NULL,
        invoice_id INTEGER,
        created_at TEXT NOT NULL
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS packing_allocations(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id INTEGER NOT NULL,
        order_id INTEGER NOT NULL,
        order_item_id INTEGER NOT NULL,
        qty INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(batch_id) REFERENCES packing_batches(id)
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_packing_batches_root_open ON packing_batches(root_order_id, invoice_id, id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_packing_allocations_batch ON packing_allocations(batch_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_invoices_issue_date ON invoices(issue_date DESC, id DESC)")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS client_search_logs(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_email TEXT,
        customer_name TEXT,
        query TEXT NOT NULL,
        product_sku TEXT,
        product_model TEXT,
        product_name TEXT,
        results_count INTEGER NOT NULL DEFAULT 0,
        source TEXT,
        created_at TEXT NOT NULL
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS email_events(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_key TEXT UNIQUE,
        event_type TEXT NOT NULL,
        ref_id TEXT,
        recipient TEXT,
        ok INTEGER NOT NULL DEFAULT 0,
        result_json TEXT,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("PRAGMA table_info(client_search_logs)")
    search_cols = {r["name"] for r in cur.fetchall()}
    if "product_sku" not in search_cols:
        cur.execute("ALTER TABLE client_search_logs ADD COLUMN product_sku TEXT")
    if "product_model" not in search_cols:
        cur.execute("ALTER TABLE client_search_logs ADD COLUMN product_model TEXT")
    if "product_name" not in search_cols:
        cur.execute("ALTER TABLE client_search_logs ADD COLUMN product_name TEXT")

    # migracja: pierwsze wersje magazynu miały w products tylko SKU, EAN i nazwę.
    # Potwierdzenie e-mail korzysta również z modelu przy awaryjnym odczycie ceny.
    cur.execute("PRAGMA table_info(products)")
    product_cols = {r[1] for r in cur.fetchall()}
    if "model" not in product_cols:
        cur.execute("ALTER TABLE products ADD COLUMN model TEXT")
    if "archived" not in product_cols:
        cur.execute("ALTER TABLE products ADD COLUMN archived INTEGER NOT NULL DEFAULT 0")

    cur.execute("PRAGMA table_info(company_profile)")
    company_cols = {r[1] for r in cur.fetchall()}
    if "bank_swift" not in company_cols:
        cur.execute("ALTER TABLE company_profile ADD COLUMN bank_swift TEXT")

    cur.execute("PRAGMA table_info(china_packages)")
    china_package_cols = {r[1] for r in cur.fetchall()}
    if "cost_amount" not in china_package_cols:
        cur.execute("ALTER TABLE china_packages ADD COLUMN cost_amount REAL NOT NULL DEFAULT 0")
    if "cost_document_no" not in china_package_cols:
        cur.execute("ALTER TABLE china_packages ADD COLUMN cost_document_no TEXT")
    china_tracking_columns = {
        "supplier": "TEXT", "ordered_at": "TEXT", "shipped_at": "TEXT",
        "arrived_at": "TEXT", "warehouse_received": "INTEGER",
        "warehouse_received_at": "TEXT", "tracking_carrier": "TEXT",
        "tracking_carrier_code": "INTEGER", "tracking_status": "TEXT",
        "tracking_substatus": "TEXT", "tracking_last_event": "TEXT",
        "tracking_last_update": "TEXT", "tracking_synced_at": "TEXT",
        "tracking_error": "TEXT", "tracking_events_json": "TEXT",
        "tracking_eta": "TEXT", "tracking_registered_at": "TEXT",
        "manual_status_at": "TEXT",
        "shipping_method": "TEXT",
    }
    for column_name, column_type in china_tracking_columns.items():
        if column_name not in china_package_cols:
            cur.execute(f"ALTER TABLE china_packages ADD COLUMN {column_name} {column_type}")

    cur.execute("PRAGMA table_info(china_documents)")
    china_document_cols = {r[1] for r in cur.fetchall()}
    if "document_type" not in china_document_cols:
        cur.execute("ALTER TABLE china_documents ADD COLUMN document_type TEXT NOT NULL DEFAULT 'order'")
    cur.execute("""
      INSERT OR IGNORE INTO china_stock_receipts(package_id,received_at,quantities_json)
      SELECT id,COALESCE(warehouse_received_at,arrived_at,created_at),'[]'
      FROM china_packages
      WHERE status='arrived' OR warehouse_received=1 OR arrived_at IS NOT NULL
    """)

    cur.execute("PRAGMA table_info(invoice_meta)")
    invoice_meta_cols = {r[1] for r in cur.fetchall()}
    if "seen_by_client" not in invoice_meta_cols:
        cur.execute("ALTER TABLE invoice_meta ADD COLUMN seen_by_client INTEGER NOT NULL DEFAULT 0")
    if "seen_at" not in invoice_meta_cols:
        cur.execute("ALTER TABLE invoice_meta ADD COLUMN seen_at TEXT")
    if "payment_reminder" not in invoice_meta_cols:
        cur.execute("ALTER TABLE invoice_meta ADD COLUMN payment_reminder INTEGER NOT NULL DEFAULT 0")
    if "paid" not in invoice_meta_cols:
        cur.execute("ALTER TABLE invoice_meta ADD COLUMN paid INTEGER NOT NULL DEFAULT 0")
    if "paid_at" not in invoice_meta_cols:
        cur.execute("ALTER TABLE invoice_meta ADD COLUMN paid_at TEXT")

    # migracja: starsze bazy mogÄ… nie mieÄ‡ kolumny NIP u klientĂłw
    cur.execute("PRAGMA table_info(customers)")
    customer_cols = {r[1] for r in cur.fetchall()}
    if "nip" not in customer_cols:
        cur.execute("ALTER TABLE customers ADD COLUMN nip TEXT")
    if "language" not in customer_cols:
        cur.execute("ALTER TABLE customers ADD COLUMN language TEXT NOT NULL DEFAULT 'pl'")
    if "price_list" not in customer_cols:
        cur.execute("ALTER TABLE customers ADD COLUMN price_list TEXT NOT NULL DEFAULT 'pln'")
    cur.execute("""
      UPDATE customers
      SET language='pl'
      WHERE language IS NULL OR LOWER(TRIM(language)) NOT IN ('pl','de','en','es','it')
    """)
    cur.execute("""
      UPDATE customers
      SET price_list='pln'
      WHERE price_list IS NULL OR LOWER(TRIM(price_list)) NOT IN ('pln','eu_eur')
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS pricing_eur(
        sku TEXT PRIMARY KEY,
        ean TEXT,
        price_eur REAL NOT NULL DEFAULT 0,
        uvp_eur REAL NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )
    """)

    # migracja: QR zamĂłwieĹ„
    cur.execute("PRAGMA table_info(orders)")
    order_cols = {r[1] for r in cur.fetchall()}
    if "qr_data_url" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN qr_data_url TEXT")
    if "warehouse_issued" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN warehouse_issued INTEGER NOT NULL DEFAULT 0")
    if "idempotency_key" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN idempotency_key TEXT")
    if "currency" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN currency TEXT NOT NULL DEFAULT 'PLN'")
    if "price_list" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN price_list TEXT NOT NULL DEFAULT 'pln'")
    if "tracking_no" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN tracking_no TEXT")
    if "carrier" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN carrier TEXT")
    if "packed_at" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN packed_at TEXT")
    if "shipped_at" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN shipped_at TEXT")
    if "inpost_shipment_id" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN inpost_shipment_id TEXT")
    if "inpost_label_format" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN inpost_label_format TEXT")
    if "inpost_dispatch_order_id" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN inpost_dispatch_order_id TEXT")

    cur.execute("PRAGMA table_info(order_items)")
    order_item_cols = {r[1] for r in cur.fetchall()}
    if "unit_net_price" not in order_item_cols:
        cur.execute("ALTER TABLE order_items ADD COLUMN unit_net_price REAL")
    if "unit_gross_price" not in order_item_cols:
        cur.execute("ALTER TABLE order_items ADD COLUMN unit_gross_price REAL")
    if "unit_retail_price" not in order_item_cols:
        cur.execute("ALTER TABLE order_items ADD COLUMN unit_retail_price REAL")
    if "currency" not in order_item_cols:
        cur.execute("ALTER TABLE order_items ADD COLUMN currency TEXT")

    # UĹ‚atwia agregowanie "w dostawie" po statusach paczek
    cur.execute("CREATE INDEX IF NOT EXISTS idx_order_items_order_id ON order_items(order_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_order_items_product_id ON order_items(product_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_orders_status_issued_created ON orders(status, warehouse_issued, created_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_china_packages_status ON china_packages(status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_china_items_package_id ON china_items(package_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_china_items_product_id ON china_items(product_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_client_search_logs_created ON client_search_logs(created_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_client_search_logs_email_query ON client_search_logs(customer_email, query)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_client_search_logs_model ON client_search_logs(product_model)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_pricing_model_norm ON pricing(TRIM(LOWER(model)))")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_pricing_eur_sku_norm ON pricing_eur(TRIM(LOWER(sku)))")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_email_events_key ON email_events(event_key)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_email_events_type_created ON email_events(event_type, created_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_cash_flow_expenses_date ON cash_flow_expenses(expense_date)")
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_orders_idempotency_key ON orders(idempotency_key) WHERE idempotency_key IS NOT NULL")

    c.commit()
    c.close()

init_db()


# =========================
# UTILS
# =========================

APP_TZ = ZoneInfo("Europe/Warsaw") if ZoneInfo else None

def app_now():
    return datetime.now(APP_TZ) if APP_TZ else datetime.now()

def now_iso():
    return app_now().strftime("%Y-%m-%d %H:%M:%S")


def overdue_invoice_rows(db_conn, *, current_time=None):
    """Widok zaleglosci oparty na wspolnej logice modulu Cash flow."""
    result = cash_flow_overdue_invoices(
        db_conn, current_time=current_time or app_now()
    )
    for invoice in result:
        invoice["order_display"] = order_display_no(
            invoice.get("source_order_id"), invoice.get("source_order_created_at"),
            invoice.get("source_order_no"), invoice.get("source_order_note")
        ) if invoice.get("source_order_id") else "-"
    return result

SHORT_ORDER_NO_RE = re.compile(r"^ZAM-(\d{6})(\d+)$", re.I)


def order_date_code(created_at: str | None = "") -> str:
    created = norm(created_at)
    if len(created) >= 10 and created[4:5] == "-" and created[7:8] == "-":
        return created[2:4] + created[5:7] + created[8:10]
    return app_now().strftime("%y%m%d")


def is_short_order_no(value: str | None) -> bool:
    return bool(SHORT_ORDER_NO_RE.match(norm(value)))


def make_order_no(order_id: int | None = None, created_at: str | None = "") -> str:
    # Format: ZAM-2607141 = ZAM- + YYMMDD + kolejny numer w danym dniu.
    date_code = order_date_code(created_at)
    day = norm(created_at)[:10] if norm(created_at) else app_now().strftime("%Y-%m-%d")
    seq = 1
    try:
        c = conn()
        cur = c.cursor()
        if order_id:
            cur.execute("SELECT order_no FROM orders WHERE substr(created_at,1,10)=? AND id<>?", (day, int(order_id)))
        else:
            cur.execute("SELECT order_no FROM orders WHERE substr(created_at,1,10)=?", (day,))
        for r in cur.fetchall():
            raw = norm(r["order_no"])
            m = SHORT_ORDER_NO_RE.match(raw)
            if m and m.group(1) == date_code:
                seq = max(seq, int(m.group(2)) + 1)
            elif raw and raw.upper() != "TEMP":
                seq += 1
        c.close()
    except Exception:
        oid = int(order_id or 1)
        seq = max(1, oid)
    return f"ZAM-{date_code}{seq}"


def canonical_order_no(order_id: int | None, created_at: str | None = "", raw_order_no: str | None = "") -> str:
    raw = norm(raw_order_no)
    if raw and raw.upper() != "TEMP":
        if raw.startswith("ORD-"):
            return "ZAM-" + raw[4:]
        return raw

    return make_order_no(order_id, created_at)


def order_display_no(order_id: int | None, created_at: str | None = "", raw_order_no: str | None = "", note: str | None = "") -> str:
    base = canonical_order_no(order_id, created_at, raw_order_no)
    note_text = norm(note)
    return f"{base} {note_text}" if note_text else base


def normalize_temp_order_numbers():
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT id, order_no, created_at FROM orders ORDER BY created_at, id")
    rows = cur.fetchall()
    changed = []
    used_seq_by_date = {}
    for r in rows:
        raw = norm(r["order_no"])
        m = SHORT_ORDER_NO_RE.match(raw)
        if m:
            used_seq_by_date[m.group(1)] = max(used_seq_by_date.get(m.group(1), 0), int(m.group(2)))

    for r in rows:
        raw = norm(r["order_no"])
        if is_short_order_no(raw):
            continue
        date_code = order_date_code(r["created_at"])
        used_seq_by_date[date_code] = used_seq_by_date.get(date_code, 0) + 1
        new_no = f"ZAM-{date_code}{used_seq_by_date[date_code]}"
        if new_no != (r["order_no"] or ""):
            cur.execute("UPDATE orders SET order_no=? WHERE id=?", (new_no, r["id"]))
            changed.append((int(r["id"]), new_no))
    c.commit()
    c.close()

    if supabase_enabled():
        for oid, ono in changed:
            try:
                supabase_update_rows("orders", {"order_no": ono}, {"id": oid})
            except Exception:
                pass
    return len(changed)

def _email_key(value: str) -> str:
    return norm(value).strip().lower()

def _order_name_is_fallback(order_name: str, email_value: str) -> bool:
    email_key = _email_key(email_value)
    if not email_key:
        return False
    local_part = email_key.split("@")[0]
    current = norm(order_name).strip().lower()
    return current in {"", email_key, local_part}

def link_orders_to_customers_by_email(sync_remote: bool = True):
    c = conn()
    cur = c.cursor()

    cur.execute("""
      SELECT id, name, address, phone, email
      FROM customers
      WHERE TRIM(COALESCE(email, '')) <> ''
      ORDER BY id
    """)
    customer_rows = [dict(r) for r in cur.fetchall()]
    customers_by_email = {_email_key(r["email"]): r for r in customer_rows if _email_key(r.get("email"))}

    cur.execute("""
      SELECT id, customer_id, customer_name, customer_address, customer_phone, customer_email
      FROM orders
      WHERE TRIM(COALESCE(customer_email, '')) <> ''
      ORDER BY id
    """)
    order_rows = [dict(r) for r in cur.fetchall()]

    changed = []
    for order_row in order_rows:
        email_key = _email_key(order_row.get("customer_email"))
        customer = customers_by_email.get(email_key)
        if not customer:
            continue

        updates = {}
        if int(order_row.get("customer_id") or 0) != int(customer["id"]):
            updates["customer_id"] = int(customer["id"])

        if _order_name_is_fallback(order_row.get("customer_name"), order_row.get("customer_email")) and norm(customer.get("name")):
            updates["customer_name"] = norm(customer.get("name"))

        if not norm(order_row.get("customer_address")) and norm(customer.get("address")):
            updates["customer_address"] = norm(customer.get("address"))

        if not norm(order_row.get("customer_phone")) and norm(customer.get("phone")):
            updates["customer_phone"] = norm(customer.get("phone"))

        if not norm(order_row.get("customer_email")) and norm(customer.get("email")):
            updates["customer_email"] = norm(customer.get("email"))

        if not updates:
            continue

        sets = ", ".join([f"{k}=?" for k in updates.keys()])
        values = list(updates.values()) + [int(order_row["id"])]
        cur.execute(f"UPDATE orders SET {sets} WHERE id=?", values)
        changed.append((int(order_row["id"]), updates))

    c.commit()
    c.close()

    if sync_remote and supabase_enabled():
        for order_id, updates in changed:
            try:
                supabase_update_rows("orders", updates, {"id": order_id})
            except Exception:
                pass

    return len(changed)


def make_qr_data_url(value: str) -> str:
    raw_value = norm(value)
    if not raw_value:
        return ""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=1
    )
    qr.add_data(raw_value)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


def next_invoice_no(issue_date: str) -> str:
    dt = datetime.strptime(issue_date, "%Y-%m-%d")
    mm = dt.strftime("%m")
    yyyy = dt.strftime("%Y")
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT COUNT(*) AS n FROM invoices WHERE substr(issue_date,1,7)=?", (f"{yyyy}-{mm}",))
    n = int(cur.fetchone()["n"] or 0) + 1
    c.close()
    return f"FVAT {n}/{mm}/{yyyy}"


def invoice_no_exists(invoice_no: str, exclude_invoice_id: int = 0) -> int:
    invoice_no = norm(invoice_no)
    if not invoice_no:
        return 0
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT id
      FROM invoices
      WHERE lower(trim(invoice_no)) = lower(trim(?))
        AND id <> ?
      LIMIT 1
    """, (invoice_no, int(exclude_invoice_id or 0)))
    row = cur.fetchone()
    c.close()
    return int(row["id"]) if row else 0


def split_address(addr: str):
    raw = (addr or "").strip()
    if not raw:
        return "", "", ""

    # wspieraj adres w wielu liniach oraz jednoliniowy (np. "ul. X 1, 00-001 Warszawa")
    parts = [x.strip() for x in raw.splitlines() if x.strip()]
    if len(parts) == 1 and "," in raw:
        comma_parts = [x.strip() for x in raw.split(",") if x.strip()]
        if len(comma_parts) >= 2:
            parts = [comma_parts[0], " ".join(comma_parts[1:])]

    street = parts[0] if parts else ""
    post_code = ""
    city = ""
    if len(parts) > 1:
        line2 = parts[1].strip()
        m = re.match(r"^(\d{2}-\d{3})\s*(.*)$", line2)
        if m:
            post_code = m.group(1).strip()
            city = m.group(2).strip()
        else:
            pc = line2.split(" ", 1)
            post_code = pc[0].strip() if pc else ""
            city = pc[1].strip() if len(pc) > 1 else ""
    return street, post_code, city


def payment_type_pl(x: str) -> str:
    v = norm(x).lower()
    mapping = {
        "cash": "gotĂłwka",
        "gotowka": "gotĂłwka",
        "transfer": "przelew",
        "card": "karta",
        "karta": "karta",
    }
    return mapping.get(v, v or "-")


VAT_23 = Decimal("0.23")
MONEY_Q = Decimal("0.01")
CURRENT_ORDER_STATUSES = {
    "new", "pending", "unconfirmed", "confirmed", "packed", "packed_partial",
    "in_delivery", "shipped", "partially_shipped",
}


def money_dec(value) -> Decimal:
    try:
        return Decimal(str(value or "0")).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal("0.00")


def money_float(value) -> float:
    return float(money_dec(value))


def vat23_from_net(net_value) -> Decimal:
    return (money_dec(net_value) * VAT_23).quantize(MONEY_Q, rounding=ROUND_HALF_UP)


def gross_from_net_23(net_value) -> Decimal:
    net = money_dec(net_value)
    return (net + vat23_from_net(net)).quantize(MONEY_Q, rounding=ROUND_HALF_UP)


def find_logo_path() -> str:
    search_dirs = [
        APP_DIR,
        os.path.join(APP_DIR, "static"),
        DATA_DIR,
    ]
    for folder in search_dirs:
        for fn in ("logo.png", "logo.jpg", "logo.jpeg", "logo.webp"):
            pth = os.path.join(folder, fn)
            if os.path.exists(pth):
                return pth
    return ""


def to_int(x, default=0):
    try:
        return int(str(x).strip())
    except:
        return default

def to_float(x, default=0.0):
    try:
        return float(str(x).strip().replace(" ", "").replace(",", "."))
    except:
        return default

def norm(s):
    if s is None:
        return ""
    return str(s).strip()

def order_status_label(status: str) -> str:
    v = norm(status).lower()
    mapping = {
        "new": "Niepotwierdzone",
        "pending": "Niepotwierdzone",
        "unconfirmed": "Niepotwierdzone",
        "confirmed": "Potwierdzone",
        "packed": "W trakcie pakowania / czeka na kuriera",
        "packed_partial": "Pakowanie częściowej wysyłki",
        "in_delivery": "W dostawie",
        "shipped": "Wysłane",
        "partially_shipped": "Wysłane częściowo",
        "issued": "W realizacji",
        "completed": "Zrealizowane",
    }
    return mapping.get(v, status or "-")

def order_status_css(status: str) -> str:
    v = norm(status).lower()
    mapping = {
        "new": "st-unconfirmed",
        "pending": "st-unconfirmed",
        "unconfirmed": "st-unconfirmed",
        "confirmed": "st-confirmed",
        "packed": "st-delivery",
        "packed_partial": "st-delivery",
        "in_delivery": "st-delivery",
        "shipped": "st-confirmed",
        "partially_shipped": "st-delivery",
        "issued": "st-delivery",
        "completed": "st-issued",
    }
    return mapping.get(v, "")

def carrier_tracking_url(carrier: str, tracking_no: str) -> str:
    carrier_key = norm(carrier).lower()
    number = re.sub(r"\s+", "", norm(tracking_no))
    encoded = urllib.parse.quote(number)
    bases = {
        "inpost": "https://inpost.pl/sledzenie-przesylek?number=",
        "dpd": "https://tracktrace.dpd.com.pl/parcelDetails?p1=",
        "fedex": "https://www.fedex.com/fedextrack/?trknbr=",
        "dhl": "https://www.dhl.com/pl-pl/home/tracking.html?tracking-id=",
        "ups": "https://www.ups.com/track?tracknum=",
    }
    return bases.get(carrier_key, "https://t.17track.net/en#nums=") + encoded

def guess_col(headers, candidates):
    h = [x.strip().lower() for x in headers]
    for cand in candidates:
        cand = cand.lower()
        if cand in h:
            return h.index(cand)
    # luĹşne dopasowanie: np. "model" w "Model uchwytu"
    for i, col in enumerate(h):
        for cand in candidates:
            if cand.lower() in col:
                return i
    return None

def ensure_stock_row(product_id):
    c = conn()
    cur = c.cursor()
    cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (product_id,))
    c.commit()
    c.close()


# =========================
# SUPABASE (cloud sync)
# =========================

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or "https://qfzawzkynmqkbjlbtkjd.supabase.co").strip().rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY", "").strip()
CLIENT_ALLOWED_ORIGINS = {
    value.strip().rstrip("/")
    for value in os.environ.get("CLIENT_ALLOWED_ORIGINS", "").split(",")
    if value.strip()
}
CLIENT_PANEL_URL = (
    os.environ.get("CLIENT_PANEL_URL")
    or "https://panel-klienta-niedzwieccy.netlify.app"
).strip().rstrip("/")
ADMIN_ACTION_TOKEN = os.environ.get("ADMIN_ACTION_TOKEN", "").strip()
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin").strip()
ADMIN_PASSWORD_HASH = os.environ.get("ADMIN_PASSWORD_HASH", "").strip()
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "").strip()
if not SUPABASE_SERVICE_ROLE_KEY:
    app.logger.error("Brak SUPABASE_SERVICE_ROLE_KEY; funkcje synchronizacji i zamówień klienta będą niedostępne.")
if not CLIENT_ALLOWED_ORIGINS:
    app.logger.warning("CLIENT_ALLOWED_ORIGINS jest puste; przeglądarkowe żądania tworzenia zamówień będą odrzucane.")
SUPABASE_STORAGE_BUCKET = (os.environ.get("SUPABASE_STORAGE_BUCKET") or "invoice-pdfs").strip()
SUPABASE_AUTO_SYNC_ON_WRITE = (os.environ.get("SUPABASE_AUTO_SYNC_ON_WRITE") or "1").strip().lower() in ("1", "true", "yes", "on")
SUPABASE_MIN_SYNC_INTERVAL_SEC = float((os.environ.get("SUPABASE_MIN_SYNC_INTERVAL_SEC") or "2").strip())
SUPABASE_MIN_PULL_INTERVAL_SEC = float((os.environ.get("SUPABASE_MIN_PULL_INTERVAL_SEC") or "2").strip())
SUPABASE_BACKGROUND_PULL_INTERVAL_SEC = float((os.environ.get("SUPABASE_BACKGROUND_PULL_INTERVAL_SEC") or "30").strip())
SEVENTEENTRACK_API_KEY = os.environ.get("SEVENTEENTRACK_API_KEY", "").strip()
SEVENTEENTRACK_ENABLED = os.environ.get("SEVENTEENTRACK_ENABLED", "0").strip()
SEVENTEENTRACK_TIMEOUT_SEC = int(os.environ.get("SEVENTEENTRACK_TIMEOUT_SEC", "15") or 15)

SUPABASE_SYNC_TABLES = [
    ("products", "id"),
    ("stock", "product_id"),
    ("customers", "id"),
    ("orders", "id"),
    ("order_items", "id"),
    ("china_packages", "id"),
    ("china_stock_receipts", "package_id"),
    ("china_items", "id"),
    ("pricing", "model"),
    ("pricing_eur", "sku"),
    ("company_profile", "id"),
    ("invoices", "id"),
    ("invoice_meta", "invoice_id"),
    ("invoice_allocations", "id"),
    ("ksef_documents", "invoice_id"),
    ("cash_flow_settings", "key"),
    ("cash_flow_expenses", "id"),
]

# Kolumna jest używana lokalnie przez moduł zamawiania odbioru InPost, ale
# starszy schemat Supabase jej nie posiada. Nie może przez to blokować
# synchronizacji całego rekordu zamówienia ani działania pulpitu.
SUPABASE_LOCAL_ONLY_COLUMNS = {
    "orders": {"inpost_dispatch_order_id"},
}


def supabase_compatible_rows(table: str, rows: list) -> list:
    omitted = SUPABASE_LOCAL_ONLY_COLUMNS.get(table, set())
    if not omitted:
        return rows
    return [{key: value for key, value in row.items() if key not in omitted} for row in rows]

# KolejnoĹ›Ä‡ PULL jest waĹĽna: najpierw rodzice, potem dzieci.
SUPABASE_PULL_TABLES = [
    ("company_profile", "id"),
    ("pricing", "model"),
    ("pricing_eur", "sku"),
    ("customers", "id"),
    ("products", "id"),
    ("product_images", "id"),
    ("product_image_assignments", "product_id"),
    ("orders", "id"),
    ("china_packages", "id"),
    ("china_stock_receipts", "package_id"),
    ("stock", "product_id"),
    ("order_items", "id"),
    ("china_items", "id"),
    ("invoices", "id"),
    ("invoice_meta", "invoice_id"),
    ("invoice_allocations", "id"),
    ("ksef_documents", "invoice_id"),
    ("cash_flow_settings", "key"),
    ("cash_flow_expenses", "id"),
]

_supabase_sync_lock = threading.Lock()
_supabase_full_io_lock = threading.Lock()
_supabase_sync_state = {
    "running": False,
    "pull_running": False,
    "last_started_ts": 0.0,
    "last_pull_finished_ts": 0.0,
    "last_result": None,
    "initial_pull_attempted": False,
}


def _local_supabase_bootstrap_complete() -> bool:
    """True only after a complete successful cloud bootstrap of this SQLite DB."""
    c = conn()
    try:
        c.execute("""
          CREATE TABLE IF NOT EXISTS local_sync_state(
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT NOT NULL
          )
        """)
        row = c.execute(
            "SELECT value FROM local_sync_state WHERE key='supabase_bootstrap_complete'"
        ).fetchone()
        c.commit()
        return bool(row and row["value"] == "1")
    finally:
        c.close()


def _mark_local_supabase_bootstrap_complete():
    c = conn()
    try:
        c.execute("""
          CREATE TABLE IF NOT EXISTS local_sync_state(
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT NOT NULL
          )
        """)
        c.execute("""
          INSERT INTO local_sync_state(key,value,updated_at)
          VALUES('supabase_bootstrap_complete','1',?)
          ON CONFLICT(key) DO UPDATE SET value='1',updated_at=excluded.updated_at
        """, (now_iso(),))
        c.commit()
    finally:
        c.close()

def supabase_enabled() -> bool:
    return bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY)

def _chunks(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]

def supabase_upsert_rows(table: str, rows: list, on_conflict: str):
    if not rows:
        return
    if not supabase_enabled():
        raise RuntimeError("Brak konfiguracji SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY")

    rows = supabase_compatible_rows(table, rows)
    qs = urllib.parse.urlencode({"on_conflict": on_conflict})
    url = f"{SUPABASE_URL}/rest/v1/{table}?{qs}"
    payload = json.dumps(rows, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=payload, method="POST")
    req.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Prefer", "resolution=merge-duplicates,return=minimal")
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            if resp.status >= 300:
                raise RuntimeError(f"Supabase HTTP {resp.status}")
    except urllib.error.HTTPError as exc:
        raw_error = exc.read().decode("utf-8", errors="replace")[:1200]
        try:
            error_payload = json.loads(raw_error)
            detail = norm(
                error_payload.get("message")
                or error_payload.get("details")
                or error_payload.get("hint")
                or raw_error
            )
        except Exception:
            detail = norm(raw_error)
        raise RuntimeError(f"Supabase HTTP {exc.code}: {detail or 'brak szczegółów'}") from exc

def sqlite_table_rows(table: str):
    c = conn()
    cur = c.cursor()
    cur.execute(f"SELECT * FROM {table}")
    rows = [dict(r) for r in cur.fetchall()]
    c.close()

    return rows

def sync_all_to_supabase():
    if not supabase_enabled():
        return {"ok": False, "error": "Brak konfiguracji SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY"}

    out = {"ok": True, "tables": {}, "synced_at": now_iso()}
    for table, conflict_col in SUPABASE_SYNC_TABLES:
        try:
            rows = sqlite_table_rows(table)
            for pack in _chunks(rows, 500):
                supabase_upsert_rows(table, pack, conflict_col)
            out["tables"][table] = {"rows": len(rows), "status": "ok"}
        except Exception as e:
            out["ok"] = False
            out["tables"][table] = {"status": "error", "error": str(e)}
    return out

def trigger_background_supabase_sync(reason: str = "write"):
    if not SUPABASE_AUTO_SYNC_ON_WRITE:
        return False, "disabled"
    if not supabase_enabled():
        return False, "not_configured"

    now_ts = time.time()
    with _supabase_sync_lock:
        if _supabase_sync_state["running"]:
            return False, "already_running"
        if (now_ts - float(_supabase_sync_state["last_started_ts"])) < SUPABASE_MIN_SYNC_INTERVAL_SEC:
            return False, "throttled"
        _supabase_sync_state["running"] = True
        _supabase_sync_state["last_started_ts"] = now_ts

    def _job():
        try:
            with _supabase_full_io_lock:
                result = sync_all_to_supabase()
            result["reason"] = reason
        except Exception as e:
            result = {"ok": False, "error": str(e), "reason": reason, "synced_at": now_iso()}
        finally:
            with _supabase_sync_lock:
                _supabase_sync_state["running"] = False
                _supabase_sync_state["last_result"] = result

    th = threading.Thread(target=_job, daemon=True)
    th.start()
    return True, "started"



def supabase_request(path: str, method: str = "GET", params: dict | None = None, payload=None, prefer: str | None = None, timeout: int = 60):
    if not supabase_enabled():
        raise RuntimeError("Brak konfiguracji SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY")

    url = f"{SUPABASE_URL}{path}"
    if params:
        qs = urllib.parse.urlencode(params, doseq=True)
        url = f"{url}?{qs}"

    data = None
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    req.add_header("Content-Type", "application/json")
    if prefer:
        req.add_header("Prefer", prefer)

    started = time.perf_counter()
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read()
            if not raw:
                return None
            ctype = (resp.headers.get("Content-Type") or "").lower()
            if "application/json" in ctype or raw[:1] in (b"[", b"{"):
                return json.loads(raw.decode("utf-8"))
            return raw.decode("utf-8", errors="replace")
    finally:
        elapsed = time.perf_counter() - started
        _perf_add("supabase_http", elapsed)
        if PERF_LOG_ENABLED and elapsed >= 0.25:
            app.logger.info("PERF SUPABASE %s %s %.2f ms", method, path, elapsed * 1000)


def supabase_storage_ref(object_path: str, bucket: str | None = None) -> str:
    bucket = bucket or SUPABASE_STORAGE_BUCKET
    return f"supabase://{bucket}/{object_path.lstrip('/')}"


def parse_supabase_storage_ref(value: str) -> tuple[str, str] | None:
    raw = norm(value)
    if not raw.startswith("supabase://"):
        return None
    rest = raw[len("supabase://"):]
    if "/" not in rest:
        return None
    bucket, object_path = rest.split("/", 1)
    return bucket, object_path


def supabase_storage_object_url(bucket: str, object_path: str) -> str:
    quoted_path = urllib.parse.quote(object_path.lstrip("/"), safe="/")
    return f"{SUPABASE_URL}/storage/v1/object/{urllib.parse.quote(bucket, safe='')}/{quoted_path}"


def ensure_supabase_storage_bucket(bucket: str | None = None):
    bucket = bucket or SUPABASE_STORAGE_BUCKET
    if not supabase_enabled() or not bucket:
        return
    payload = json.dumps({"id": bucket, "name": bucket, "public": False}).encode("utf-8")
    req = urllib.request.Request(f"{SUPABASE_URL}/storage/v1/bucket", data=payload, method="POST")
    req.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    req.add_header("Content-Type", "application/json")
    try:
        urllib.request.urlopen(req, timeout=30).read()
    except urllib.error.HTTPError as e:
        if e.code not in (400, 409):
            raise


def supabase_storage_upload_file(local_path: str, object_path: str, bucket: str | None = None, content_type: str = "application/pdf") -> str:
    if not supabase_enabled():
        raise RuntimeError("Brak konfiguracji Supabase")
    bucket = bucket or SUPABASE_STORAGE_BUCKET
    ensure_supabase_storage_bucket(bucket)
    with open(local_path, "rb") as f:
        data = f.read()
    req = urllib.request.Request(supabase_storage_object_url(bucket, object_path), data=data, method="POST")
    req.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    req.add_header("Content-Type", content_type)
    req.add_header("x-upsert", "true")
    with urllib.request.urlopen(req, timeout=90) as resp:
        resp.read()
    return supabase_storage_ref(object_path, bucket)


def supabase_storage_download_bytes(storage_ref: str) -> tuple[bytes, str]:
    parsed = parse_supabase_storage_ref(storage_ref)
    if not parsed:
        raise RuntimeError("Nieprawidłowa ścieżka Supabase Storage")
    bucket, object_path = parsed
    req = urllib.request.Request(supabase_storage_object_url(bucket, object_path), method="GET")
    req.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    with urllib.request.urlopen(req, timeout=90) as resp:
        return resp.read(), os.path.basename(object_path)


def supabase_insert_row(table: str, row: dict):
    res = supabase_request(
        f"/rest/v1/{table}",
        method="POST",
        payload=[row],
        prefer="return=representation",
    )
    if isinstance(res, list):
        return res[0] if res else None
    return res


def supabase_update_rows(table: str, values: dict, filters: dict):
    params = {k: f"eq.{v}" for k, v in filters.items()}
    return supabase_request(
        f"/rest/v1/{table}",
        method="PATCH",
        params=params,
        payload=values,
        prefer="return=minimal",
    )


def supabase_delete_rows(table: str, filters: dict):
    params = {k: f"eq.{v}" for k, v in filters.items()}
    return supabase_request(
        f"/rest/v1/{table}",
        method="DELETE",
        params=params,
        prefer="return=minimal",
    )


def supabase_select_rows(table: str, order_by: str = "id", page_size: int = 1000, extra_params: dict | None = None):
    rows = []
    offset = 0
    while True:
        params = {"select": "*", "limit": page_size, "offset": offset}
        if order_by:
            params["order"] = f"{order_by}.asc"
        if extra_params:
            params.update(extra_params)
        chunk = supabase_request(f"/rest/v1/{table}", method="GET", params=params) or []
        if not isinstance(chunk, list):
            raise RuntimeError(f"NieprawidĹ‚owa odpowiedĹş Supabase dla tabeli {table}")
        rows.extend(chunk)
        if len(chunk) < page_size:
            break
        offset += page_size
    return rows


def local_client_search_rows(limit: int = 5000):
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT customer_email, customer_name, query, product_sku, product_model, product_name, results_count, source, created_at
      FROM client_search_logs
      ORDER BY created_at DESC, id DESC
      LIMIT ?
    """, (limit,))
    rows = [dict(r) for r in cur.fetchall()]
    c.close()
    return rows


def supabase_client_search_rows(limit: int = 5000):
    if not supabase_enabled():
        return []
    rows = supabase_request(
        "/rest/v1/client_search_logs",
        method="GET",
        params={
            "select": "customer_email,customer_name,query,product_sku,product_model,product_name,results_count,source,created_at",
            "order": "created_at.desc",
            "limit": str(limit),
        },
        timeout=30,
    ) or []
    return rows if isinstance(rows, list) else []


_client_search_cache_lock = threading.Lock()
_client_search_cloud_cache = {"rows": [], "running": False, "loaded_at": 0.0}


def _trigger_client_search_cache_refresh(limit: int = 5000):
    if not supabase_enabled():
        return False
    with _client_search_cache_lock:
        if _client_search_cloud_cache["running"] or time.time() - _client_search_cloud_cache["loaded_at"] < 60:
            return False
        _client_search_cloud_cache["running"] = True

    def _job():
        try:
            rows = supabase_client_search_rows(limit=limit)
            with _client_search_cache_lock:
                _client_search_cloud_cache["rows"] = rows
                _client_search_cloud_cache["loaded_at"] = time.time()
        except Exception:
            pass
        finally:
            with _client_search_cache_lock:
                _client_search_cloud_cache["running"] = False

    threading.Thread(target=_job, daemon=True).start()
    return True


def load_client_search_rows(limit: int = 5000):
    local_rows = local_client_search_rows(limit=limit)
    with _client_search_cache_lock:
        cloud_rows = list(_client_search_cloud_cache["rows"])
        cloud_ok = bool(_client_search_cloud_cache["loaded_at"])

    # Po restarcie Render lokalna baza i pamięć procesu mogą być jeszcze puste.
    # Pierwsza wersja uruchamiała wtedy wyłącznie odczyt w tle, więc użytkownik
    # widział pusty raport aż do ręcznego odświeżenia strony. Tylko w tym zimnym
    # przypadku pobieramy jedną, konkretną tabelę synchronicznie. Kolejne wejścia
    # korzystają z lokalnych danych/cache, a odświeżenie chmury odbywa się w tle.
    if not local_rows and not cloud_ok and supabase_enabled():
        try:
            cloud_rows = supabase_client_search_rows(limit=limit)
            with _client_search_cache_lock:
                _client_search_cloud_cache["rows"] = list(cloud_rows)
                _client_search_cloud_cache["loaded_at"] = time.time()
            cloud_ok = True
        except Exception as exc:
            app.logger.warning(
                "Nie udalo sie zaladowac historii wyszukiwan przy pierwszym wejsciu: %s",
                type(exc).__name__,
            )

    # Po błędzie pierwszego odczytu pozostaw również lekki retry w tle.
    _trigger_client_search_cache_refresh(limit=limit)

    merged = []
    seen = set()
    for row in list(cloud_rows) + list(local_rows):
        cleaned = {
            "customer_email": norm((row or {}).get("customer_email")).lower(),
            "customer_name": norm((row or {}).get("customer_name")),
            "query": norm((row or {}).get("query")),
            "product_sku": norm((row or {}).get("product_sku")),
            "product_model": norm((row or {}).get("product_model")),
            "product_name": norm((row or {}).get("product_name")),
            "results_count": to_int((row or {}).get("results_count"), 0),
            "source": norm((row or {}).get("source")) or "stock",
            "created_at": norm((row or {}).get("created_at")),
        }
        if not cleaned["query"]:
            continue
        key = (
            cleaned["customer_email"],
            cleaned["customer_name"],
            cleaned["query"].lower(),
            cleaned["product_sku"].lower(),
            cleaned["product_model"].lower(),
            cleaned["product_name"].lower(),
            cleaned["results_count"],
            cleaned["source"],
            cleaned["created_at"],
        )
        if key in seen:
            continue
        seen.add(key)
        merged.append(cleaned)

    merged.sort(key=lambda r: r.get("created_at") or "", reverse=True)
    source_label = "Supabase + kopia lokalna" if cloud_ok else "Kopia lokalna"
    return merged[:limit], source_label


def save_client_search_log_local(row: dict):
    c = conn()
    cur = c.cursor()
    cur.execute("""
      INSERT INTO client_search_logs(customer_email, customer_name, query, product_sku, product_model, product_name, results_count, source, created_at)
      VALUES(?,?,?,?,?,?,?,?,?)
    """, (
        row.get("customer_email", ""),
        row.get("customer_name", ""),
        row.get("query", ""),
        row.get("product_sku", ""),
        row.get("product_model", ""),
        row.get("product_name", ""),
        to_int(row.get("results_count"), 0),
        row.get("source", "stock"),
        row.get("created_at") or now_iso(),
    ))
    c.commit()
    c.close()


def save_client_search_log_supabase(row: dict) -> bool:
    if not supabase_enabled():
        return False
    payload = {
        "customer_email": row.get("customer_email", ""),
        "customer_name": row.get("customer_name", ""),
        "query": row.get("query", ""),
        "product_sku": row.get("product_sku", ""),
        "product_model": row.get("product_model", ""),
        "product_name": row.get("product_name", ""),
        "results_count": to_int(row.get("results_count"), 0),
        "source": row.get("source", "stock"),
        "created_at": row.get("created_at") or now_iso(),
    }
    supabase_insert_row("client_search_logs", payload)
    return True


def sqlite_table_columns(table: str):
    c = conn()
    cur = c.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = [r[1] for r in cur.fetchall()]
    c.close()
    return cols


def sqlite_upsert_rows(table: str, rows: list, conflict_col: str):
    if not rows:
        return 0

    table_cols = sqlite_table_columns(table)
    usable_cols = [c for c in table_cols if any(c in row for row in rows)]
    if not usable_cols:
        return 0

    placeholders = ",".join(["?"] * len(usable_cols))
    update_cols = [c for c in usable_cols if c != conflict_col]
    if update_cols:
        # Dostarczona/przyjęta paczka jest stanem monotonicznym. Starszy rekord
        # z Supabase nie może cofnąć jej do shipped ani wyzerować przyjęcia.
        if table == "china_packages":
            assignments = []
            for col in update_cols:
                if col == "status":
                    assignments.append("status=CASE WHEN china_packages.status='arrived' THEN china_packages.status ELSE excluded.status END")
                elif col == "warehouse_received":
                    assignments.append("warehouse_received=CASE WHEN china_packages.warehouse_received=1 THEN 1 ELSE excluded.warehouse_received END")
                elif col in {"warehouse_received_at", "arrived_at"}:
                    assignments.append(f"{col}=COALESCE(china_packages.{col},excluded.{col})")
                else:
                    assignments.append(f"{col}=excluded.{col}")
            update_sql = ", ".join(assignments)
        else:
            update_sql = ", ".join([f"{c}=excluded.{c}" for c in update_cols])
        sql = f"INSERT INTO {table}({','.join(usable_cols)}) VALUES({placeholders}) ON CONFLICT({conflict_col}) DO UPDATE SET {update_sql}"
    else:
        sql = f"INSERT INTO {table}({','.join(usable_cols)}) VALUES({placeholders}) ON CONFLICT({conflict_col}) DO NOTHING"

    c = conn()
    cur = c.cursor()
    cnt = 0
    for row in rows:
        values = [row.get(col) for col in usable_cols]
        cur.execute(sql, values)
        cnt += 1
    c.commit()
    c.close()
    return cnt


def sqlite_delete_missing_rows(table: str, conflict_col: str, remote_keys: list):
    c = conn()
    cur = c.cursor()
    if not remote_keys:
        cur.execute(f"DELETE FROM {table}")
        deleted = cur.rowcount if cur.rowcount is not None else 0
        c.commit()
        c.close()
        return deleted

    cur.execute(f"SELECT {conflict_col} FROM {table}")
    local_keys = [r[0] for r in cur.fetchall()]
    remote_set = {str(x) for x in remote_keys}
    to_delete = [x for x in local_keys if str(x) not in remote_set]
    deleted = 0
    if to_delete:
        for i in range(0, len(to_delete), 800):
            pack = to_delete[i:i+800]
            ph = ",".join(["?"] * len(pack))
            cur.execute(f"DELETE FROM {table} WHERE {conflict_col} IN ({ph})", tuple(pack))
            deleted += cur.rowcount if cur.rowcount is not None else 0
    c.commit()
    c.close()
    return deleted


def pull_shared_tables_from_supabase(force: bool = False, delete_missing: bool = True):
    if not supabase_enabled():
        return {"ok": False, "error": "not_configured"}

    now_ts = time.time()
    with _supabase_sync_lock:
        last_started = float(_supabase_sync_state.get("last_pull_started_ts") or 0.0)
        if (not force) and (now_ts - last_started) < SUPABASE_MIN_PULL_INTERVAL_SEC:
            return {"ok": True, "status": "throttled"}
        _supabase_sync_state["last_pull_started_ts"] = now_ts

    result = {"ok": True, "tables": {}, "pulled_at": now_iso()}
    fetched = {}

    for table, conflict_col in SUPABASE_PULL_TABLES:
        table_started = time.perf_counter()
        try:
            fetched[(table, conflict_col)] = supabase_select_rows(table, order_by=conflict_col)
        except Exception as e:
            result["ok"] = False
            result["tables"][table] = {"status": "error", "stage": "fetch", "error": str(e)}
        finally:
            elapsed = time.perf_counter() - table_started
            result["tables"].setdefault(table, {})["fetch_ms"] = round(elapsed * 1000, 2)

    for table, conflict_col in SUPABASE_PULL_TABLES:
        if (table, conflict_col) not in fetched:
            continue
        table_started = time.perf_counter()
        try:
            remote_rows = fetched[(table, conflict_col)]
            sqlite_upsert_rows(table, remote_rows, conflict_col)
            result["tables"].setdefault(table, {})["rows"] = len(remote_rows)
            result["tables"][table]["upsert"] = "ok"
        except Exception as e:
            result["ok"] = False
            result["tables"].setdefault(table, {})
            result["tables"][table].update({"status": "error", "stage": "upsert", "error": str(e)})
        finally:
            elapsed = time.perf_counter() - table_started
            result["tables"].setdefault(table, {})["upsert_ms"] = round(elapsed * 1000, 2)

    for table, conflict_col in reversed(SUPABASE_PULL_TABLES):
        if (table, conflict_col) not in fetched:
            continue
        if table in {"ksef_documents", "china_stock_receipts"}:
            result["tables"].setdefault(table, {})
            result["tables"][table]["deleted_local"] = 0
            if result["tables"][table].get("upsert") == "ok":
                result["tables"][table]["status"] = "ok"
            continue
        if not delete_missing:
            result["tables"].setdefault(table, {})["deleted_local"] = 0
            if result["tables"][table].get("upsert") == "ok":
                result["tables"][table]["status"] = "ok"
            continue
        try:
            remote_rows = fetched[(table, conflict_col)]
            remote_keys = [row.get(conflict_col) for row in remote_rows if row.get(conflict_col) is not None]
            deleted = sqlite_delete_missing_rows(table, conflict_col, remote_keys)
            result["tables"].setdefault(table, {})
            result["tables"][table]["deleted_local"] = deleted
            if result["tables"][table].get("upsert") == "ok":
                result["tables"][table]["status"] = "ok"
        except Exception as e:
            result["ok"] = False
            result["tables"].setdefault(table, {})
            result["tables"][table].update({"status": "error", "stage": "cleanup", "error": str(e)})

    try:
        normalize_temp_order_numbers()
        link_orders_to_customers_by_email(sync_remote=True)
    except Exception:
        pass
    return result


def _run_post_pull_reconciliation():
    started = time.perf_counter()
    try:
        reconcile_legacy_shipped_order_statuses()
        reconcile_paid_order_statuses()
        reconcile_legacy_orders_by_age()
    finally:
        elapsed = time.perf_counter() - started
        _perf_add("reconciliation", elapsed)
        if PERF_LOG_ENABLED:
            app.logger.info("PERF reconciliation %.2f ms", elapsed * 1000)


def trigger_background_supabase_pull(reason: str = "read"):
    """Schedule a full refresh without putting network I/O on a GET request."""
    if not supabase_enabled():
        return False, "not_configured"
    now_ts = time.time()
    with _supabase_sync_lock:
        if _supabase_sync_state.get("pull_running"):
            return False, "already_running"
        last_finished = float(_supabase_sync_state.get("last_pull_finished_ts") or 0.0)
        if now_ts - last_finished < SUPABASE_BACKGROUND_PULL_INTERVAL_SEC:
            return False, "throttled"
        _supabase_sync_state["pull_running"] = True

    def _job():
        started = time.perf_counter()
        try:
            with _supabase_full_io_lock:
                # Background refresh only merges remote rows. Deletions are
                # propagated by their explicit write endpoints; cleanup from a
                # possibly stale snapshot could otherwise erase a newer local write.
                result = pull_shared_tables_from_supabase(force=True, delete_missing=False)
                if result.get("ok"):
                    _run_post_pull_reconciliation()
            result["reason"] = reason
        except Exception as exc:
            result = {"ok": False, "error": str(exc), "reason": reason}
        finally:
            elapsed = time.perf_counter() - started
            result["total_ms"] = round(elapsed * 1000, 2)
            with _supabase_sync_lock:
                _supabase_sync_state["pull_running"] = False
                _supabase_sync_state["last_pull_finished_ts"] = time.time()
                _supabase_sync_state["last_pull_result"] = result
            if PERF_LOG_ENABLED:
                app.logger.info("PERF background_supabase_pull %s", json.dumps(result, ensure_ascii=False, sort_keys=True))

    threading.Thread(target=_job, daemon=True).start()
    return True, "started"


def maybe_pull_shared_from_supabase(force: bool = False):
    """Keep GET fast; protected write paths may still request a blocking refresh."""
    try:
        if request.method == "GET":
            # Po zimnym starcie Rendera lokalny SQLite bywa pusty. Pokazanie
            # użytkownikowi pulpitu z samymi zerami jest gorsze niż jednorazowe
            # oczekiwanie na odtworzenie danych. Blokujemy wyłącznie pierwszy
            # odczyt pustej instalacji; późniejsze GET-y nadal synchronizują się
            # w tle i nie czekają na sieć.
            with _supabase_sync_lock:
                initial_attempted = bool(_supabase_sync_state.get("initial_pull_attempted"))
            if not initial_attempted:
                if _local_supabase_bootstrap_complete():
                    with _supabase_sync_lock:
                        _supabase_sync_state["initial_pull_attempted"] = True
                else:
                    started = time.perf_counter()
                    with _supabase_full_io_lock:
                        with _supabase_sync_lock:
                            already_attempted = bool(_supabase_sync_state.get("initial_pull_attempted"))
                        if not already_attempted:
                            result = pull_shared_tables_from_supabase(force=True, delete_missing=False)
                            if result.get("ok"):
                                _run_post_pull_reconciliation()
                                _mark_local_supabase_bootstrap_complete()
                                with _supabase_sync_lock:
                                    _supabase_sync_state["initial_pull_attempted"] = True
                    _perf_add("supabase_initial_bootstrap", time.perf_counter() - started)
                    return result if not already_attempted else None
            return trigger_background_supabase_pull(reason=f"GET {request.path}")
        if force:
            started = time.perf_counter()
            with _supabase_full_io_lock:
                result = pull_shared_tables_from_supabase(force=True)
                if result.get("ok"):
                    _run_post_pull_reconciliation()
            _perf_add("supabase_pull_blocking", time.perf_counter() - started)
            return result
    except Exception as exc:
        app.logger.warning("Synchronizacja Supabase nie powiodła się: %s", type(exc).__name__)
    return None


def sync_local_rows_to_supabase(table: str, conflict_col: str, ids: list):
    ids = [x for x in ids if x is not None]
    if not ids or not supabase_enabled():
        return 0

    c = conn()
    cur = c.cursor()
    ph = ",".join(["?"] * len(ids))
    cur.execute(f"SELECT * FROM {table} WHERE {conflict_col} IN ({ph})", tuple(ids))
    rows = [dict(r) for r in cur.fetchall()]
    c.close()
    if rows:
        supabase_upsert_rows(table, rows, conflict_col)
    return len(rows)


def sync_local_table_to_supabase(table: str, conflict_col: str, chunk_size: int = 250):
    """Synchronizuj całą tabelę partiami, bez przekraczania limitu żądania."""
    if not supabase_enabled():
        return 0
    rows = sqlite_table_rows(table)
    for pack in _chunks(rows, max(1, int(chunk_size or 250))):
        supabase_upsert_rows(table, pack, conflict_col)
    return len(rows)


def sync_order_to_supabase(order_id: int):
    sync_local_rows_to_supabase("orders", "id", [order_id])
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT id FROM order_items WHERE order_id=?", (order_id,))
    item_ids = [int(r["id"]) for r in cur.fetchall()]
    c.close()
    if item_ids:
        sync_local_rows_to_supabase("order_items", "id", item_ids)


def remote_first_create_customer(
    name: str,
    address: str,
    phone: str,
    email: str,
    nip: str,
    language: str = "pl",
    price_list: str = "pln",
):
    language = normalize_client_language(language)
    price_list = normalize_client_price_list(price_list)
    created = supabase_insert_row("customers", {
        "name": name,
        "address": address,
        "phone": phone,
        "email": email,
        "nip": nip,
        "language": language,
        "price_list": price_list,
        "created_at": now_iso(),
    })
    if not created or "id" not in created:
        raise RuntimeError("Supabase nie zwrĂłciĹ‚ ID dla klienta")

    customer_id = int(created["id"])
    c = conn()
    cur = c.cursor()
    cur.execute(
        "INSERT INTO customers(id, name, address, phone, email, nip, language, price_list, created_at) VALUES (?,?,?,?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET name=excluded.name, address=excluded.address, phone=excluded.phone, email=excluded.email, nip=excluded.nip, language=excluded.language, price_list=excluded.price_list, created_at=excluded.created_at",
        (customer_id, name, address, phone, email, nip, language, price_list, created.get("created_at") or now_iso())
    )
    c.commit()
    c.close()
    return customer_id


def remote_first_create_order(
    customer_id,
    customer_name,
    customer_address,
    customer_phone,
    customer_email,
    note,
    items,
    idempotency_key=None,
    price_list="pln",
    currency="PLN",
):
    created_at = now_iso()
    price_list = normalize_client_price_list(price_list)
    currency = normalize_order_currency(currency or price_list_currency(price_list))
    order_payload = {
        "order_no": "TEMP",
        "customer_id": customer_id if customer_id else None,
        "customer_name": customer_name,
        "customer_address": customer_address,
        "customer_phone": customer_phone,
        "customer_email": customer_email,
        "status": "new",
        "note": note,
        "created_at": created_at,
        "qr_data_url": "",
        "price_list": price_list,
        "currency": currency,
    }
    if idempotency_key:
        order_payload["idempotency_key"] = idempotency_key
    created_order = supabase_insert_row("orders", order_payload)
    if not created_order or "id" not in created_order:
        raise RuntimeError("Supabase nie zwrĂłciĹ‚ ID dla zamĂłwienia")

    order_id = int(created_order["id"])
    order_no = make_order_no(order_id, created_at)
    qr_data_url = ""
    supabase_update_rows("orders", {"order_no": order_no, "qr_data_url": qr_data_url}, {"id": order_id})

    c = conn()
    try:
        cur = c.cursor()
        cur.execute(
            "INSERT INTO orders(id, order_no, customer_id, customer_name, customer_address, customer_phone, customer_email, status, note, created_at, qr_data_url, idempotency_key, price_list, currency) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET order_no=excluded.order_no, customer_id=excluded.customer_id, customer_name=excluded.customer_name, customer_address=excluded.customer_address, customer_phone=excluded.customer_phone, customer_email=excluded.customer_email, status=excluded.status, note=excluded.note, created_at=excluded.created_at, qr_data_url=excluded.qr_data_url, idempotency_key=excluded.idempotency_key, price_list=excluded.price_list, currency=excluded.currency",
            (order_id, order_no, customer_id if customer_id else None, customer_name, customer_address, customer_phone, customer_email, "new", note, created_at, qr_data_url, idempotency_key, price_list, currency)
        )

        for raw_item in items:
            if isinstance(raw_item, dict):
                pid = to_int(raw_item.get("product_id"), 0)
                qty = to_int(raw_item.get("qty"), 0)
                unit_net_price = money_float(raw_item.get("unit_net_price"))
                unit_gross_price = money_float(raw_item.get("unit_gross_price"))
                unit_retail_price = money_float(raw_item.get("unit_retail_price"))
                item_currency = normalize_order_currency(raw_item.get("currency") or currency)
            else:
                pid, qty = raw_item[:2]
                unit_net_price = unit_gross_price = unit_retail_price = None
                item_currency = currency
            cur.execute("SELECT sku FROM products WHERE id=?", (pid,))
            p = cur.fetchone()
            if not p:
                raise ValueError(f"Nie istnieje produkt ID {pid}")
            item_payload = {
                "order_id": order_id,
                "product_id": pid,
                "sku": p["sku"],
                "qty": qty,
                "created_at": now_iso(),
                "currency": item_currency,
            }
            if unit_net_price is not None:
                item_payload.update({
                    "unit_net_price": unit_net_price,
                    "unit_gross_price": unit_gross_price,
                    "unit_retail_price": unit_retail_price,
                })
            created_item = supabase_insert_row("order_items", item_payload)
            if not created_item or "id" not in created_item:
                raise RuntimeError("Supabase nie zwrócił ID dla pozycji zamówienia")
            cur.execute(
                "INSERT INTO order_items(id, order_id, product_id, sku, qty, unit_net_price, unit_gross_price, unit_retail_price, currency, created_at) VALUES (?,?,?,?,?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET order_id=excluded.order_id, product_id=excluded.product_id, sku=excluded.sku, qty=excluded.qty, unit_net_price=excluded.unit_net_price, unit_gross_price=excluded.unit_gross_price, unit_retail_price=excluded.unit_retail_price, currency=excluded.currency, created_at=excluded.created_at",
                (int(created_item["id"]), order_id, pid, p["sku"], qty, unit_net_price, unit_gross_price, unit_retail_price, item_currency, created_item.get("created_at") or now_iso())
            )
        c.commit()
    except Exception:
        c.rollback()
        try:
            supabase_delete_rows("order_items", {"order_id": order_id})
            supabase_delete_rows("orders", {"id": order_id})
        except Exception as rollback_exc:
            app.logger.error("Niepełny rollback zamówienia order_id=%s: %s", order_id, rollback_exc)
        raise
    finally:
        c.close()
    try:
        normalize_temp_order_numbers()
    except Exception:
        pass
    return order_id


def get_stock(product_id):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT qty FROM stock WHERE product_id=?", (product_id,))
    r = cur.fetchone()
    c.close()
    return int(r["qty"]) if r else 0

def change_stock(product_id, delta):
    c = conn()
    cur = c.cursor()
    cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (product_id,))
    cur.execute("UPDATE stock SET qty = qty + ? WHERE product_id=?", (delta, product_id))
    c.commit()
    c.close()

def safe_filename(s):
    s = re.sub(r"[^a-zA-Z0-9_\-\.]+", "_", s)
    return s[:80] if s else "file"


def invoice_dir_for_customer(customer_name: str) -> str:
    root = os.path.join(DATA_DIR, "faktury")
    os.makedirs(root, exist_ok=True)
    customer_dir = os.path.join(root, safe_filename(customer_name or "klient"))
    os.makedirs(customer_dir, exist_ok=True)
    return customer_dir


def get_pdf_font_names():
    regular = "Helvetica"
    bold = "Helvetica-Bold"

    # Szukaj czcionek Unicode takĹĽe po wildcardach i lokalnym katalogu app/fonts.
    regular_candidates = [
        # Lokalne fonty aplikacji (najwyĹĽszy priorytet)
        ("AppFont-Regular", os.path.join(APP_DIR, "fonts", "regular.ttf")),
        ("AppFontRoot-Regular", os.path.join(APP_DIR, "regular.ttf")),

        # Linux
        ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        ("DejaVuSansCondensed", "/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed.ttf"),
        ("LiberationSans", "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
        ("NotoSans", "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf"),

        # Windows
        ("Arial", r"C:\Windows\Fonts\arial.ttf"),
        ("Calibri", r"C:\Windows\Fonts\calibri.ttf"),
        ("Tahoma", r"C:\Windows\Fonts\tahoma.ttf"),

        # macOS
        ("ArialMT", "/System/Library/Fonts/Supplemental/Arial.ttf"),
        ("HelveticaNeue", "/System/Library/Fonts/Helvetica.ttc"),
    ]
    bold_candidates = [
        ("AppFont-Bold", os.path.join(APP_DIR, "fonts", "bold.ttf")),
        ("AppFontRoot-Bold", os.path.join(APP_DIR, "bold.ttf")),

        # Linux
        ("DejaVuSans-Bold", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("DejaVuSansCondensed-Bold", "/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed-Bold.ttf"),
        ("LiberationSans-Bold", "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
        ("NotoSans-Bold", "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf"),

        # Windows
        ("Arial-Bold", r"C:\Windows\Fonts\arialbd.ttf"),
        ("Calibri-Bold", r"C:\Windows\Fonts\calibrib.ttf"),
        ("Tahoma-Bold", r"C:\Windows\Fonts\tahomabd.ttf"),

        # macOS
        ("Arial-BoldMT", "/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
    ]

    # Dodatkowe wildcardy gdy Ĺ›cieĹĽki systemowe rĂłĹĽniÄ… siÄ™ miÄ™dzy maszynami.
    for path in glob.glob('/usr/share/fonts/**/*DejaVuSans*.ttf', recursive=True)[:6]:
        regular_candidates.append((f"AutoReg-{safe_filename(os.path.basename(path))}", path))
    for path in glob.glob('/usr/share/fonts/**/*NotoSans*.ttf', recursive=True)[:6]:
        regular_candidates.append((f"AutoReg-{safe_filename(os.path.basename(path))}", path))
    for path in glob.glob('/usr/share/fonts/**/*LiberationSans*.ttf', recursive=True)[:6]:
        regular_candidates.append((f"AutoReg-{safe_filename(os.path.basename(path))}", path))

    for path in glob.glob('/usr/share/fonts/**/*Bold*.ttf', recursive=True)[:10]:
        bold_candidates.append((f"AutoBold-{safe_filename(os.path.basename(path))}", path))

    def register_first(candidates):
        for name, path in candidates:
            if not path or not os.path.exists(path):
                continue
            try:
                if name not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(name, path))
                return name
            except Exception:
                continue
        return None

    reg = register_first(regular_candidates)
    bld = register_first(bold_candidates)

    if reg:
        regular = reg
    if bld:
        bold = bld
    elif reg:
        bold = reg

    return regular, bold


CLIENT_LANGUAGES = {"pl", "de", "en", "es", "it"}
CLIENT_PRICE_LISTS = {"pln", "eu_eur"}

ORDER_PDF_TRANSLATIONS = {
    "pl": {
        "title": "ZAMÓWIENIE", "order_number": "Numer zamówienia", "date": "Data",
        "customer": "Klient", "delivery_address": "Adres dostawy", "status": "Status",
        "email": "E-mail", "phone": "Telefon", "notes": "Uwagi klienta",
        "sku": "SKU", "product": "Model / nazwa", "quantity": "Ilość",
        "unit_net": "Cena jedn. netto", "line_net": "Wartość netto",
        "unit_gross": "Cena jedn. brutto", "line_gross": "Wartość brutto",
        "net_total": "Razem netto", "vat": "VAT", "gross_total": "Razem brutto",
        "currency": "PLN", "page": "Strona",
        "status_unconfirmed": "Niepotwierdzone", "status_confirmed": "Potwierdzone",
        "status_in_delivery": "W dostawie", "status_completed": "Zrealizowane",
        "status_cancelled": "Anulowane", "status_other": "Inny",
    },
    "de": {
        "title": "BESTELLUNG", "order_number": "Bestellnummer", "date": "Datum",
        "customer": "Kunde", "delivery_address": "Lieferadresse", "status": "Status",
        "email": "E-Mail", "phone": "Telefon", "notes": "Kundenhinweise",
        "sku": "SKU", "product": "Modell / Bezeichnung", "quantity": "Menge",
        "unit_net": "Nettostückpreis", "line_net": "Nettowert",
        "unit_gross": "Bruttostückpreis", "line_gross": "Bruttowert",
        "net_total": "Nettobetrag", "vat": "MwSt.", "gross_total": "Gesamtbetrag",
        "currency": "PLN", "page": "Seite",
        "status_unconfirmed": "Unbestätigt", "status_confirmed": "Bestätigt",
        "status_in_delivery": "In Lieferung", "status_completed": "Abgeschlossen",
        "status_cancelled": "Storniert", "status_other": "Sonstiger",
    },
    "en": {
        "title": "ORDER", "order_number": "Order number", "date": "Date",
        "customer": "Customer", "delivery_address": "Delivery address", "status": "Status",
        "email": "Email", "phone": "Phone", "notes": "Customer notes",
        "sku": "SKU", "product": "Model / product", "quantity": "Quantity",
        "unit_net": "Net unit price", "line_net": "Net value",
        "unit_gross": "Gross unit price", "line_gross": "Gross value",
        "net_total": "Net total", "vat": "VAT", "gross_total": "Gross total",
        "currency": "PLN", "page": "Page",
        "status_unconfirmed": "Unconfirmed", "status_confirmed": "Confirmed",
        "status_in_delivery": "In delivery", "status_completed": "Completed",
        "status_cancelled": "Cancelled", "status_other": "Other",
    },
    "es": {
        "title": "PEDIDO", "order_number": "Número de pedido", "date": "Fecha",
        "customer": "Cliente", "delivery_address": "Dirección de entrega", "status": "Estado",
        "email": "Correo electrónico", "phone": "Teléfono", "notes": "Notas del cliente",
        "sku": "SKU", "product": "Modelo / producto", "quantity": "Cantidad",
        "unit_net": "Precio unitario neto", "line_net": "Importe neto",
        "unit_gross": "Precio unitario bruto", "line_gross": "Importe bruto",
        "net_total": "Total neto", "vat": "IVA", "gross_total": "Total bruto",
        "currency": "PLN", "page": "Página",
        "status_unconfirmed": "Sin confirmar", "status_confirmed": "Confirmado",
        "status_in_delivery": "En entrega", "status_completed": "Completado",
        "status_cancelled": "Cancelado", "status_other": "Otro",
    },
    "it": {
        "title": "ORDINE", "order_number": "Numero ordine", "date": "Data",
        "customer": "Cliente", "delivery_address": "Indirizzo di consegna", "status": "Stato",
        "email": "E-mail", "phone": "Telefono", "notes": "Note del cliente",
        "sku": "SKU", "product": "Modello / prodotto", "quantity": "Quantità",
        "unit_net": "Prezzo unitario netto", "line_net": "Valore netto",
        "unit_gross": "Prezzo unitario lordo", "line_gross": "Valore lordo",
        "net_total": "Totale netto", "vat": "IVA", "gross_total": "Totale lordo",
        "currency": "PLN", "page": "Pagina",
        "status_unconfirmed": "Non confermato", "status_confirmed": "Confermato",
        "status_in_delivery": "In consegna", "status_completed": "Completato",
        "status_cancelled": "Annullato", "status_other": "Altro",
    },
}

PACKING_LIST_TRANSLATIONS = {
    "pl": {
        "title": "LISTA PAKOWANIA", "invoice": "Faktura", "continued": "dalszy ciąg",
        "order": "ZAMÓWIENIE", "date": "DATA", "customer": "KLIENT", "checked": "✓",
        "line": "LP.", "product": "MODEL / NAZWA", "source": "ZAMÓWIENIE / NOTATKA",
        "quantity": "ILOŚĆ", "positions": "Pozycje", "total_qty": "Razem sztuk",
        "packages": "Liczba pudełek", "packed_by": "Spakował(a)", "signature": "Podpis",
    },
    "de": {
        "title": "PACKLISTE", "invoice": "Rechnung", "continued": "Fortsetzung",
        "order": "BESTELLUNG", "date": "DATUM", "customer": "KUNDE", "checked": "✓",
        "line": "POS.", "product": "MODELL / BEZEICHNUNG", "source": "BESTELLUNG / NOTIZ",
        "quantity": "MENGE", "positions": "Positionen", "total_qty": "Gesamtmenge",
        "packages": "Anzahl Pakete", "packed_by": "Gepackt von", "signature": "Unterschrift",
    },
    "en": {
        "title": "PACKING LIST", "invoice": "Invoice", "continued": "continued",
        "order": "ORDER", "date": "DATE", "customer": "CUSTOMER", "checked": "✓",
        "line": "NO.", "product": "MODEL / PRODUCT", "source": "ORDER / NOTE",
        "quantity": "QTY", "positions": "Items", "total_qty": "Total quantity",
        "packages": "Packages", "packed_by": "Packed by", "signature": "Signature",
    },
    "es": {
        "title": "LISTA DE EMBALAJE", "invoice": "Factura", "continued": "continuación",
        "order": "PEDIDO", "date": "FECHA", "customer": "CLIENTE", "checked": "✓",
        "line": "N.º", "product": "MODELO / PRODUCTO", "source": "PEDIDO / NOTA",
        "quantity": "CANT.", "positions": "Artículos", "total_qty": "Cantidad total",
        "packages": "Número de paquetes", "packed_by": "Preparado por", "signature": "Firma",
    },
    "it": {
        "title": "LISTA DI IMBALLAGGIO", "invoice": "Fattura", "continued": "continua",
        "order": "ORDINE", "date": "DATA", "customer": "CLIENTE", "checked": "✓",
        "line": "N.", "product": "MODELLO / PRODOTTO", "source": "ORDINE / NOTA",
        "quantity": "Q.TÀ", "positions": "Articoli", "total_qty": "Quantità totale",
        "packages": "Numero colli", "packed_by": "Preparato da", "signature": "Firma",
    },
}


def normalize_client_language(value) -> str:
    language = norm(value).lower()
    return language if language in CLIENT_LANGUAGES else "pl"


def normalize_client_price_list(value) -> str:
    price_list = norm(value).lower()
    return price_list if price_list in CLIENT_PRICE_LISTS else "pln"


def price_list_for_language(language) -> str:
    """Polish accounts use PLN; every supported foreign language uses EUR."""
    return "pln" if normalize_client_language(language) == "pl" else "eu_eur"


def price_list_currency(value) -> str:
    return "EUR" if normalize_client_price_list(value) == "eu_eur" else "PLN"


def normalize_order_currency(value) -> str:
    currency = norm(value).upper()
    return currency if re.fullmatch(r"[A-Z]{3}", currency or "") else "PLN"


EU_VAT_PREFIXES = {"AT","BE","BG","HR","CY","CZ","DK","EE","FI","FR","DE","EL","HU","IE","IT","LV","LT","LU","MT","NL","PT","RO","SK","SI","ES","SE"}


def automatic_invoice_tax_context(order, buyer_tax_no="", buyer_country=""):
    """Derive new invoice settings from the persisted order, never language."""
    order = order or {}
    currency = normalize_order_currency(order.get("currency") or price_list_currency(order.get("price_list")))
    price_list = normalize_client_price_list(order.get("price_list"))
    tax_id = re.sub(r"[\s.\-]+", "", norm(buyer_tax_no).upper())
    country = norm(buyer_country).upper()
    aliases = {"DEUTSCHLAND":"DE", "GERMANY":"DE", "NIEMCY":"DE", "POLSKA":"PL", "POLAND":"PL"}
    country = aliases.get(country, country)
    prefix = tax_id[:2] if len(tax_id) >= 2 else ""
    if prefix in EU_VAT_PREFIXES:
        country = prefix
    invoice_type = "wdt" if price_list == "eu_eur" or currency == "EUR" else "domestic"
    return invoice_type, currency, country


def order_pdf_text(language: str, key: str) -> str:
    language = normalize_client_language(language)
    return ORDER_PDF_TRANSLATIONS.get(language, {}).get(key) or ORDER_PDF_TRANSLATIONS["pl"].get(key, "")


def packing_list_text(language: str, key: str) -> str:
    language = normalize_client_language(language)
    return PACKING_LIST_TRANSLATIONS.get(language, {}).get(key) or PACKING_LIST_TRANSLATIONS["pl"].get(key, "")


def localized_pdf_money(value, language: str) -> str:
    amount = money_dec(value)
    raw = f"{amount:,.2f}"
    if normalize_client_language(language) == "en":
        return raw
    return raw.replace(",", "\u0000").replace(".", ",").replace("\u0000", " ")


def localized_pdf_date(value, language: str) -> str:
    raw = norm(value).replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(raw)
    except Exception:
        try:
            parsed = datetime.strptime(raw[:10], "%Y-%m-%d")
        except Exception:
            return norm(value)
    if normalize_client_language(language) in {"es", "en"}:
        return parsed.strftime("%d/%m/%Y")
    return parsed.strftime("%d.%m.%Y")


def order_pdf_status(status, language: str) -> str:
    key = norm(status).lower()
    if key in {"new", "pending", "unconfirmed"}:
        label_key = "status_unconfirmed"
    elif key in {"confirmed", "packed", "packed_partial", "issued"}:
        label_key = "status_confirmed"
    elif key in {"in_delivery", "shipped", "partially_shipped"}:
        label_key = "status_in_delivery"
    elif key == "completed":
        label_key = "status_completed"
    elif key == "cancelled":
        label_key = "status_cancelled"
    else:
        label_key = "status_other"
    return order_pdf_text(language, label_key)


def generate_client_order_pdf(
    order_row: dict,
    items: list[dict],
    language: str,
    retail_prices: bool = False,
) -> tuple[io.BytesIO, str]:
    """Create one localized order PDF without persisting a second document copy."""
    language = normalize_client_language(language)
    currency = normalize_order_currency(order_row.get("currency"))
    regular_font, bold_font = get_pdf_font_names()
    page_width, page_height = 210 * mm, 297 * mm
    left, right = 15 * mm, 195 * mm
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(page_width, page_height))
    # Wersja przyjazna drukarkom laserowym: wysoki kontrast, neutralne szarosci
    # i grubsze linie zamiast delikatnych ekranowych odcieni niebieskiego.
    ink = (0.0, 0.0, 0.0)
    muted_ink = (0.12, 0.12, 0.12)
    pale_gray = (0.91, 0.91, 0.91)
    rule_gray = (0.38, 0.38, 0.38)

    c = conn()
    try:
        company_row = c.execute("SELECT * FROM company_profile WHERE id=1").fetchone()
        company = dict(company_row) if company_row else {}
    finally:
        c.close()

    def txt(value) -> str:
        return fix_polish_mojibake(norm(value))

    def fit(value, font_name, size, max_width) -> str:
        value = txt(value)
        if pdfmetrics.stringWidth(value, font_name, size) <= max_width:
            return value
        suffix = "..."
        while value and pdfmetrics.stringWidth(value + suffix, font_name, size) > max_width:
            value = value[:-1]
        return value + suffix if value else ""

    def wrap(value, font_name, size, max_width, max_lines=3):
        words = txt(value).split()
        lines, current = [], ""
        for word in words:
            candidate = word if not current else f"{current} {word}"
            if pdfmetrics.stringWidth(candidate, font_name, size) <= max_width:
                current = candidate
            else:
                if current:
                    lines.append(current)
                current = fit(word, font_name, size, max_width)
                if len(lines) >= max_lines:
                    break
        if current and len(lines) < max_lines:
            lines.append(current)
        return lines or ["-"]

    def draw_complete(value, x, y, font_name, size, max_width, min_size=6.2):
        """Draw the complete value, reducing the font instead of adding ellipsis."""
        value = txt(value)
        fitted_size = size
        while fitted_size > min_size and pdfmetrics.stringWidth(value, font_name, fitted_size) > max_width:
            fitted_size -= 0.2
        pdf.setFont(font_name, max(fitted_size, min_size))
        pdf.drawString(x, y, value)

    page_no = 0
    order_number = canonical_order_no(
        order_row.get("id"), order_row.get("created_at"), order_row.get("order_no")
    )

    def footer():
        pdf.setStrokeColorRGB(*rule_gray)
        pdf.setLineWidth(0.7)
        pdf.line(left, 12 * mm, right, 12 * mm)
        pdf.setFont(bold_font, 7.5)
        pdf.setFillColorRGB(*muted_ink)
        pdf.drawString(left, 7.5 * mm, txt(company.get("company_name") or "Niedźwieccy"))
        pdf.drawRightString(right, 7.5 * mm, f"{order_pdf_text(language, 'page')} {page_no}")

    def new_page(first=False):
        nonlocal page_no
        if not first:
            footer()
            pdf.showPage()
        page_no += 1
        # Czysty, biały nagłówek. Logo i typografia tworzą hierarchię bez
        # ciężkiego, ciemnego pasa zajmującego całą szerokość dokumentu.
        logo_path = find_logo_path()
        if logo_path:
            try:
                pdf.drawImage(ImageReader(logo_path), left, page_height - 42 * mm, 54.4 * mm, 34 * mm,
                              preserveAspectRatio=True, anchor="w", mask="auto")
            except Exception:
                pass
        pdf.setFillColorRGB(*ink)
        pdf.setFont(bold_font, 18)
        pdf.drawRightString(right, page_height - 15 * mm, order_pdf_text(language, "title"))
        pdf.setFillColorRGB(*muted_ink)
        pdf.setFont(bold_font, 9)
        pdf.drawRightString(right, page_height - 21 * mm, fit(order_number, regular_font, 9, 80 * mm))
        pdf.setStrokeColorRGB(*rule_gray)
        pdf.setLineWidth(0.8)
        pdf.line(left, page_height - 45 * mm, right, page_height - 45 * mm)
        return page_height - 55 * mm

    def draw_table_header(y):
        pdf.setFillColorRGB(*pale_gray)
        pdf.setStrokeColorRGB(*rule_gray)
        pdf.setLineWidth(0.8)
        pdf.roundRect(left, y - 7 * mm, right - left, 8 * mm, 2 * mm, fill=1, stroke=1)
        pdf.setFillColorRGB(*ink)
        pdf.setFont(bold_font, 8.0)
        pdf.drawString(left + 2 * mm, y - 4 * mm, order_pdf_text(language, "sku"))
        pdf.drawString(left + 43 * mm, y - 4 * mm, order_pdf_text(language, "product"))
        pdf.drawRightString(left + 122 * mm, y - 4 * mm, order_pdf_text(language, "quantity"))
        unit_price_key = "unit_gross" if retail_prices else "unit_net"
        line_value_key = "line_gross" if retail_prices else "line_net"
        pdf.drawRightString(left + 152 * mm, y - 4 * mm, order_pdf_text(language, unit_price_key))
        pdf.drawRightString(right - 2 * mm, y - 4 * mm, order_pdf_text(language, line_value_key))
        return y - 10 * mm

    y = new_page(first=True)
    pdf.setFillColorRGB(*ink)
    pdf.setFont(bold_font, 9)
    pdf.drawString(left, y, order_pdf_text(language, "order_number"))
    pdf.drawString(left + 72 * mm, y, order_pdf_text(language, "date"))
    pdf.drawString(left + 117 * mm, y, order_pdf_text(language, "status"))
    pdf.setFont(bold_font, 9.5)
    pdf.drawString(left, y - 5 * mm, fit(order_number, bold_font, 9.5, 65 * mm))
    pdf.drawString(left + 72 * mm, y - 5 * mm, localized_pdf_date(order_row.get("created_at"), language))
    pdf.drawString(left + 117 * mm, y - 5 * mm, order_pdf_status(order_row.get("status"), language))
    y -= 15 * mm

    customer_name = order_row.get("customer_name") or "-"
    customer_address = order_row.get("customer_address") or ""
    pdf.setFont(bold_font, 9)
    pdf.drawString(left, y, order_pdf_text(language, "customer"))
    pdf.drawString(left + 95 * mm, y, order_pdf_text(language, "delivery_address"))
    pdf.setFont(bold_font, 8.8)
    customer_lines = wrap(customer_name, bold_font, 8.8, 82 * mm, 2)
    contact = " - ".join(x for x in [txt(order_row.get("customer_email")), txt(order_row.get("customer_phone"))] if x)
    if contact:
        customer_lines.extend(wrap(contact, regular_font, 7.5, 82 * mm, 2))
    address_lines = wrap(customer_address or "-", regular_font, 8.5, 82 * mm, 3)
    max_lines = max(len(customer_lines), len(address_lines))
    for idx in range(max_lines):
        if idx < len(customer_lines):
            pdf.drawString(left, y - (idx + 1) * 4.5 * mm, customer_lines[idx])
        if idx < len(address_lines):
            pdf.drawString(left + 95 * mm, y - (idx + 1) * 4.5 * mm, address_lines[idx])
    y -= (max_lines + 2) * 4.5 * mm

    y = draw_table_header(y)
    total_net = Decimal("0.00")
    total_gross = Decimal("0.00")
    for item in items:
        qty = to_int(item.get("qty"), 0)
        unit_net = money_dec(item.get("net_price"))
        unit_gross = money_dec(item.get("gross_price"))
        if retail_prices:
            # Cennik detaliczny jest zawsze wyliczany od ceny netto B2B.
            # Brutto liczymy od wartości źródłowej, a kwoty zaokrąglamy dopiero
            # na wyjściu — dokładnie tak samo jak ceny detaliczne w panelu.
            retail_net = unit_net * Decimal("1.45")
            unit_net = retail_net.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
            unit_gross = (retail_net * Decimal("1.23")).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        line_net = (unit_net * qty).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        line_gross = (unit_gross * qty).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        total_net += line_net
        total_gross += line_gross
        if y < 31 * mm:
            y = new_page()
            y = draw_table_header(y)
        pdf.setFillColorRGB(*ink)
        product_label = " / ".join(x for x in [txt(item.get("model")), txt(item.get("name"))] if x) or "-"
        draw_complete(item.get("sku"), left + 2 * mm, y, bold_font, 8.2, 38 * mm)
        pdf.setFont(bold_font, 8.2)
        pdf.drawString(left + 43 * mm, y, fit(product_label, bold_font, 8.2, 45 * mm))
        pdf.drawRightString(left + 122 * mm, y, str(qty))
        displayed_unit_price = unit_gross if retail_prices else unit_net
        displayed_line_value = line_gross if retail_prices else line_net
        pdf.drawRightString(left + 152 * mm, y, f"{localized_pdf_money(displayed_unit_price, language)} {currency}")
        pdf.drawRightString(right - 2 * mm, y, f"{localized_pdf_money(displayed_line_value, language)} {currency}")
        pdf.setStrokeColorRGB(*rule_gray)
        pdf.setLineWidth(0.7)
        pdf.line(left, y - 2.5 * mm, right, y - 2.5 * mm)
        y -= 7 * mm

    total_net = total_net.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    total_gross = total_gross.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    vat_value = (total_gross - total_net).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    if y < 48 * mm:
        y = new_page()
    y -= 3 * mm
    pdf.setFillColorRGB(*ink)
    totals_to_draw = (
        (("gross_total", total_gross, True),)
        if retail_prices
        else (("net_total", total_net, False), ("vat", vat_value, False), ("gross_total", total_gross, True))
    )
    for key, value, is_bold in totals_to_draw:
        pdf.setFont(bold_font if is_bold else regular_font, 10 if is_bold else 9)
        pdf.drawRightString(right, y, f"{order_pdf_text(language, key)}: {localized_pdf_money(value, language)} {currency}")
        y -= 5.5 * mm

    note = norm(order_row.get("note"))
    if note:
        if y < 34 * mm:
            y = new_page()
        pdf.setFont(bold_font, 9)
        pdf.drawString(left, y, order_pdf_text(language, "notes"))
        pdf.setFont(regular_font, 8.5)
        for line in wrap(note, regular_font, 8.5, right - left, 5):
            y -= 4.5 * mm
            pdf.drawString(left, y, line)

    footer()
    pdf.save()
    buffer.seek(0)
    pdf_variant = "_RETAIL" if retail_prices else ""
    filename = f"ORDER_{safe_filename(order_number)}{pdf_variant}_{language}.pdf"
    return buffer, filename


def generate_sales_invoice(order_row, items):
    customer_dir = invoice_dir_for_customer(order_row["customer_name"])
    fname = f"FV_{safe_filename(canonical_order_no(order_row['id'] if 'id' in order_row.keys() else None, order_row['created_at'] if 'created_at' in order_row.keys() else '', order_row['order_no']))}.pdf"
    fpath = os.path.join(customer_dir, fname)

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM company_profile WHERE id=1")
    company = cur.fetchone()
    cur.execute("SELECT model, net_price, gross_price FROM pricing")
    pricing_rows = cur.fetchall()
    cur.execute("SELECT sku, model, name FROM products")
    product_rows = cur.fetchall()
    customer_language = "pl"
    try:
        customer_id = int(order_row["customer_id"] or 0) if order_row and "customer_id" in order_row.keys() else 0
        if customer_id:
            cur.execute("SELECT language FROM customers WHERE id=? LIMIT 1", (customer_id,))
            lang_row = cur.fetchone()
            if lang_row:
                customer_language = normalize_client_language(lang_row["language"])
    except Exception:
        customer_language = "pl"
    c.close()

    pricing_map = {norm(r["model"]): r for r in pricing_rows}
    product_map = {norm(r["sku"]): r for r in product_rows}

    def pdf_txt(value) -> str:
        return fix_polish_mojibake(norm(value))

    def fit_pdf_text(value, font_name, font_size, max_width, suffix="...") -> str:
        text = pdf_txt(value)
        if pdfmetrics.stringWidth(text, font_name, font_size) <= max_width:
            return text
        while text and pdfmetrics.stringWidth(text + suffix, font_name, font_size) > max_width:
            text = text[:-1]
        return (text + suffix) if text else ""

    w = 210 * mm
    h = 297 * mm
    cpdf = canvas.Canvas(fpath, pagesize=(w, h))

    pdf_font, pdf_font_bold = get_pdf_font_names()

    y = h - 18 * mm
    cpdf.setFont(pdf_font_bold, 14)
    cpdf.drawString(15 * mm, y, f"Faktura sprzedaĹĽowa: {canonical_order_no(order_row['id'] if 'id' in order_row.keys() else None, order_row['created_at'] if 'created_at' in order_row.keys() else '', order_row['order_no'])}")

    y -= 8 * mm
    cpdf.setFont(pdf_font, 10)
    cpdf.drawString(15 * mm, y, f"Data: {order_row['created_at']}")

    y -= 9 * mm
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawString(15 * mm, y, "Sprzedawca:")
    y -= 6 * mm
    cpdf.setFont(pdf_font, 8.5)
    if company:
        cpdf.drawString(15 * mm, y, f"{company['company_name'] or '-'}")
        y -= 5 * mm
        for ln in (company["address"] or "-").splitlines():
            cpdf.drawString(15 * mm, y, ln)
            y -= 5 * mm
        cpdf.drawString(15 * mm, y, f"NIP: {company['nip'] or '-'}")
        y -= 5 * mm
        cpdf.drawString(15 * mm, y, f"Tel: {company['phone'] or '-'}  Email: {company['email'] or '-'}")
        y -= 5 * mm
        cpdf.drawString(15 * mm, y, f"Konto: {company['bank_account'] or '-'}")
    else:
        cpdf.drawString(15 * mm, y, "Brak danych firmy (uzupeĹ‚nij w zakĹ‚adce: Dane mojej firmy)")

    y -= 8 * mm
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawString(15 * mm, y, "Nabywca:")
    y -= 6 * mm
    cpdf.setFont(pdf_font, 9)
    cpdf.drawString(15 * mm, y, f"{order_row['customer_name'] or '-'}")
    y -= 5 * mm
    for ln in (order_row["customer_address"] or "-").splitlines():
        cpdf.drawString(15 * mm, y, ln)
        y -= 5 * mm
    cpdf.drawString(15 * mm, y, f"Tel: {order_row['customer_phone'] or '-'}  Email: {order_row['customer_email'] or '-'}")

    y -= 10 * mm
    cpdf.setFont(pdf_font_bold, 9)
    cpdf.drawString(15 * mm, y, "SKU")
    cpdf.drawString(45 * mm, y, "Model")
    cpdf.drawString(95 * mm, y, "IloĹ›Ä‡")
    cpdf.drawString(112 * mm, y, "Netto/szt")
    cpdf.drawString(140 * mm, y, "Brutto/szt")
    cpdf.drawString(170 * mm, y, "WartoĹ›Ä‡ brutto")
    y -= 5 * mm

    total_net = 0.0
    total_gross = 0.0
    cpdf.setFont(pdf_font, 9)

    for it in items:
        model = norm(it["model"])
        pr = pricing_map.get(model)
        net = float(pr["net_price"]) if pr else 0.0
        gross = float(pr["gross_price"]) if pr else 0.0
        qty = int(it["qty"])
        line_net = net * qty
        line_gross = gross * qty
        total_net += line_net
        total_gross += line_gross

        cpdf.drawString(15 * mm, y, it["sku"])
        cpdf.drawString(45 * mm, y, (model or "-")[:24])
        cpdf.drawRightString(108 * mm, y, str(qty))
        cpdf.drawRightString(136 * mm, y, f"{net:.2f}")
        cpdf.drawRightString(164 * mm, y, f"{gross:.2f}")
        cpdf.drawRightString(195 * mm, y, f"{line_gross:.2f}")
        y -= 5 * mm

        if y < 28 * mm:
            cpdf.showPage()
            y = h - 20 * mm
            cpdf.setFont(pdf_font, 9)

    y -= 6 * mm
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawRightString(195 * mm, y, f"Suma netto: {total_net:.2f} PLN")
    y -= 5 * mm
    cpdf.drawRightString(195 * mm, y, f"Suma brutto: {total_gross:.2f} PLN")

    y -= 8 * mm
    cpdf.setFont(pdf_font, 9)
    cpdf.drawString(15 * mm, y, "Ceny pobrane z zakĹ‚adki Cennik (model, netto, brutto).")

    cpdf.save()
    return fpath


def _legacy_generate_order_invoice_pdf(order_row, items, meta):
    customer_dir = invoice_dir_for_customer(meta.get("buyer_name") or (order_row["customer_name"] if order_row and "customer_name" in order_row.keys() else "") or "Klient")
    fname = f"{safe_filename(meta['invoice_no'])}.pdf"
    fpath = os.path.join(customer_dir, fname)

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM company_profile WHERE id=1")
    company = cur.fetchone()
    cur.execute("SELECT model, net_price, gross_price FROM pricing")
    pricing_rows = cur.fetchall()
    cur.execute("SELECT sku, model, name FROM products")
    product_rows = cur.fetchall()
    customer_language = "pl"
    try:
        customer_id = int(order_row["customer_id"] or 0) if order_row and "customer_id" in order_row.keys() else 0
        if customer_id:
            cur.execute("SELECT language FROM customers WHERE id=? LIMIT 1", (customer_id,))
            lang_row = cur.fetchone()
            if lang_row:
                customer_language = normalize_client_language(lang_row["language"])
    except Exception:
        customer_language = "pl"
    c.close()

    pricing_map = {norm(r["model"]): r for r in pricing_rows}
    product_map = {norm(r["sku"]): r for r in product_rows}
    document_type = resolve_invoice_type(meta, items)
    document_vat_rate = to_int((items[0].get("vat_rate") if items else meta.get("vat_rate")), 23)
    document_currency = norm((items[0].get("currency") if items else meta.get("currency")) or "PLN").upper()
    invoice_language = customer_language
    if invoice_language == "pl" and document_vat_rate == 0:
        country = norm(meta.get("buyer_country")).upper()
        invoice_language = {"DE": "de", "AT": "de", "CH": "de", "ES": "es", "IT": "it"}.get(country, "en")
    PDF_COPY = {
        "pl": {"place":"Miejsce","issue":"Data wystawienia","sell":"Data sprzedaży","payment":"Forma płatności","due":"Termin płatności","seller":"Sprzedawca","buyer":"Nabywca","name":"Nazwa/SKU","qty":"Ilość","net_unit":"Netto/szt","gross_unit":"Brutto/szt","net_value":"Wartość netto","discount":"Rabat","net_total":"Suma netto","gross_total":"Suma brutto","wdt":"Wewnątrzwspólnotowa dostawa towarów (WDT) — stawka VAT 0%.","wdt_basis":"Podstawa: art. 42 ustawy o VAT; zastosowanie stawki wymaga spełnienia warunków ustawowych.","account":"konto","phone":"tel","email":"email"},
        "de": {"place":"Ort","issue":"Rechnungsdatum","sell":"Lieferdatum","payment":"Zahlungsart","due":"Zahlungsfrist","seller":"Verkäufer","buyer":"Käufer","name":"Bezeichnung/SKU","qty":"Menge","net_unit":"Netto/Stk.","gross_unit":"Brutto/Stk.","net_value":"Nettowert","discount":"Rabatt","net_total":"Nettosumme","gross_total":"Bruttosumme","wdt":"Innergemeinschaftliche Lieferung – Umsatzsteuersatz 0 %.","wdt_basis":"Rechtsgrundlage: Art. 42 des polnischen Umsatzsteuergesetzes i. V. m. Art. 138 der Richtlinie 2006/112/EG.","account":"Konto","phone":"Tel.","email":"E-Mail"},
        "en": {"place":"Place","issue":"Invoice date","sell":"Supply date","payment":"Payment method","due":"Payment due","seller":"Seller","buyer":"Buyer","name":"Description/SKU","qty":"Qty","net_unit":"Net/unit","gross_unit":"Gross/unit","net_value":"Net value","discount":"Discount","net_total":"Net total","gross_total":"Gross total","wdt":"Intra-Community supply — VAT rate 0%.","wdt_basis":"Legal basis: Article 42 of the Polish VAT Act; the 0% rate applies subject to statutory conditions.","account":"account","phone":"tel.","email":"email"},
        "es": {"place":"Lugar","issue":"Fecha de factura","sell":"Fecha de entrega","payment":"Forma de pago","due":"Vencimiento","seller":"Vendedor","buyer":"Comprador","name":"Descripción/SKU","qty":"Cantidad","net_unit":"Neto/ud.","gross_unit":"Bruto/ud.","net_value":"Valor neto","discount":"Descuento","net_total":"Total neto","gross_total":"Total bruto","wdt":"Entrega intracomunitaria — IVA 0 %.","wdt_basis":"Base legal: art. 42 de la Ley polaca del IVA; el tipo 0 % se aplica si se cumplen los requisitos legales.","account":"cuenta","phone":"tel.","email":"email"},
        "it": {"place":"Luogo","issue":"Data fattura","sell":"Data consegna","payment":"Metodo di pagamento","due":"Scadenza","seller":"Venditore","buyer":"Acquirente","name":"Descrizione/SKU","qty":"Quantità","net_unit":"Netto/pz.","gross_unit":"Lordo/pz.","net_value":"Valore netto","discount":"Sconto","net_total":"Totale netto","gross_total":"Totale lordo","wdt":"Cessione intracomunitaria — IVA 0%.","wdt_basis":"Base giuridica: art. 42 della legge polacca sull’IVA; l’aliquota 0% si applica se sono soddisfatte le condizioni di legge.","account":"conto","phone":"tel.","email":"email"},
    }
    pdf_copy = PDF_COPY.get(invoice_language, PDF_COPY["pl"])
    PAYMENT_COPY = {
        "pl": {"cash":"gotówka","gotowka":"gotówka","gotówka":"gotówka","transfer":"przelew","przelew":"przelew","card":"karta","karta":"karta"},
        "de": {"cash":"Barzahlung","gotowka":"Barzahlung","gotówka":"Barzahlung","transfer":"Überweisung","przelew":"Überweisung","card":"Karte","karta":"Karte"},
        "en": {"cash":"cash","gotowka":"cash","gotówka":"cash","transfer":"bank transfer","przelew":"bank transfer","card":"card","karta":"card"},
        "es": {"cash":"efectivo","gotowka":"efectivo","gotówka":"efectivo","transfer":"transferencia bancaria","przelew":"transferencia bancaria","card":"tarjeta","karta":"tarjeta"},
        "it": {"cash":"contanti","gotowka":"contanti","gotówka":"contanti","transfer":"bonifico bancario","przelew":"bonifico bancario","card":"carta","karta":"carta"},
    }
    payment_raw = norm(meta.get("payment_type")).lower()
    payment_label = PAYMENT_COPY.get(invoice_language, PAYMENT_COPY["pl"]).get(payment_raw, payment_raw or "-")

    def pdf_txt(value) -> str:
        return fix_polish_mojibake(norm(value))

    def fit_pdf_text(value, font_name, font_size, max_width, suffix="...") -> str:
        text = pdf_txt(value)
        if pdfmetrics.stringWidth(text, font_name, font_size) <= max_width:
            return text
        while text and pdfmetrics.stringWidth(text + suffix, font_name, font_size) > max_width:
            text = text[:-1]
        return (text + suffix) if text else ""

    def wrap_pdf_text(value, font_name, font_size, max_width, max_lines=None):
        text = pdf_txt(value)
        if not text:
            return []
        out = []
        for raw_line in str(text).replace("\r", "\n").split("\n"):
            words = raw_line.split()
            if not words:
                out.append("")
                continue
            line = ""
            for word in words:
                candidate = word if not line else f"{line} {word}"
                if pdfmetrics.stringWidth(candidate, font_name, font_size) <= max_width:
                    line = candidate
                    continue
                if line:
                    out.append(line)
                    line = word
                else:
                    out.append(fit_pdf_text(word, font_name, font_size, max_width))
                    line = ""
                if max_lines and len(out) >= max_lines:
                    out[-1] = fit_pdf_text(out[-1], font_name, font_size, max_width)
                    return out
            if line:
                out.append(line)
            if max_lines and len(out) >= max_lines:
                out = out[:max_lines]
                out[-1] = fit_pdf_text(out[-1], font_name, font_size, max_width)
                return out
        return out

    w = 210 * mm
    h = 297 * mm
    cpdf = canvas.Canvas(fpath, pagesize=(w, h))
    pdf_font, pdf_font_bold = get_pdf_font_names()

    header_y = h - 20 * mm
    if invoice_language == "de":
        document_title = ("Rechnung – Ausfuhrlieferung 0 %" if document_type == "export" else "Rechnung – innergemeinschaftliche Lieferung 0 %") if document_vat_rate == 0 else "Rechnung"
        invoice_no_label = "Rechnungsnummer"
    elif invoice_language == "en":
        document_title = ("Invoice – export of goods 0%" if document_type == "export" else "Invoice – intra-Community supply 0%") if document_vat_rate == 0 else "VAT invoice"
        invoice_no_label = "Invoice number"
    elif invoice_language == "es":
        document_title = ("Factura – exportación de bienes 0 %" if document_type == "export" else "Factura – entrega intracomunitaria 0 %") if document_vat_rate == 0 else "Factura"
        invoice_no_label = "Número de factura"
    elif invoice_language == "it":
        document_title = ("Fattura – esportazione di beni 0%" if document_type == "export" else "Fattura – cessione intracomunitaria 0%") if document_vat_rate == 0 else "Fattura"
        invoice_no_label = "Numero fattura"
    else:
        document_title = ("Faktura eksportowa 0%" if document_type == "export" else "Faktura WDT 0%") if document_vat_rate == 0 else "Faktura VAT"
        invoice_no_label = "Numer faktury"

    # Numer faktury musi być w pełni widoczny. Wcześniej tytuł i numer były
    # rysowane w jednym długim wierszu, a następnie logo przykrywało środek
    # numeru (na PDF zostawało np. samo „26” z końcówki roku 2026).
    # Niemiecki tytuł WDT jest odrobinę dłuższy niż dostępne 125 mm przy 13 pt.
    # Używamy 12.5 pt, żeby cały napis (łącznie z „%”) mieścił się bez ucinania
    # i nadal zachował bezpieczny odstęp od logo.
    header_title_size = 12.5
    cpdf.setFont(pdf_font_bold, header_title_size)
    cpdf.drawString(15 * mm, header_y, fit_pdf_text(document_title, pdf_font_bold, header_title_size, 125 * mm, suffix=""))
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawString(15 * mm, header_y - 6 * mm, f"{invoice_no_label}: {pdf_txt(meta['invoice_no'])}")

    y = h - 38 * mm
    logo = find_logo_path()
    if logo:
        try:
            logo_img = ImageReader(logo)
            img_w, img_h = logo_img.getSize()
            # Logo na fakturze było zbyt małe. Powiększenie pola dokładnie
            # o 70%, z zachowaniem proporcji obrazu.
            max_w = 102 * mm
            max_h = 40.8 * mm
            scale = min(max_w / float(img_w), max_h / float(img_h)) if img_w and img_h else 1.0
            draw_w = float(img_w) * scale
            draw_h = float(img_h) * scale
            draw_x = 195 * mm - draw_w
            draw_y = h - 10 * mm - draw_h
            cpdf.drawImage(logo_img, draw_x, draw_y, width=draw_w, height=draw_h, preserveAspectRatio=True, mask="auto")
            y = min(y, draw_y - 5 * mm)
        except Exception:
            pass

    y -= 7 * mm
    cpdf.setFont(pdf_font, 10)
    cpdf.drawString(15 * mm, y, f"{pdf_copy['place']}: {pdf_txt(meta.get('place') or '-')}")
    cpdf.drawString(85 * mm, y, f"{pdf_copy['issue']}: {pdf_txt(meta['issue_date'])}")
    cpdf.drawString(150 * mm, y, f"{pdf_copy['sell']}: {pdf_txt(meta['sell_date'])}")

    y -= 7 * mm
    is_paid = int(meta.get("paid") or 0) == 1
    paid_labels = {"pl":"Opłacona", "de":"Bezahlt", "en":"Paid", "es":"Pagada", "it":"Pagata"}
    if is_paid:
        cpdf.drawString(15 * mm, y, paid_labels.get(invoice_language, "Paid"))
    else:
        cpdf.drawString(15 * mm, y, f"{pdf_copy['payment']}: {pdf_txt(payment_label)}")
        cpdf.drawString(85 * mm, y, f"{pdf_copy['due']}: {pdf_txt(meta.get('payment_to') or '-')}")

    y -= 10 * mm
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawString(15 * mm, y, pdf_copy["seller"])
    cpdf.drawString(110 * mm, y, pdf_copy["buyer"])

    y -= 6 * mm
    cpdf.setFont(pdf_font, 9)
    seller_name = pdf_txt((company["company_name"] if company else "") or "-")
    seller_nip = pdf_txt((company["nip"] if company else "") or "-")
    seller_addr = pdf_txt((company["address"] if company else "") or "-")
    seller_phone = pdf_txt((company["phone"] if company else "") or "")
    seller_email = pdf_txt((company["email"] if company else "") or "")
    seller_bank = pdf_txt((company["bank_account"] if company else "") or "")

    buyer_name = pdf_txt(meta.get("buyer_name") or (order_row["customer_name"] if order_row and "customer_name" in order_row.keys() else "") or "-")
    buyer_tax_no = pdf_txt(meta.get("buyer_tax_no") or "-")
    buyer_street = pdf_txt(meta.get("buyer_street") or "-")
    buyer_post = pdf_txt(meta.get("buyer_post_code") or "")
    buyer_city = pdf_txt(meta.get("buyer_city") or "")
    buyer_country = pdf_txt(meta.get("buyer_country") or "PL")
    buyer_email = pdf_txt(meta.get("buyer_email") or "")
    buyer_phone = pdf_txt(meta.get("buyer_phone") or "")

    if document_vat_rate == 0:
        seller_vat_eu = re.sub(r"[\s.-]+", "", seller_nip).upper()
        if seller_vat_eu and seller_vat_eu != "-" and not seller_vat_eu.startswith("PL"):
            seller_vat_eu = f"PL{seller_vat_eu}"
        seller_tax_line = f"VAT UE: {seller_vat_eu}"
        buyer_tax_line = f"VAT UE: {buyer_tax_no}"
    else:
        seller_tax_line = f"NIP: {seller_nip}"
        buyer_tax_line = f"NIP: {buyer_tax_no}"

    seller_lines = [seller_name, seller_tax_line, seller_addr]
    if seller_phone:
        seller_lines.append(f"{pdf_copy['phone']}: {seller_phone}")
    if seller_email:
        seller_lines.append(f"{pdf_copy['email']}: {seller_email}")
    if seller_bank and not is_paid:
        seller_lines.append(f"{pdf_copy['account']}: {seller_bank}")

    buyer_lines = [buyer_name, buyer_tax_line, buyer_street, f"{buyer_post} {buyer_city}".strip(), buyer_country]
    if buyer_phone:
        buyer_lines.append(f"{pdf_copy['phone']}: {buyer_phone}")
    if buyer_email:
        buyer_lines.append(f"{pdf_copy['email']}: {buyer_email}")

    seller_x = 15 * mm
    buyer_x = 108 * mm
    seller_width = 84 * mm
    buyer_width = 87 * mm
    line_gap = 4.8 * mm
    seller_wrapped = []
    buyer_wrapped = []
    for line in seller_lines:
        seller_wrapped.extend(wrap_pdf_text(line, pdf_font, 8.7, seller_width, max_lines=2))
    for line in buyer_lines:
        buyer_wrapped.extend(wrap_pdf_text(line, pdf_font, 8.7, buyer_width, max_lines=2))

    max_len = max(len(seller_wrapped), len(buyer_wrapped))
    cpdf.setFont(pdf_font, 8.7)
    for i in range(max_len):
        if i < len(seller_wrapped):
            cpdf.drawString(seller_x, y, seller_wrapped[i])
        if i < len(buyer_wrapped):
            cpdf.drawString(buyer_x, y, buyer_wrapped[i])
        y -= line_gap

    y -= 4 * mm
    table_left = 12 * mm
    table_right = 198 * mm
    row_h = 12 * mm
    # L.p. | Nazwa/SKU | Ilo?? | Netto/szt | Brutto/szt | Wart. netto | VAT
    col_x = [12 * mm, 20 * mm, 100 * mm, 113 * mm, 136 * mm, 159 * mm, 182 * mm, 198 * mm]

    def cell_center(x1, x2):
        return (x1 + x2) / 2.0

    def cell_baseline(y_top, h_cell, font_name, font_size):
        asc = pdfmetrics.getAscent(font_name, font_size)
        desc = pdfmetrics.getDescent(font_name, font_size)
        text_h = asc - desc
        y_bottom = y_top - h_cell + 1
        return y_bottom + (h_cell - text_h) / 2.0 - desc

    cpdf.setFillColorRGB(0.96, 0.96, 0.96)
    cpdf.rect(table_left, y - row_h + 1, table_right - table_left, row_h, stroke=0, fill=1)
    cpdf.setFillColorRGB(0, 0, 0)
    header_font = 7.6
    cpdf.setFont(pdf_font_bold, header_font)
    header_y = cell_baseline(y, row_h, pdf_font_bold, header_font)
    cpdf.drawCentredString(cell_center(col_x[0], col_x[1]), header_y, "L.p.")
    cpdf.drawCentredString(cell_center(col_x[1], col_x[2]), header_y, pdf_copy["name"])
    cpdf.drawCentredString(cell_center(col_x[2], col_x[3]), header_y, pdf_copy["qty"])
    cpdf.drawCentredString(cell_center(col_x[3], col_x[4]), header_y, pdf_copy["net_unit"])
    cpdf.drawCentredString(cell_center(col_x[4], col_x[5]), header_y, pdf_copy["gross_unit"])
    cpdf.drawCentredString(cell_center(col_x[5], col_x[6]), header_y, pdf_copy["net_value"])
    cpdf.drawCentredString(cell_center(col_x[6], col_x[7]), header_y, "VAT")
    cpdf.line(table_left, y + 1, table_right, y + 1)
    cpdf.line(table_left, y - row_h + 1, table_right, y - row_h + 1)
    for cx in col_x:
        cpdf.line(cx, y + 1, cx, y - row_h + 1)
    y -= row_h

    total_net = 0.0
    total_net_dec = Decimal("0.00")
    discount_pct = max(0.0, to_float(meta.get("discount_percent"), 0.0))
    body_font = 8.2
    cpdf.setFont(pdf_font, body_font)

    lp = 1
    for it in items:
        sku = pdf_txt(it.get("sku"))
        product_row = product_map.get(norm(sku))
        model = pdf_txt(it.get("model") or (product_row["model"] if product_row else ""))
        name = pdf_txt(it.get("name") or (product_row["name"] if product_row else ""))
        common_name = name or model
        pr = pricing_map.get(model) or pricing_map.get(sku)
        # Faktura musi zachowac cene zapisana w chwili jej utworzenia.
        # Aktualny cennik jest tylko awaryjnym zrodlem dla bardzo starych
        # pozycji, w ktorych ceny jeszcze nie byly utrwalane w JSON-ie.
        saved_net_price = it.get("net_price")
        if saved_net_price is None or norm(saved_net_price) == "":
            saved_net_price = pr["net_price"] if pr else 0
        net_dec = money_dec(saved_net_price)
        qty = int(it["qty"])
        line_net_dec = (net_dec * Decimal(qty)).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        if discount_pct > 0:
            line_net_dec = (line_net_dec * (Decimal("100.0") - Decimal(str(discount_pct))) / Decimal("100.0")).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        vat_rate = to_int(it.get("vat_rate"), to_int(meta.get("vat_rate"), 23))
        unit_gross_dec = net_dec if vat_rate == 0 else gross_from_net_23(net_dec)

        net = money_float(net_dec)
        gross = money_float(unit_gross_dec)
        line_net = money_float(line_net_dec)

        total_net += line_net
        total_net_dec += line_net_dec

        text_y = cell_baseline(y, row_h, pdf_font, body_font)
        cpdf.drawCentredString(cell_center(col_x[0], col_x[1]), text_y, str(lp))
        name_left = col_x[1] + 1.5 * mm
        name_width = (col_x[2] - col_x[1]) - 3 * mm
        cpdf.setFont(pdf_font_bold, body_font)
        cpdf.drawString(name_left, y - 4.4 * mm, fit_pdf_text(sku or "-", pdf_font_bold, body_font, name_width))
        cpdf.setFont(pdf_font, body_font)
        if common_name:
            label = common_name if common_name.lower() == model.lower() else f"{common_name} / {model}".strip(" /")
            cpdf.drawString(name_left, y - 8.7 * mm, fit_pdf_text(label, pdf_font, body_font, name_width))
        cpdf.drawCentredString(cell_center(col_x[2], col_x[3]), text_y, str(qty))
        cpdf.drawRightString(col_x[4] - 1.5 * mm, text_y, f"{net:.2f}")
        cpdf.drawRightString(col_x[5] - 1.5 * mm, text_y, f"{gross:.2f}")
        cpdf.drawRightString(col_x[6] - 1.5 * mm, text_y, f"{line_net:.2f}")
        cpdf.drawCentredString(cell_center(col_x[6], col_x[7]), text_y, f"{vat_rate}%")
        cpdf.line(table_left, y - row_h + 1, table_right, y - row_h + 1)
        for cx in col_x:
            cpdf.line(cx, y + 1, cx, y - row_h + 1)
        y -= row_h
        lp += 1
        if y < 26 * mm:
            cpdf.showPage()
            y = h - 20 * mm
            cpdf.setFont(pdf_font, body_font)

    total_net_dec = total_net_dec.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    total_tax_dec = Decimal("0.00") if document_vat_rate == 0 else vat23_from_net(total_net_dec)
    total_gross_dec = (total_net_dec + total_tax_dec).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    total_net = money_float(total_net_dec)
    total_tax = money_float(total_tax_dec)
    total_gross = money_float(total_gross_dec)
    y -= 6 * mm
    cpdf.setFont(pdf_font_bold, 10)
    if discount_pct > 0:
        cpdf.drawRightString(198 * mm, y, f"{pdf_copy['discount']}: {discount_pct:.2f}%")
        y -= 5 * mm
    cpdf.drawRightString(198 * mm, y, f"{pdf_copy['net_total']}: {total_net:.2f} {document_currency}")
    y -= 5 * mm
    cpdf.drawRightString(198 * mm, y, f"VAT {document_vat_rate}%: {total_tax:.2f} {document_currency}")
    y -= 5 * mm
    cpdf.drawRightString(198 * mm, y, f"{pdf_copy['gross_total']}: {total_gross:.2f} {document_currency}")
    if document_vat_rate == 0:
        y -= 9 * mm
        cpdf.setFont(pdf_font, 8.5)
        legal_width = 180 * mm
        export_copy = {
            "pl":"Eksport towarów — stawka VAT 0%.",
            "de":"Ausfuhrlieferung — Umsatzsteuersatz 0 %.",
            "en":"Export of goods — VAT rate 0%.",
            "es":"Exportación de bienes — IVA 0 %.",
            "it":"Esportazione di beni — IVA 0%.",
        }
        export_basis = {
            "pl":"Podstawa: art. 41 ust. 4–11 ustawy o VAT; stawka 0% wymaga dokumentów potwierdzających wywóz.",
            "de":"Rechtsgrundlage: Art. 41 Abs. 4–11 des polnischen Umsatzsteuergesetzes; der Ausfuhrnachweis ist erforderlich.",
            "en":"Legal basis: Article 41(4–11) of the Polish VAT Act; proof of export is required.",
            "es":"Base legal: art. 41, apdos. 4–11, de la Ley polaca del IVA; se exige prueba de exportación.",
            "it":"Base giuridica: art. 41, commi 4–11, della legge polacca sull’IVA; è richiesta la prova dell’esportazione.",
        }
        legal_title = export_copy.get(invoice_language, export_copy["en"]) if document_type == "export" else pdf_copy["wdt"]
        legal_basis = export_basis.get(invoice_language, export_basis["en"]) if document_type == "export" else pdf_copy["wdt_basis"]
        for legal_line in wrap_pdf_text(legal_title, pdf_font, 8.5, legal_width):
            cpdf.drawString(15 * mm, y, legal_line)
            y -= 4.5 * mm
        for legal_line in wrap_pdf_text(legal_basis, pdf_font, 8.5, legal_width):
            cpdf.drawString(15 * mm, y, legal_line)
            y -= 4.5 * mm

    ksef_number = norm(meta.get("ksef_number") or "")
    if ksef_number:
        y -= 10 * mm
        if y < 22 * mm:
            cpdf.showPage()
            y = h - 20 * mm
        cpdf.setFont(pdf_font_bold, 9)
        cpdf.drawString(15 * mm, y, "KSeF")
        y -= 5 * mm
        cpdf.setFont(pdf_font, 8.5)
        cpdf.drawString(15 * mm, y, "Faktura została wystawiona i jest dostępna w Krajowym Systemie e-Faktur.")
        y -= 5 * mm
        cpdf.setFont(pdf_font_bold, 8.5)
        cpdf.drawString(15 * mm, y, f"Numer KSeF: {pdf_txt(ksef_number)}")

    cpdf.save()
    return fpath, round(total_net,2), round(total_gross,2)


def generate_order_invoice_pdf(order_row, items, meta):
    """Pure PDF dispatcher; it performs no stock/status/allocation writes."""
    invoice_type = resolve_invoice_type(meta, items)
    payload = dict(meta)
    payload["invoice_type"] = invoice_type
    if invoice_type == "domestic":
        return invoice_domestic.generate(
            order_row, items, payload, renderer=_legacy_generate_order_invoice_pdf
        )
    if invoice_type in {"wdt", "export"}:
        return invoice_foreign.generate(
            order_row, items, payload, renderer=_legacy_generate_order_invoice_pdf
        )
    raise ValueError("Nieobsługiwany typ faktury")


def packing_list_pdf_path_for_invoice(invoice_pdf_path: str, invoice_no: str) -> str:
    if parse_supabase_storage_ref(invoice_pdf_path):
        invoice_pdf_path = ""
    base_dir = os.path.dirname(invoice_pdf_path) if invoice_pdf_path else os.path.join(DATA_DIR, "faktury")
    return os.path.join(base_dir, f"{safe_filename(invoice_no)}_lista_pakowania.pdf")


def invoice_storage_object_path(invoice_id: int, invoice_no: str, suffix: str = ".pdf") -> str:
    return f"invoices/{int(invoice_id)}/{safe_filename(invoice_no)}{suffix}"


def invoice_packing_storage_object_path(invoice_id: int, invoice_no: str) -> str:
    return invoice_storage_object_path(invoice_id, invoice_no, "_lista_pakowania.pdf")


def upload_invoice_pdfs_to_supabase(invoice_id: int, invoice_no: str, invoice_pdf_path: str, packing_pdf_path: str = "") -> str:
    if not supabase_enabled():
        return invoice_pdf_relpath(invoice_pdf_path)
    invoice_ref = supabase_storage_upload_file(
        invoice_pdf_path,
        invoice_storage_object_path(invoice_id, invoice_no),
        content_type="application/pdf",
    )
    if packing_pdf_path and os.path.exists(packing_pdf_path):
        try:
            supabase_storage_upload_file(
                packing_pdf_path,
                invoice_packing_storage_object_path(invoice_id, invoice_no),
                content_type="application/pdf",
            )
        except Exception:
            pass
    return invoice_ref


def generate_invoice_packing_list_pdf(order_row, items, meta, invoice_pdf_path: str = "") -> str:
    customer_dir = invoice_dir_for_customer(meta.get("buyer_name") or (order_row["customer_name"] if order_row and "customer_name" in order_row.keys() else "") or "Klient")
    fpath = packing_list_pdf_path_for_invoice(invoice_pdf_path or os.path.join(customer_dir, f"{safe_filename(meta['invoice_no'])}.pdf"), meta["invoice_no"])
    w, h = 210 * mm, 297 * mm
    cpdf = canvas.Canvas(fpath, pagesize=(w, h))
    pdf_font, pdf_font_bold = get_pdf_font_names()
    # Lista pakowania jest dokumentem roboczym. Kolory są celowo dużo
    # ciemniejsze niż w ekranowych PDF-ach, aby pozostały czytelne także na
    # kolorowych drukarkach laserowych i przy druku ekonomicznym.
    navy, muted, blue = (0.0, 0.0, 0.0), (0.12, 0.12, 0.12), (0.0, 0.0, 0.0)
    pale_blue, line_color = (0.91, 0.91, 0.91), (0.38, 0.38, 0.38)

    def order_value(key, default=""):
        if not order_row:
            return default
        try:
            return order_row[key] if key in order_row.keys() else default
        except (AttributeError, KeyError, TypeError):
            return order_row.get(key, default) if isinstance(order_row, dict) else default

    def fit_text(text, max_width, font_name, font_size):
        text = norm(text or "") or "-"
        if cpdf.stringWidth(text, font_name, font_size) <= max_width:
            return text
        while text and cpdf.stringWidth(text + "...", font_name, font_size) > max_width:
            text = text[:-1]
        return (text + "...") if text else "..."

    def fit_font_size(text, max_width, font_name, preferred_size, minimum_size=6.0):
        """Zmniejsza font bez obcinania tekstu (uzywane dla numeru zamowienia)."""
        text = norm(text or "") or "-"
        size = float(preferred_size)
        while size > minimum_size and cpdf.stringWidth(text, font_name, size) > max_width:
            size -= 0.25
        return max(size, minimum_size)

    def strip_note_from_order_no(order_no, note):
        order_no, note = norm(order_no or ""), norm(note or "")
        if note and order_no.lower().endswith((" " + note).lower()):
            return order_no[:-(len(note) + 1)].strip()
        return order_no

    packing_order_numbers = []
    for item in items:
        source_no = strip_note_from_order_no(
            item.get("source_order_no"),
            item.get("source_order_note"),
        )
        if source_no and source_no not in packing_order_numbers:
            packing_order_numbers.append(source_no)

    language = normalize_client_language(meta.get("language") or order_value("language"))
    if not (meta.get("language") or order_value("language")):
        customer_email = norm(order_value("customer_email") or meta.get("buyer_email"))
        if customer_email:
            try:
                language = normalize_client_language(_client_profile_for_email(customer_email).get("language"))
            except Exception as exc:
                app.logger.warning("Nie udało się ustalić języka listy pakowania dla %s: %s", customer_email, type(exc).__name__)
    tr = lambda key: packing_list_text(language, key)

    def draw_header(continuation=False):
        logo_path = find_logo_path()
        if logo_path:
            try:
                cpdf.drawImage(ImageReader(logo_path), 15 * mm, h - 42 * mm, 54.4 * mm, 34 * mm,
                               preserveAspectRatio=True, anchor="w", mask="auto")
            except Exception:
                pass
        cpdf.setFillColorRGB(*navy)
        cpdf.setFont(pdf_font_bold, 18)
        cpdf.drawRightString(195 * mm, h - 15 * mm, tr("title"))
        cpdf.setFillColorRGB(*muted)
        cpdf.setFont(pdf_font_bold, 9)
        document_label_key = norm(meta.get("document_label_key")) or "invoice"
        subtitle_value = norm(meta.get("invoice_no") or "-")
        if document_label_key == "order" and packing_order_numbers:
            subtitle_value = ", ".join(packing_order_numbers)
        subtitle = f"{tr(document_label_key)}: {subtitle_value}"
        subtitle_size = fit_font_size(subtitle, 105 * mm, pdf_font_bold, 9, minimum_size=5.5)
        cpdf.setFont(pdf_font_bold, subtitle_size)
        if continuation:
            subtitle += f"  |  {tr('continued')}"
        cpdf.drawRightString(195 * mm, h - 23 * mm, subtitle)
        cpdf.setStrokeColorRGB(*line_color)
        cpdf.setLineWidth(0.8)
        cpdf.line(15 * mm, h - 45 * mm, 195 * mm, h - 45 * mm)
        return h - 55 * mm

    def draw_table_header(current_y):
        cpdf.setFillColorRGB(*pale_blue)
        cpdf.setStrokeColorRGB(*line_color)
        cpdf.setLineWidth(0.8)
        cpdf.roundRect(15 * mm, current_y - 7 * mm, 180 * mm, 10 * mm, 2 * mm, fill=1, stroke=1)
        cpdf.setFillColorRGB(*blue)
        cpdf.setFont(pdf_font_bold, 9)
        for label, x_mm in ((tr("checked"), 19), (tr("line"), 29), ("SKU", 42), (tr("product"), 91), (tr("source"), 132)):
            cpdf.drawString(x_mm * mm, current_y - 3.5 * mm, label)
        cpdf.drawRightString(190 * mm, current_y - 3.5 * mm, tr("quantity"))
        return current_y - 13 * mm

    y = draw_header()
    customer_name = norm(meta.get("buyer_name") or order_value("customer_name") or "-")
    main_order_no = ", ".join(packing_order_numbers) or norm(order_value("order_no") or order_value("number") or "-")
    # Data w naglowku pochodzi z zamowienia. Osobna data wydruku jest
    # umieszczana na dole przy polu osoby pakujacej.
    order_created_at = norm(order_value("created_at") or "")
    order_date = order_created_at[:10] if len(order_created_at) >= 10 else app_now().strftime("%Y-%m-%d")
    print_date = app_now().strftime("%Y-%m-%d")
    cpdf.setFillColorRGB(*muted)
    cpdf.setFont(pdf_font_bold, 9)
    cpdf.drawString(15 * mm, y, tr("order"))
    cpdf.drawString(72 * mm, y, tr("date"))
    cpdf.drawString(122 * mm, y, tr("customer"))
    y -= 6 * mm
    cpdf.setFillColorRGB(*navy)
    cpdf.setFont(pdf_font_bold, 10.5)
    main_order_font_size = fit_font_size(main_order_no, 50 * mm, pdf_font_bold, 10.5, minimum_size=5.5)
    cpdf.setFont(pdf_font_bold, main_order_font_size)
    cpdf.drawString(15 * mm, y, main_order_no)
    cpdf.setFont(pdf_font_bold, 10.5)
    cpdf.drawString(72 * mm, y, order_date)
    cpdf.drawString(122 * mm, y, fit_text(customer_name, 73 * mm, pdf_font, 10))
    y = draw_table_header(y - 11 * mm)

    total_qty = 0
    item_count = 0
    for it in items:
        qty = int(it.get("qty") or 0)
        if qty <= 0:
            continue
        item_count += 1
        total_qty += qty
        source_order_no = norm(it.get("source_order_no") or "")
        note = norm(it.get("source_order_note") or "")
        source_order_no = strip_note_from_order_no(source_order_no, note)
        source_text = " · ".join(x for x in (source_order_no, note) if x) or "-"
        sku = norm(it.get("sku") or "")
        model_name = norm(it.get("model") or it.get("name") or "")
        if y < 36 * mm:
            cpdf.showPage()
            y = draw_table_header(draw_header(continuation=True))
        cpdf.setStrokeColorRGB(0.05, 0.05, 0.05)
        cpdf.setLineWidth(1.2)
        cpdf.roundRect(18 * mm, y - 2.5 * mm, 5 * mm, 5 * mm, 1 * mm, fill=0, stroke=1)
        cpdf.setFillColorRGB(*navy)
        cpdf.setFont(pdf_font_bold, 9.5)
        cpdf.drawString(29 * mm, y, str(item_count))
        cpdf.setFont(pdf_font_bold, 10)
        cpdf.drawString(42 * mm, y, fit_text(sku, 44 * mm, pdf_font_bold, 10))
        cpdf.setFont(pdf_font_bold, 9.5)
        cpdf.drawString(91 * mm, y, fit_text(model_name, 36 * mm, pdf_font_bold, 9.5))
        cpdf.setFillColorRGB(*muted)
        # Numer zamowienia jest identyfikatorem roboczym i nigdy nie moze byc
        # zakonczony wielokropkiem. W razie dluzszego numeru zmniejszamy font,
        # zachowujac cala wartosc wraz z dopiskiem/notatka.
        source_font_size = fit_font_size(source_text, 43 * mm, pdf_font_bold, 8.5)
        cpdf.setFont(pdf_font_bold, source_font_size)
        cpdf.drawString(132 * mm, y, source_text)
        cpdf.setFillColorRGB(*navy)
        cpdf.setFont(pdf_font_bold, 12)
        cpdf.drawRightString(190 * mm, y, str(qty))
        cpdf.setStrokeColorRGB(*line_color)
        cpdf.setLineWidth(0.7)
        cpdf.line(15 * mm, y - 5.5 * mm, 195 * mm, y - 5.5 * mm)
        y -= 11 * mm

    if y < 50 * mm:
        cpdf.showPage()
        y = draw_header(continuation=True)
    y -= 2 * mm
    cpdf.setFillColorRGB(*pale_blue)
    cpdf.setStrokeColorRGB(*line_color)
    cpdf.setLineWidth(0.8)
    cpdf.roundRect(15 * mm, y - 13 * mm, 180 * mm, 18 * mm, 3 * mm, fill=1, stroke=1)
    cpdf.setFillColorRGB(*navy)
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawString(21 * mm, y - 5 * mm, f"{tr('positions')}: {item_count}")
    cpdf.drawString(70 * mm, y - 5 * mm, f"{tr('total_qty')}: {total_qty}")
    cpdf.drawString(128 * mm, y - 5 * mm, f"{tr('packages')}:  ______")
    y -= 28 * mm
    cpdf.setFillColorRGB(*muted)
    cpdf.setFont(pdf_font_bold, 9)
    cpdf.drawString(15 * mm, y, f"{tr('packed_by')}:")
    cpdf.drawString(91 * mm, y, f"{tr('date').title()}:")
    cpdf.drawString(104 * mm, y, print_date)
    cpdf.drawString(145 * mm, y, f"{tr('signature')}:")
    cpdf.setStrokeColorRGB(*line_color)
    cpdf.setLineWidth(0.8)
    cpdf.line(37 * mm, y - 1 * mm, 82 * mm, y - 1 * mm)
    cpdf.line(103 * mm, y - 1 * mm, 134 * mm, y - 1 * mm)
    cpdf.line(159 * mm, y - 1 * mm, 195 * mm, y - 1 * mm)
    cpdf.save()
    return fpath


def invoice_pdf_relpath(abs_path: str) -> str:
    try:
        return os.path.relpath(abs_path, DATA_DIR)
    except Exception:
        return abs_path

def invoice_pdf_abspath(rel_path: str) -> str:
    return os.path.join(DATA_DIR, rel_path)

def find_invoice_pdf_fallback(invoice_no: str) -> str:
    root = os.path.join(DATA_DIR, "faktury")
    target = f"{safe_filename(invoice_no or '')}.pdf"
    if not target or target == ".pdf" or not os.path.isdir(root):
        return ""
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            if fn == target:
                return os.path.join(dirpath, fn)
    return ""

def invoice_pdf_exists(pdf_path: str, invoice_no: str = "") -> tuple[bool, str]:
    abs_path = ""
    raw_pdf = norm(pdf_path)
    if parse_supabase_storage_ref(raw_pdf):
        try:
            supabase_storage_download_bytes(raw_pdf)
            return True, raw_pdf
        except Exception:
            return False, raw_pdf
    if raw_pdf:
        abs_path = raw_pdf if os.path.isabs(raw_pdf) else invoice_pdf_abspath(raw_pdf)
    if abs_path and os.path.exists(abs_path):
        return True, abs_path
    fallback = find_invoice_pdf_fallback(invoice_no)
    if fallback and os.path.exists(fallback):
        return True, fallback
    return False, ""


def load_invoice_meta(invoice_id: int):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM invoice_meta WHERE invoice_id=?", (invoice_id,))
    row = cur.fetchone()
    c.close()
    return dict(row) if row else None

def upsert_invoice_meta(
    invoice_id: int,
    pdf_path: str = "",
    invoice_items_json: str = "",
    sent_to_client: int | None = None,
    seen_by_client: int | None = None,
    seen_at: str | None = None,
    payment_reminder: int | None = None,
    paid: int | None = None,
    paid_at: str | None = None
):
    current = load_invoice_meta(invoice_id) or {}
    if sent_to_client is None:
        sent_to_client = int(current.get("sent_to_client") or 0)
    if seen_by_client is None:
        seen_by_client = int(current.get("seen_by_client") or 0)
    if seen_at is None:
        seen_at = current.get("seen_at")
    if payment_reminder is None:
        payment_reminder = int(current.get("payment_reminder") or 0)
    if paid is None:
        paid = int(current.get("paid") or 0)
    if paid_at is None:
        paid_at = current.get("paid_at")

    c = conn()
    cur = c.cursor()
    cur.execute("""
      INSERT INTO invoice_meta(invoice_id, pdf_path, invoice_items_json, sent_to_client, seen_by_client, payment_reminder, paid, paid_at, seen_at, updated_at)
      VALUES(?,?,?,?,?,?,?,?,?,?)
      ON CONFLICT(invoice_id) DO UPDATE SET
        pdf_path=excluded.pdf_path,
        invoice_items_json=excluded.invoice_items_json,
        sent_to_client=excluded.sent_to_client,
        seen_by_client=excluded.seen_by_client,
        payment_reminder=excluded.payment_reminder,
        paid=excluded.paid,
        paid_at=excluded.paid_at,
        seen_at=excluded.seen_at,
        updated_at=excluded.updated_at
    """, (invoice_id, pdf_path, invoice_items_json, int(sent_to_client), int(seen_by_client), int(payment_reminder), int(paid), paid_at, seen_at, now_iso()))
    c.commit()
    c.close()


def sync_invoice_meta_to_supabase(invoice_id: int):
    if not supabase_enabled():
        return
    meta = load_invoice_meta(invoice_id)
    if not meta:
        return
    try:
        sync_local_rows_to_supabase("invoice_meta", "invoice_id", [invoice_id])
        return
    except Exception:
        pass

    # Fallback dla Supabase bez najnowszych kolumn payment_reminder/paid/paid_at.
    legacy = {
        "invoice_id": meta.get("invoice_id"),
        "pdf_path": meta.get("pdf_path") or "",
        "invoice_items_json": meta.get("invoice_items_json") or "",
        "sent_to_client": int(meta.get("sent_to_client") or 0),
        "seen_by_client": int(meta.get("seen_by_client") or 0),
        "seen_at": meta.get("seen_at"),
        "updated_at": meta.get("updated_at") or now_iso(),
    }
    supabase_upsert_rows("invoice_meta", [legacy], "invoice_id")

def prepare_invoice_items(order_items: list[dict], form):
    prepared = []
    for it in order_items:
        remaining_qty = int(it.get("remaining_qty") if it.get("remaining_qty") is not None else it.get("qty") or 0)
        qty = to_int(form.get(f"invoice_qty_{it['id']}"), 0)
        if qty <= 0:
            continue
        qty = min(qty, remaining_qty)
        if qty <= 0:
            continue
        row = dict(it)
        row["order_item_id"] = int(it.get("id") or 0)
        row["source_order_id"] = int(it.get("order_id") or it.get("source_order_id") or 0)
        row["source_order_no"] = it.get("source_order_no") or ""
        row["source_order_note"] = it.get("source_order_note") or ""
        row["ordered_qty"] = int(it.get("qty") or 0)
        row["invoiced_qty_before"] = int(it.get("invoiced_qty") or 0)
        row["qty"] = qty
        line_net = money_dec(row.get("net_price")) * Decimal(qty)
        line_net = line_net.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        line_vat = vat23_from_net(line_net)
        line_gross = (line_net + line_vat).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        row["gross_price"] = money_float(gross_from_net_23(row.get("net_price")))
        row["vat_rate"] = 23
        row["line_value_net"] = money_float(line_net)
        row["line_value_vat"] = money_float(line_vat)
        row["line_value_gross"] = money_float(line_gross)
        prepared.append(row)
    return prepared


def invoiced_qty_by_order_item_ids(order_item_ids: list[int]):
    ids = [int(x) for x in order_item_ids if x is not None]
    out = {x: 0 for x in ids}
    if not ids:
        return out

    c = conn()
    cur = c.cursor()
    ph = ",".join(["?"] * len(ids))
    cur.execute(f"""
      SELECT order_item_id, COALESCE(SUM(qty),0) AS qty
      FROM invoice_allocations
      WHERE order_item_id IN ({ph})
      GROUP BY order_item_id
    """, tuple(ids))
    for r in cur.fetchall():
        out[int(r["order_item_id"])] = int(r["qty"] or 0)
    c.close()
    return out


def replace_invoice_allocations(invoice_id: int, invoice_items: list[dict]):
    c = conn()
    cur = c.cursor()
    cur.execute("DELETE FROM invoice_allocations WHERE invoice_id=?", (invoice_id,))
    allocation_ids = []
    for it in invoice_items:
        order_item_id = int(it.get("order_item_id") or it.get("id") or 0)
        order_id = int(it.get("source_order_id") or it.get("order_id") or 0)
        qty = int(it.get("qty") or 0)
        if order_item_id <= 0 or order_id <= 0 or qty <= 0:
            continue
        cur.execute("""
          INSERT INTO invoice_allocations(invoice_id, order_id, order_item_id, product_id, sku, qty, created_at)
          VALUES(?,?,?,?,?,?,?)
        """, (
            invoice_id,
            order_id,
            order_item_id,
            int(it.get("product_id") or 0) or None,
            it.get("sku") or "",
            qty,
            now_iso()
        ))
        allocation_ids.append(int(cur.lastrowid))
    c.commit()
    c.close()
    return allocation_ids


def order_fully_invoiced(cur, order_id: int) -> bool:
    cur.execute("SELECT id, qty FROM order_items WHERE order_id=?", (order_id,))
    rows = [dict(r) for r in cur.fetchall()]
    if not rows:
        return False
    item_ids = [int(r["id"]) for r in rows]
    ph = ",".join(["?"] * len(item_ids))
    cur.execute(f"""
      SELECT order_item_id, COALESCE(SUM(qty),0) AS qty
      FROM invoice_allocations
      WHERE order_item_id IN ({ph})
      GROUP BY order_item_id
    """, tuple(item_ids))
    done = {int(r["order_item_id"]): int(r["qty"] or 0) for r in cur.fetchall()}
    return all(int(row["qty"] or 0) > 0 and int(done.get(int(row["id"]), 0)) >= int(row["qty"] or 0) for row in rows)


def save_packing_selection(root_order_id: int, packing_items: list[dict]) -> int:
    """Persist one packing choice until it is consumed by an invoice."""
    rows = []
    for item in packing_items or []:
        order_id = to_int(item.get("source_order_id") or item.get("order_id"), 0)
        item_id = to_int(item.get("order_item_id") or item.get("id"), 0)
        qty = max(0, to_int(item.get("qty"), 0))
        if order_id > 0 and item_id > 0 and qty > 0:
            rows.append((order_id, item_id, qty))
    if not rows:
        return 0
    c = conn()
    cur = c.cursor()
    cur.execute(
        "INSERT INTO packing_batches(root_order_id, invoice_id, created_at) VALUES(?,NULL,?)",
        (int(root_order_id), now_iso()),
    )
    batch_id = int(cur.lastrowid)
    cur.executemany(
        """INSERT INTO packing_allocations(batch_id, order_id, order_item_id, qty, created_at)
           VALUES(?,?,?,?,?)""",
        [(batch_id, order_id, item_id, qty, now_iso()) for order_id, item_id, qty in rows],
    )
    c.commit()
    c.close()
    return batch_id


def load_open_packing_selection(root_order_id: int) -> dict:
    c = conn()
    cur = c.cursor()
    cur.execute(
        """SELECT id FROM packing_batches
           WHERE root_order_id=? AND invoice_id IS NULL
           ORDER BY id DESC LIMIT 1""",
        (int(root_order_id),),
    )
    batch = cur.fetchone()
    if not batch:
        c.close()
        return {}
    batch_id = int(batch["id"])
    cur.execute(
        "SELECT order_id, order_item_id, qty FROM packing_allocations WHERE batch_id=? ORDER BY id",
        (batch_id,),
    )
    rows = [dict(row) for row in cur.fetchall()]
    c.close()
    return {
        "batch_id": batch_id,
        "root_order_id": int(root_order_id),
        "order_ids": sorted({int(row["order_id"]) for row in rows}),
        "items": [[int(row["order_item_id"]), int(row["qty"])] for row in rows],
    }


def consume_packing_selection(batch_id: int, invoice_id: int):
    if batch_id <= 0 or invoice_id <= 0:
        return
    c = conn()
    c.execute(
        "UPDATE packing_batches SET invoice_id=? WHERE id=? AND invoice_id IS NULL",
        (int(invoice_id), int(batch_id)),
    )
    c.commit()
    c.close()


def issue_order_stock(cur, order_id: int) -> list[int]:
    """Odejmij pełne zamówienie dokładnie raz w bieżącej transakcji."""
    cur.execute(
        "UPDATE orders SET warehouse_issued=1 WHERE id=? AND COALESCE(warehouse_issued,0)=0",
        (order_id,),
    )
    if cur.rowcount != 1:
        return []

    cur.execute("SELECT product_id, qty FROM order_items WHERE order_id=? ORDER BY id", (order_id,))
    changed_product_ids = []
    for item in cur.fetchall():
        product_id = int(item["product_id"])
        qty = int(item["qty"] or 0)
        if qty <= 0:
            continue
        cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (product_id,))
        cur.execute("UPDATE stock SET qty=qty-? WHERE product_id=?", (qty, product_id))
        changed_product_ids.append(product_id)

    return changed_product_ids


def finalize_fully_invoiced_orders(order_ids: list[int]):
    touched = sorted({int(x) for x in order_ids if x})
    if not touched:
        return [], []

    c = conn()
    cur = c.cursor()
    changed_order_ids = []
    changed_product_ids = []

    for order_id in touched:
        if not order_fully_invoiced(cur, order_id):
            continue
        cur.execute("SELECT id, status, warehouse_issued FROM orders WHERE id=?", (order_id,))
        order_row = cur.fetchone()
        if not order_row:
            continue

        warehouse_issued = int(order_row["warehouse_issued"] or 0)
        if warehouse_issued == 0:
            changed_product_ids.extend(issue_order_stock(cur, order_id))
            warehouse_issued = 1

        current_status = norm(order_row["status"]).lower()
        # Wystawienie faktury oznacza realizację, a nie zakończenie zamówienia.
        # Status wysyłki zachowujemy; „Zrealizowane” ustawia dopiero opłacenie
        # wszystkich faktur przypisanych do zamówienia.
        preserved = {"shipped", "partially_shipped", "completed", "issued", "cancelled"}
        next_status = current_status if current_status in preserved else "packed"
        if current_status != next_status or int(order_row["warehouse_issued"] or 0) != warehouse_issued:
            if next_status == "packed":
                cur.execute("""
                  UPDATE orders
                  SET status=?, warehouse_issued=?, packed_at=COALESCE(packed_at, ?)
                  WHERE id=?
                """, (next_status, warehouse_issued, now_iso(), order_id))
            else:
                cur.execute("UPDATE orders SET status=?, warehouse_issued=? WHERE id=?", (next_status, warehouse_issued, order_id))
            changed_order_ids.append(order_id)

    c.commit()
    c.close()

    if supabase_enabled():
        if changed_order_ids:
            try:
                sync_local_rows_to_supabase("orders", "id", changed_order_ids)
            except Exception:
                pass
        if changed_product_ids:
            try:
                sync_local_rows_to_supabase("stock", "product_id", list(set(changed_product_ids)))
            except Exception:
                pass

    return changed_order_ids, list(set(changed_product_ids))


def reconcile_legacy_shipped_order_statuses():
    """Uzupełnia precyzyjny status starszych wysyłek bez zmiany dokumentów.

    Dawny kod zapisywał ``shipped`` również wtedy, gdy faktura/wysyłka
    obejmowała tylko część pozycji. Źródłem prawdy o zrealizowanych ilościach
    są istniejące ``invoice_allocations``. Rekord bez żadnej alokacji zostaje
    nietknięty, bo nie da się bezpiecznie ustalić, czy był wysłany częściowo.
    """
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT id, status FROM orders WHERE LOWER(COALESCE(status,'')) IN ('shipped','partially_shipped')")
    shipped_rows = [dict(row) for row in cur.fetchall()]
    changed_to_partial = []

    for order_row in shipped_rows:
        order_id = int(order_row["id"])
        if norm(order_row.get("status")).lower() != "shipped":
            continue

        cur.execute("""
          SELECT
            COALESCE(SUM(oi.qty), 0) AS ordered_qty,
            COALESCE(SUM(MIN(oi.qty, COALESCE(done.invoiced_qty, 0))), 0) AS completed_qty
          FROM order_items oi
          LEFT JOIN (
            SELECT order_item_id, SUM(qty) AS invoiced_qty
            FROM invoice_allocations
            GROUP BY order_item_id
          ) done ON done.order_item_id=oi.id
          WHERE oi.order_id=?
        """, (order_id,))
        progress = cur.fetchone()
        ordered_qty = int(progress["ordered_qty"] or 0) if progress else 0
        completed_qty = int(progress["completed_qty"] or 0) if progress else 0

        if ordered_qty > 0 and 0 < completed_qty < ordered_qty:
            cur.execute("UPDATE orders SET status='partially_shipped' WHERE id=?", (order_id,))
            changed_to_partial.append(order_id)

    c.commit()
    c.close()

    if changed_to_partial and supabase_enabled():
        for order_id in changed_to_partial:
            try:
                supabase_update_rows("orders", {"status": "partially_shipped"}, {"id": order_id})
            except Exception:
                app.logger.exception("Nie udało się zsynchronizować częściowej wysyłki zamówienia %s", order_id)

    shipped_order_ids = [int(row["id"]) for row in shipped_rows]
    if shipped_order_ids:
        finalize_fully_invoiced_orders(shipped_order_ids)

    return changed_to_partial


def finalize_legacy_shipped_orders_with_full_invoice():
    """Zachowuje zgodność ze starszym wywołaniem naprawy statusów."""
    return reconcile_legacy_shipped_order_statuses()


def reconcile_orders_after_invoice_change(order_ids: list[int]):
    touched = sorted({int(x) for x in order_ids if x})
    if not touched:
        return [], []

    c = conn()
    cur = c.cursor()
    changed_order_ids = []
    changed_product_ids = []

    for order_id in touched:
        cur.execute("SELECT id, status, warehouse_issued FROM orders WHERE id=?", (order_id,))
        order_row = cur.fetchone()
        if not order_row:
            continue

        fully = order_fully_invoiced(cur, order_id)
        warehouse_issued = int(order_row["warehouse_issued"] or 0)
        current_status = norm(order_row["status"]).lower()

        if fully and warehouse_issued == 0:
            cur.execute("SELECT product_id, qty FROM order_items WHERE order_id=?", (order_id,))
            for it in cur.fetchall():
                pid = int(it["product_id"])
                qty = int(it["qty"] or 0)
                cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
                cur.execute("UPDATE stock SET qty = qty - ? WHERE product_id=?", (qty, pid))
                changed_product_ids.append(pid)
            preserved = {"shipped", "partially_shipped", "completed", "issued", "cancelled"}
            next_status = current_status if current_status in preserved else "packed"
            if next_status == "packed":
                cur.execute("""
                  UPDATE orders
                  SET status=?, warehouse_issued=1, packed_at=COALESCE(packed_at, ?)
                  WHERE id=?
                """, (next_status, now_iso(), order_id))
            else:
                cur.execute("UPDATE orders SET status=?, warehouse_issued=1 WHERE id=?", (next_status, order_id))
            changed_order_ids.append(order_id)

        elif not fully and warehouse_issued == 1:
            cur.execute("SELECT product_id, qty FROM order_items WHERE order_id=?", (order_id,))
            for it in cur.fetchall():
                pid = int(it["product_id"])
                qty = int(it["qty"] or 0)
                cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
                cur.execute("UPDATE stock SET qty = qty + ? WHERE product_id=?", (qty, pid))
                changed_product_ids.append(pid)
            next_status = "confirmed" if current_status in {"issued", "packed", "packed_partial"} else (current_status or "confirmed")
            cur.execute("UPDATE orders SET status=?, warehouse_issued=0 WHERE id=?", (next_status, order_id))
            changed_order_ids.append(order_id)

    c.commit()
    c.close()

    if supabase_enabled():
        if changed_order_ids:
            try:
                sync_local_rows_to_supabase("orders", "id", changed_order_ids)
            except Exception:
                pass
        if changed_product_ids:
            try:
                sync_local_rows_to_supabase("stock", "product_id", list(set(changed_product_ids)))
            except Exception:
                pass

    return changed_order_ids, list(set(changed_product_ids))


def invoice_edit_items(invoice_id: int, invoice_row: dict):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT invoice_items_json FROM invoice_meta WHERE invoice_id=?", (invoice_id,))
    saved_meta = cur.fetchone()
    saved_by_item = {}
    if saved_meta and norm(saved_meta["invoice_items_json"]):
        try:
            for saved in json.loads(saved_meta["invoice_items_json"]):
                saved_id = to_int(saved.get("order_item_id") or saved.get("id"), 0)
                if saved_id:
                    saved_by_item[saved_id] = saved
        except Exception:
            saved_by_item = {}
    cur.execute("SELECT order_id, order_item_id, qty FROM invoice_allocations WHERE invoice_id=?", (invoice_id,))
    current_alloc_rows = [dict(r) for r in cur.fetchall()]
    current_qty_by_item = {int(r["order_item_id"]): int(r["qty"] or 0) for r in current_alloc_rows}
    allocated_order_ids = {int(r["order_id"]) for r in current_alloc_rows if int(r.get("order_id") or 0)}

    email = _email_key(invoice_row.get("buyer_email"))
    if not email and invoice_row.get("order_id"):
        cur.execute("SELECT customer_email FROM orders WHERE id=?", (invoice_row.get("order_id"),))
        rr = cur.fetchone()
        email = _email_key(rr["customer_email"]) if rr else ""

    order_ids = set(allocated_order_ids)
    if invoice_row.get("order_id"):
        order_ids.add(int(invoice_row.get("order_id")))
    if email:
        status_ph = ",".join(["?"] * len(CURRENT_ORDER_STATUSES))
        cur.execute(f"""
          SELECT id
          FROM orders
          WHERE LOWER(COALESCE(customer_email,'')) = ?
            AND (
              LOWER(COALESCE(status,'')) IN ({status_ph})
              OR id IN (SELECT order_id FROM invoice_allocations WHERE invoice_id=?)
            )
          ORDER BY created_at DESC, id DESC
        """, (email, *sorted(CURRENT_ORDER_STATUSES), invoice_id))
        order_ids.update(int(r["id"]) for r in cur.fetchall())

    if not order_ids:
        c.close()
        return []

    ids = sorted(order_ids)
    ph = ",".join(["?"] * len(ids))
    cur.execute(f"""
      SELECT oi.*, p.model, p.name, COALESCE(s.qty,0) AS stock_qty,
             oo.order_no AS source_order_no,
             oo.created_at AS source_order_created_at,
             oo.note AS source_order_note,
             COALESCE(oi.unit_net_price, pr.net_price, 0) AS net_price,
             COALESCE(oi.unit_gross_price, oi.unit_net_price, pr.gross_price, pr.net_price, 0) AS gross_price,
             COALESCE(oi.currency, oo.currency, 'PLN') AS currency
      FROM order_items oi
      JOIN orders oo ON oo.id=oi.order_id
      JOIN products p ON p.id=oi.product_id
      LEFT JOIN stock s ON s.product_id=oi.product_id
      LEFT JOIN pricing pr ON (TRIM(LOWER(pr.model)) = TRIM(LOWER(p.model)) OR TRIM(LOWER(pr.model)) = TRIM(LOWER(p.sku)))
      WHERE oi.order_id IN ({ph})
      ORDER BY oo.created_at DESC, oo.id DESC, oi.id
    """, ids)
    items = [dict(r) for r in cur.fetchall()]
    c.close()

    invoiced_by_item = invoiced_qty_by_order_item_ids([int(it["id"]) for it in items])
    out = []
    for it in items:
        item_id = int(it["id"])
        current_qty = int(current_qty_by_item.get(item_id, 0))
        ordered_qty = int(it.get("qty") or 0)
        invoiced_total = int(invoiced_by_item.get(item_id, 0))
        invoiced_other = max(0, invoiced_total - current_qty)
        max_qty = max(0, ordered_qty - invoiced_other)
        row = dict(it)
        # Dla pozycji już znajdującej się na fakturze zapisany JSON jest
        # źródłem prawdy o cenie. Nie wolno zastępować EUR bieżącym cennikiem PLN.
        saved = saved_by_item.get(item_id) or {}
        order_has_foreign_snapshot = (
            normalize_invoice_type(invoice_row.get("invoice_type")) in {"wdt", "export"}
            and normalize_order_currency(it.get("currency")) != "PLN"
            and money_float(it.get("net_price")) > 0
        )
        if saved and not order_has_foreign_snapshot:
            if saved.get("net_price") not in (None, ""):
                row["net_price"] = money_float(saved.get("net_price"))
            if saved.get("gross_price") not in (None, ""):
                row["gross_price"] = money_float(saved.get("gross_price"))
            if norm(saved.get("currency")):
                row["currency"] = normalize_order_currency(saved.get("currency"))
            if saved.get("vat_rate") not in (None, ""):
                row["vat_rate"] = to_int(saved.get("vat_rate"), 23)
        row["order_item_id"] = item_id
        row["source_order_id"] = int(it.get("order_id") or 0)
        row["source_order_no"] = order_display_no(
            row["source_order_id"],
            it.get("source_order_created_at"),
            it.get("source_order_no"),
            it.get("source_order_note") or ""
        )
        row["source_order_note"] = it.get("source_order_note") or ""
        row["ordered_qty"] = ordered_qty
        row["invoiced_other_qty"] = invoiced_other
        row["current_invoice_qty"] = current_qty
        row["remaining_qty"] = max_qty
        if max_qty > 0 or current_qty > 0:
            out.append(row)
    return out


def prepare_invoice_edit_items(edit_items: list[dict], form, invoice_type="domestic", currency="PLN"):
    prepared = []
    for it in edit_items:
        max_qty = int(it.get("remaining_qty") or 0)
        qty = to_int(form.get(f"invoice_qty_{it['id']}"), 0)
        qty = max(0, min(qty, max_qty))
        if qty <= 0:
            continue
        row = dict(it)
        row["order_item_id"] = int(it.get("id") or it.get("order_item_id") or 0)
        row["source_order_id"] = int(it.get("order_id") or it.get("source_order_id") or 0)
        row["ordered_qty"] = int(it.get("ordered_qty") or it.get("qty") or 0)
        row["invoiced_qty_before"] = int(it.get("invoiced_other_qty") or 0)
        row["qty"] = qty
        line_net = money_dec(row.get("net_price")) * Decimal(qty)
        line_net = line_net.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        foreign_zero = normalize_invoice_type(invoice_type) in {"wdt", "export"}
        line_vat = Decimal("0.00") if foreign_zero else vat23_from_net(line_net)
        line_gross = (line_net + line_vat).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        row["gross_price"] = money_float(row.get("net_price")) if foreign_zero else money_float(gross_from_net_23(row.get("net_price")))
        row["vat_rate"] = 0 if foreign_zero else 23
        row["currency"] = normalize_order_currency(currency)
        row["line_value_net"] = money_float(line_net)
        row["line_value_vat"] = money_float(line_vat)
        row["line_value_gross"] = money_float(line_gross)
        prepared.append(row)
    return prepared


# =========================
# TEMPLATES (BASE as "file")
# =========================

CLIENT_API_PATHS = {
    "/api/client_stock_catalog", "/api/client_search_log", "/api/client/orders",
    "/api/order_lookup", "/api/client_invoices", "/api/client_order_email",
    "/api/client/profile",
}
_rate_lock = threading.Lock()
_rate_hits = {}
_client_auth_cache_lock = threading.Lock()
_client_auth_cache = {}
CLIENT_AUTH_CACHE_TTL_SEC = 60


def _rate_limit(bucket: str, limit: int, window_seconds: int):
    now = time.time()
    key = (bucket, request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip())
    with _rate_lock:
        hits = [ts for ts in _rate_hits.get(key, []) if now - ts < window_seconds]
        if len(hits) >= limit:
            return False
        hits.append(now)
        _rate_hits[key] = hits
    return True


def _admin_password_ok(candidate: str) -> bool:
    if ADMIN_PASSWORD_HASH:
        try:
            return check_password_hash(ADMIN_PASSWORD_HASH, candidate)
        except Exception:
            return False
    return bool(ADMIN_PASSWORD) and hmac.compare_digest(ADMIN_PASSWORD, candidate)


@app.before_request
def _start_request_performance_trace():
    if PERF_LOG_ENABLED:
        g.perf_started = time.perf_counter()
        g.perf_stages = {}


@app.after_request
def _log_request_performance(response):
    started = getattr(g, "perf_started", None)
    if PERF_LOG_ENABLED and started is not None:
        total = time.perf_counter() - started
        stages = dict(getattr(g, "perf_stages", {}))
        # supabase_http i reconciliation są podetapami pulla, więc nie mogą
        # zostać drugi raz odjęte od czasu całego requestu.
        top_level = stages.get("render_html", 0.0) + stages.get("supabase_pull_blocking", 0.0)
        stages["view_logic_sql"] = max(0.0, total - top_level)
        payload = {name: round(value * 1000, 2) for name, value in stages.items()}
        payload["total"] = round(total * 1000, 2)
        app.logger.info("PERF %s %s %s", request.method, request.path, json.dumps(payload, ensure_ascii=False, sort_keys=True))
    return response


@app.before_request
def security_gate():
    path = request.path
    if path == "/webhooks/17track":
        # Endpoint nie korzysta z sesji; sam weryfikuje podpis 17TRACK na
        # surowym body przed dotknięciem danych logistycznych.
        return None
    if path == "/login":
        if request.method == "POST" and not _rate_limit("admin_login", 8, 15 * 60):
            return "Zbyt wiele prób logowania. Spróbuj później.", 429
        return None

    is_client_api = (
        path in CLIENT_API_PATHS
        or path.startswith("/api/invoices/")
        or path.startswith("/api/client/orders/")
        or path.startswith("/api/client/product-images/")
    )
    if is_client_api:
        if request.method == "OPTIONS":
            return None
        if path == "/api/client/orders" and not _rate_limit("client_orders", 12, 10 * 60):
            return jsonify(ok=False, error="Zbyt wiele prób złożenia zamówienia"), 429
        if not _rate_limit("client_api", 180, 60):
            return jsonify(ok=False, error="Zbyt wiele żądań"), 429
        user = _authenticated_client_user()
        if not user:
            # Stare e-maile wskazuja endpoint PDF bez naglowka Authorization.
            # Przekierowujemy do panelu; dokument nadal wymaga tokenu po loginie.
            old_invoice_link = re.fullmatch(r"/api/invoices/(\d+)/download", path)
            if request.method == "GET" and old_invoice_link:
                invoice_id = int(old_invoice_link.group(1))
                panel_url = (
                    f"{CLIENT_PANEL_URL}/?"
                    + urllib.parse.urlencode({"section": "invoices", "invoice": invoice_id})
                )
                return redirect(panel_url, code=302)
            return jsonify(ok=False, error="Brak autoryzacji"), 401
        g.client_user = user
        return None

    if path == "/logout":
        return None
    if not session.get("admin_authenticated"):
        if path.startswith("/api/"):
            return jsonify(ok=False, error="Brak autoryzacji administratora"), 401
        return redirect(url_for("login", next=request.full_path if request.query_string else path))

    if request.method in {"POST", "PUT", "PATCH", "DELETE"}:
        if request.is_json:
            origin = norm(request.headers.get("Origin")).rstrip("/")
            expected = request.host_url.rstrip("/")
            if origin and origin != expected:
                return jsonify(ok=False, error="Nieprawidłowe źródło żądania"), 403
        else:
            supplied = norm(request.form.get("csrf_token") or request.headers.get("X-CSRF-Token"))
            if not supplied or not hmac.compare_digest(supplied, session.get("csrf_token", "")):
                return "Nieprawidłowy token bezpieczeństwa formularza. Odśwież stronę.", 403

@app.after_request
def security_headers_and_csrf(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(self), microphone=(), geolocation=()")
    response.headers.setdefault("Content-Security-Policy", "default-src 'self'; img-src 'self' data: blob:; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline' https://unpkg.com https://cdn.jsdelivr.net; connect-src 'self' https://*.supabase.co https://api.resend.com")
    if session.get("admin_authenticated") and response.content_type and response.content_type.startswith("text/html"):
        body = response.get_data(as_text=True)
        token = session.get("csrf_token", "")
        if token:
            hidden = f'<input type="hidden" name="csrf_token" value="{token}">'
            body = re.sub(r'(<form\b[^>]*\bmethod=["\']post["\'][^>]*>)', r'\1' + hidden, body, flags=re.I)
            response.set_data(body)
    return response

BASE = r"""
<!doctype html>
<html lang="pl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{{ title or "NiedĹşwieccy Orders" }}</title>
  <style>
    :root{--navy:#12213d;--navy2:#0b1730;--blue:#5577ee;--blue2:#3f63dc;--mint:#31b98b;--amber:#f5a524;--red:#e05263;--ink:#17233c;--muted:#718096;--bg:#f5f6fa;--line:#e7eaf2;--card:#fff;--radius:22px;--shadow:0 12px 35px rgba(31,45,78,.07)}
    *{box-sizing:border-box}html{background:var(--bg)}body{font-family:Inter,ui-sans-serif,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;margin:0;background:radial-gradient(circle at 85% -10%,#eaf0ff 0,transparent 28%),var(--bg);color:var(--ink);line-height:1.45}
    .top{position:fixed;inset:10px auto 10px 10px;width:238px;background:linear-gradient(165deg,var(--navy),var(--navy2));color:#fff;padding:22px 14px;border-radius:26px;display:flex;flex-direction:column;z-index:1100;box-shadow:0 24px 50px rgba(10,24,54,.22);overflow-y:auto}
    .brand{font-size:19px;font-weight:800;letter-spacing:-.3px;padding:4px 10px 20px;display:flex;align-items:center;gap:10px}.brand:before{content:"◇";display:grid;place-items:center;width:36px;height:36px;border:1px solid rgba(255,255,255,.32);border-radius:12px;background:rgba(255,255,255,.08);font-size:20px}
    .nav{display:flex!important;flex-direction:column;align-items:stretch!important;gap:5px!important;flex-wrap:nowrap!important;width:100%}.nav a,.nav-drop-btn{display:flex;align-items:center;color:#dce5f7;text-decoration:none;padding:11px 12px;border:0;border-radius:13px;background:transparent;font:inherit;font-size:14px;font-weight:600;cursor:pointer;transition:.18s ease}.nav a:hover,.nav-drop-btn:hover,.nav a.active{background:rgba(93,128,246,.24);color:#fff;transform:translateX(2px)}
    .nav a:before{width:25px;font-size:16px;opacity:.9}.nav a:nth-child(1):before{content:"⌂"}.nav a:nth-child(2):before{content:"▣"}.nav a:nth-child(3):before{content:"＋"}.nav a:nth-child(4):before{content:"▤"}.nav a:nth-child(5):before{content:"K"}.nav a:nth-child(6):before{content:"⌕"}.nav a:nth-child(7):before{content:"▦"}.nav a:nth-child(8):before{content:"◇"}.nav a:nth-child(9):before{content:"▧"}
    .nav-dropdown{position:relative;display:block}.nav-drop-btn{width:100%;text-align:left}.nav-drop-btn:before{content:"⚙";width:25px}.nav-dropdown-menu{display:none;margin:4px 0 2px 12px;border-left:1px solid rgba(255,255,255,.16);padding:2px 0 2px 8px}.nav-dropdown:hover .nav-dropdown-menu,.nav-dropdown:focus-within .nav-dropdown-menu{display:grid;gap:2px}.nav-dropdown-menu a{font-size:13px;padding:8px 10px}.nav-dropdown-menu a:before{display:none}
    .top>.right{margin:auto 8px 0!important;padding-top:16px;border-top:1px solid rgba(255,255,255,.13);color:#91a1bd!important;font-size:10px;overflow-wrap:anywhere}
    .mobile-toggle{display:none}.wrap{max-width:1500px;margin:0 0 0 258px;padding:28px 28px 18px;min-height:100vh}
    .card{background:rgba(255,255,255,.94);border:1px solid rgba(226,230,239,.9);border-radius:var(--radius);padding:20px;box-shadow:var(--shadow);margin-bottom:16px;overflow-x:auto}.card:hover{border-color:#dce2f1}
    .row{display:grid;grid-template-columns:1fr 1fr;gap:16px}h1{font-size:26px;letter-spacing:-.7px;margin:0 0 16px}h2{font-size:17px;letter-spacing:-.2px;margin:0 0 13px}.muted{color:var(--muted);font-size:12px}
    .btn{display:inline-flex;align-items:center;justify-content:center;gap:6px;padding:10px 14px;border:1px solid #dce1eb;border-radius:13px;background:#fff;color:var(--ink);font-weight:650;text-decoration:none;cursor:pointer;box-shadow:0 3px 10px rgba(30,44,75,.04);transition:.18s ease}.btn:hover{transform:translateY(-1px);border-color:#bfc9df;box-shadow:0 7px 16px rgba(30,44,75,.09)}.btn.primary{background:linear-gradient(135deg,var(--blue),var(--blue2));color:#fff;border-color:transparent}.btn.danger{background:#fff0f2;color:#b92d43;border-color:#ffd6dc}.btn.ok{background:#e9faf4;color:#14835f;border-color:#c6f0e2}
    input,select,textarea{width:100%;padding:11px 13px;border:1px solid #dfe3ec;border-radius:13px;background:#fbfcfe;color:var(--ink);font:inherit;font-size:14px;outline:none;transition:.18s}input:focus,select:focus,textarea:focus{border-color:#7892f3;background:#fff;box-shadow:0 0 0 4px rgba(85,119,238,.11)}textarea{min-height:90px}
    table{width:100%;border-collapse:separate;border-spacing:0;min-width:660px}th,td{border-bottom:1px solid #edf0f5;padding:12px 11px;text-align:left;vertical-align:middle}th{background:#f8f9fc;color:#64718a;font-size:11px;text-transform:uppercase;letter-spacing:.45px;font-weight:750}thead th:first-child{border-radius:12px 0 0 12px}thead th:last-child{border-radius:0 12px 12px 0}tbody tr{transition:.15s}tbody tr:hover{background:#fafbff}
    .badge{display:inline-block;padding:5px 10px;border-radius:999px;border:1px solid #dfe4ef;background:#f8faff;color:#526079;font-size:11px;font-weight:700}.st-confirmed,.badge-paid{background:#e8f9f3!important;color:#16835f!important;border-color:#c9efe2!important}.st-unconfirmed{background:#fff1f2!important;color:#be3b50!important;border-color:#ffd7dc!important}.st-delivery{background:#edf3ff!important;color:#4166d3!important;border-color:#d9e4ff!important}.st-issued{background:#f0f2f6!important;color:#667085!important}
    .flex{display:flex;gap:10px;flex-wrap:wrap;align-items:center}.right{margin-left:auto}.small{font-size:12px}.grid3{display:grid;grid-template-columns:2fr 1fr 1fr;gap:16px}.line{height:1px;background:#edf0f5;margin:16px 0}.hint{background:#fff9e9;border:1px solid #f8e6ae;padding:12px 14px;border-radius:14px;color:#7e641b;font-size:13px}.kpi{display:flex;gap:10px;flex-wrap:wrap}.kpi .pill{background:#f4f7ff;border:1px solid #e1e8fb;padding:8px 11px;border-radius:999px;color:#516582;font-size:12px}.items-row{display:grid;grid-template-columns:2fr 120px 120px 120px;gap:10px;align-items:center}
    @media(max-width:980px){.top{transform:translateX(-110%);transition:.25s}.top.open{transform:translateX(0)}.mobile-toggle{display:grid;place-items:center;position:fixed;right:14px;bottom:14px;z-index:1200;width:52px;height:52px;border:0;border-radius:17px;background:var(--navy);color:#fff;font-size:22px;box-shadow:0 12px 28px rgba(12,28,58,.28)}.wrap{margin-left:0;padding:18px 14px 80px}.row,.grid3{grid-template-columns:1fr}.items-row{grid-template-columns:1fr 1fr}}
    @media(max-width:560px){.card{padding:15px;border-radius:18px}.items-row{grid-template-columns:1fr}.flex>.right{margin-left:0}h1{font-size:22px}}
  </style>
</head>
<body>
  <button class="mobile-toggle" type="button" onclick="document.querySelector('.top').classList.toggle('open')">☰</button>
  <div class="top">
    <div class="brand">Niedźwieccy</div>
    <div class="nav flex">
      <a class="{% if request.endpoint == 'home' %}active{% endif %}" href="{{ url_for('home') }}">Pulpit</a>
      <a class="{% if request.endpoint in ['orders','order_view'] %}active{% endif %}" href="{{ url_for('orders') }}">Zamówienia</a>
      <a class="{% if request.endpoint == 'order_new' %}active{% endif %}" href="{{ url_for('order_new') }}">Nowe zamówienie</a>
      <a href="{{ url_for('invoices') }}">Faktury</a>
      <a href="{{ url_for('ksef_dashboard') }}">KSeF</a>
      <a href="{{ url_for('client_searches') }}">Wyszukiwania</a>
      <a href="{{ url_for('stock') }}">Stan magazynu</a>
      <a href="{{ url_for('china') }}">Dostawy (P/O)</a>
      <a href="{{ url_for('order_scan') }}">Skan QR</a>
      <div class="nav-dropdown">
        <button class="nav-drop-btn" type="button">Ustawienia ▾</button>
        <div class="nav-dropdown-menu">
          <a href="{{ url_for('products') }}">Produkty</a>
          <a href="{{ url_for('customers') }}">Klienci</a>
          <a href="{{ url_for('pricing') }}">Cennik</a>
          <a href="{{ url_for('company') }}">Dane mojej firmy</a>
          <a href="{{ url_for('cash_flow') }}">Cash flow</a>
          <a href="{{ url_for('email_test') }}">Test maili</a>
        </div>
      </div>
      <a href="{{ url_for('logout') }}">Wyloguj</a>
    </div>
    <div class="right muted">Magazyn główny<br>{{ base_url }}</div>
  </div>

  <div class="wrap">
    {% block content %}{% endblock %}
    <div class="muted small" style="margin:14px 2px;">Dane na dysku: <b>{{ db_path }}</b></div>
  </div>

<script>
async function refreshStock(productId, targetId){
  if(!productId){ document.getElementById(targetId).innerText = "-"; return; }
  const r = await fetch("/api/product/"+productId);
  const j = await r.json();
  document.getElementById(targetId).innerText = (j.stock ?? "-");
}

function addItemRow(){
  const tpl = document.getElementById("itemRowTpl");
  const container = document.getElementById("itemsContainer");
  const node = tpl.content.cloneNode(true);
  container.appendChild(node);
}

function removeRow(btn){
  const row = btn.closest(".items-row");
  if(row) row.remove();
}
</script>

</body>
</html>
"""

# loader: BASE dostÄ™pny jako "base.html"
app.jinja_loader = ChoiceLoader([
    DictLoader({"base.html": BASE}),
    FileSystemLoader(os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")),
])
app.jinja_env.globals["canonical_order_no"] = canonical_order_no
app.jinja_env.globals["order_display_no"] = order_display_no
app.jinja_env.globals["order_status_label"] = order_status_label if "order_status_label" in globals() else None
app.jinja_env.globals["order_status_css"] = order_status_css if "order_status_css" in globals() else None
app.jinja_env.globals["carrier_tracking_url"] = carrier_tracking_url


# =========================
# PAGES
# =========================
def client_searches_v2():
    q = norm(request.args.get("q"))
    rows, source_label = load_client_search_rows(limit=5000)

    customer_name_by_email = {}
    order_name_by_email = {}
    product_name_by_sku = {}
    known_product_names = {}
    try:
        c = conn()
        cur = c.cursor()
        cur.execute("SELECT email, name FROM customers WHERE TRIM(COALESCE(email,''))<>''")
        for rr in cur.fetchall():
            email_key = _email_key(rr["email"])
            company_name = norm(rr["name"])
            if email_key and company_name and not _order_name_is_fallback(company_name, email_key):
                customer_name_by_email[email_key] = company_name

        # Render moze miec niepelna lokalna kopie klientow. Supabase jest tutaj
        # zrodlem prawdy, wiec uzupelniamy mape nazw jednym zbiorczym odczytem.
        if supabase_enabled():
            try:
                cloud_customers = supabase_request(
                    "/rest/v1/customers",
                    params={"select": "email,name", "limit": 5000},
                    timeout=20,
                ) or []
                for rr in cloud_customers:
                    email_key = _email_key(rr.get("email"))
                    company_name = norm(rr.get("name"))
                    if email_key and company_name and not _order_name_is_fallback(company_name, email_key):
                        customer_name_by_email[email_key] = company_name
            except Exception as exc:
                app.logger.warning("Nie udalo sie pobrac nazw klientow dla historii wyszukiwan: %s", type(exc).__name__)

        cur.execute("""
          SELECT customer_email, customer_name
          FROM orders
          WHERE TRIM(COALESCE(customer_email,''))<>''
          ORDER BY id DESC
        """)
        for rr in cur.fetchall():
            email_key = _email_key(rr["customer_email"])
            company_name = norm(rr["customer_name"])
            if email_key and company_name and email_key not in order_name_by_email and not _order_name_is_fallback(company_name, email_key):
                order_name_by_email[email_key] = company_name

        cur.execute("SELECT sku, name FROM products WHERE COALESCE(archived,0)=0 AND TRIM(COALESCE(name,''))<>''")
        for rr in cur.fetchall():
            sku = norm(rr["sku"])
            product_name = norm(rr["name"])
            if sku and product_name:
                product_name_by_sku[sku.lower()] = product_name
                known_product_names.setdefault(product_name.lower(), product_name)
        c.close()
    except Exception:
        try:
            c.close()
        except Exception:
            pass

    def display_customer_name(row):
        email = _email_key(row.get("customer_email"))
        raw_name = norm(row.get("customer_name"))
        for candidate in (customer_name_by_email.get(email), order_name_by_email.get(email), raw_name):
            candidate = norm(candidate)
            if candidate and not _order_name_is_fallback(candidate, email) and "@" not in candidate:
                return candidate
        return "-"

    def canonical_product_name(row):
        product_name = norm(row.get("product_name"))
        if product_name and product_name != "-":
            return product_name
        product_sku = norm(row.get("product_sku")).lower()
        product_model = norm(row.get("product_model")).lower()
        query_key = norm(row.get("query")).lower()
        if product_sku and product_sku in product_name_by_sku:
            return product_name_by_sku[product_sku]
        if product_model and product_model in product_name_by_sku:
            return product_name_by_sku[product_model]
        if query_key and query_key in known_product_names:
            return known_product_names[query_key]
        return ""

    for row in rows:
        row["_client_label"] = display_customer_name(row)
        row["_product_label"] = canonical_product_name(row)

    if q:
        needle = q.lower()
        rows = [
            r for r in rows
            if needle in (r.get("query") or "").lower()
            or needle in (r.get("customer_name") or "").lower()
            or needle in (r.get("_client_label") or "").lower()
            or needle in (r.get("_product_label") or "").lower()
            or needle in (r.get("product_sku") or "").lower()
            or needle in (r.get("product_model") or "").lower()
        ]

    phrase_stats = {}
    model_stats = {}
    client_stats = {}
    phrase_events_seen = set()
    model_events_seen = set()
    latest_events = []

    for r in rows:
        query = norm(r.get("query"))
        if not query:
            continue
        email = norm(r.get("customer_email")).lower()
        client_label = norm(r.get("_client_label"))
        client_key = email or client_label or "anon"
        product_name = norm(r.get("_product_label"))
        product_model = norm(r.get("product_model"))
        product_sku = norm(r.get("product_sku"))
        results_count = to_int(r.get("results_count"), 0)
        created_at = norm(r.get("created_at"))

        # Ranking produktowy jest rankingiem modeli. SKU jest tylko technicznym
        # wariantem produktu i nie może tworzyć osobnego "wyszukania".
        model_label = product_model
        model_key = model_label.lower()
        model_event_key = (client_key, query.lower(), model_key, created_at)
        if model_key and results_count > 0 and model_event_key not in model_events_seen:
            model_events_seen.add(model_event_key)
            item = model_stats.setdefault(model_key, {
                "product_model": model_label,
                "product_sku": product_sku,
                "product_name": product_name,
                "searches_count": 0,
                "clients": set(),
                "last_at": "",
            })
            item["searches_count"] += 1
            item["clients"].add(client_key)
            if created_at > item["last_at"]:
                item["last_at"] = created_at

        phrase_event_key = (client_key, query.lower(), created_at)
        if phrase_event_key in phrase_events_seen:
            continue
        phrase_events_seen.add(phrase_event_key)
        latest_events.append(r)

        phrase = phrase_stats.setdefault(query, {
            "query": query,
            "searches_count": 0,
            "clients": set(),
            "no_result_count": 0,
            "max_results": 0,
            "last_at": "",
        })
        phrase["searches_count"] += 1
        phrase["clients"].add(client_key)
        if results_count == 0:
            phrase["no_result_count"] += 1
        phrase["max_results"] = max(phrase["max_results"], results_count)
        if created_at > phrase["last_at"]:
            phrase["last_at"] = created_at

        summary_name = client_label if client_label and client_label != "-" else "Nieznany klient"
        skey = summary_name
        summary = client_stats.setdefault(skey, {
            "client_label": summary_name,
            "searches_count": 0,
            "phrases": set(),
            "models": set(),
            "no_result_count": 0,
            "max_results": 0,
            "last_at": "",
        })
        summary["searches_count"] += 1
        summary["phrases"].add(query)
        if model_label:
            summary["models"].add(model_label)
        if results_count == 0:
            summary["no_result_count"] += 1
        summary["max_results"] = max(summary["max_results"], results_count)
        if created_at > summary["last_at"]:
            summary["last_at"] = created_at

    name_rows = []
    for r in model_stats.values():
        item = dict(r)
        item["clients_count"] = len(item.pop("clients"))
        name_rows.append(item)
    name_rows.sort(key=lambda r: (r["searches_count"], r["last_at"]), reverse=True)
    name_rows = name_rows[:10]

    phrase_rows = []
    for r in phrase_stats.values():
        item = dict(r)
        item["clients_count"] = len(item.pop("clients"))
        phrase_rows.append(item)
    phrase_rows.sort(key=lambda r: (r["searches_count"], r["last_at"]), reverse=True)
    phrase_rows = phrase_rows[:10]

    summary_rows = []
    for r in client_stats.values():
        item = dict(r)
        item["phrases_count"] = len(item.pop("phrases"))
        item["models_count"] = len(item.pop("models"))
        summary_rows.append(item)
    summary_rows.sort(key=lambda r: (r["searches_count"], r["last_at"]), reverse=True)
    summary_rows = summary_rows[:50]

    latest_events.sort(key=lambda r: norm(r.get("created_at")), reverse=True)
    latest_rows = latest_events[:50]
    total_count = len(phrase_events_seen)

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">Top wyszukiwania</h1>
          <span class="badge">Łącznie: {{ total_count }}</span>
          <span class="badge">{{ source_label }}</span>
        </div>
        <form method="get" class="grid3" style="margin-top:10px;">
          <input name="q" value="{{ q }}" placeholder="Szukaj: klient / fraza / nazwa">
          <button class="btn primary" type="submit">Szukaj</button>
          <a class="btn" href="{{ url_for('client_searches') }}">Wyczyść</a>
        </form>
      </div>

      <div class="card">
        <h2>Wyszukiwania według klienta</h2>
        <div class="muted" style="margin-bottom:8px;">Jedno wpisanie frazy przez klienta jest liczone jako jedno wyszukanie, niezależnie od liczby wyników.</div>
        <table>
          <thead><tr><th>Klient</th><th>Wyszukań</th><th>Różnych fraz</th><th>Bez wyników</th><th>Ostatnio</th></tr></thead>
          <tbody>
            {% for r in summary_rows %}<tr><td><b>{{ r.client_label }}</b></td><td><span class="badge">{{ r.searches_count }}</span></td><td>{{ r.phrases_count }}</td><td>{{ r.no_result_count or '-' }}</td><td class="muted">{{ r.last_at }}</td></tr>{% endfor %}
            {% if not summary_rows %}<tr><td colspan="5" class="muted">Brak zapisanych wyszukiwań.</td></tr>{% endif %}
          </tbody>
        </table>
      </div>

      <div class="card">
        <h2>Najczęściej wyszukiwane modele</h2>
        <div class="muted" style="margin-bottom:8px;">
          Ranking według modelu. Różne SKU tego samego modelu nie zwiększają licznika.
        </div>
        <table>
          <thead>
            <tr><th>Model</th><th>Nazwa</th><th>Wyszukań</th><th>Klientów</th><th>Ostatnio</th></tr>
          </thead>
          <tbody>
            {% for r in name_rows %}
              <tr>
                <td><b>{{ r.product_model or '-' }}</b></td>
                <td>{{ r.product_name or '-' }}</td>
                <td><span class="badge">{{ r.searches_count }}</span></td>
                <td>{{ r.clients_count }}</td>
                <td class="muted">{{ r.last_at }}</td>
              </tr>
            {% endfor %}
            {% if not name_rows %}
              <tr><td colspan="5" class="muted">Brak wyszukiwań przypisanych do modeli.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>

      <div class="card">
        <h2>Ostatnie wyszukiwania</h2>
        <table>
          <thead>
            <tr><th>Czas</th><th>Klient</th><th>Fraza</th><th>Nazwa</th><th>Model</th><th>Wyniki</th></tr>
          </thead>
          <tbody>
            {% for r in latest_rows %}
              <tr>
                <td class="muted">{{ r.created_at }}</td>
                <td>{{ r._client_label or '-' }}</td>
                <td><b>{{ r.query }}</b></td>
                <td>{{ r._product_label or '-' }}</td>
                <td>{{ r.product_model or '-' }}</td>
                <td>{{ r.results_count }}</td>
              </tr>
            {% endfor %}
            {% if not latest_rows %}
              <tr><td colspan="6" class="muted">Brak wpisów.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>

      <details class="card">
        <summary style="cursor:pointer;font-weight:700;font-size:16px;">Pokaż dodatkowe zestawienia: frazy i klienci</summary>

        <div style="margin-top:14px;">
          <h2>Frazy klientów</h2>
          <div class="muted" style="margin-bottom:8px;">
            Tu zostają wpisane teksty klienta. Pomaga sprawdzić, jak klienci szukają produktów i gdzie pojawiają się literówki albo brakujące nazwy.
          </div>
          <table>
            <thead>
              <tr><th>Fraza</th><th>Wyszukań</th><th>Klientów</th><th>Bez wyników</th><th>Najwięcej wyników</th><th>Ostatnio</th></tr>
            </thead>
            <tbody>
              {% for r in phrase_rows %}
                <tr>
                  <td><b>{{ r.query }}</b></td>
                  <td><span class="badge">{{ r.searches_count }}</span></td>
                  <td>{{ r.clients_count }}</td>
                  <td>{% if r.no_result_count %}<span class="badge">{{ r.no_result_count }}</span>{% else %}-{% endif %}</td>
                  <td>{{ r.max_results }}</td>
                  <td class="muted">{{ r.last_at }}</td>
                </tr>
              {% endfor %}
              {% if not phrase_rows %}
                <tr><td colspan="6" class="muted">Brak zapisanych fraz.</td></tr>
              {% endif %}
            </tbody>
          </table>
        </div>

        <div style="margin-top:18px;">
          <h2>Podsumowanie klientów</h2>
          <div class="muted" style="margin-bottom:8px;">Zbiorcze statystyki unikalnych wyszukań klientów.</div>
          <table>
            <thead>
              <tr><th>Klient</th><th>Wyszukań</th><th>Różnych fraz</th><th>Bez wyników</th><th>Ostatnio</th></tr>
            </thead>
            <tbody>
              {% for r in summary_rows %}
                <tr>
                  <td><b>{{ r.client_label }}</b></td>
                  <td><span class="badge">{{ r.searches_count }}</span></td>
                  <td>{{ r.phrases_count }}</td>
                  <td>{% if r.no_result_count %}<span class="badge">{{ r.no_result_count }}</span>{% else %}-{% endif %}</td>
                  <td class="muted">{{ r.last_at }}</td>
                </tr>
              {% endfor %}
              {% if not summary_rows %}
                <tr><td colspan="5" class="muted">Brak zapisanych wyszukiwań.</td></tr>
              {% endif %}
            </tbody>
          </table>
        </div>

      </details>
    {% endblock %}
    """
    return render_template_string(tpl, title="Top wyszukiwania", base_url=BASE_URL, db_path=DB_PATH,
                                  name_rows=name_rows, phrase_rows=phrase_rows, summary_rows=summary_rows, latest_rows=latest_rows,
                                  total_count=total_count, q=q, source_label=source_label)


app.view_functions["client_searches"] = client_searches_v2


register_cash_flow(app, {
    "conn": conn,
    "now_iso": now_iso,
    "app_now": app_now,
    "to_float": to_float,
    "maybe_pull_shared_from_supabase": maybe_pull_shared_from_supabase,
    "supabase_enabled": supabase_enabled,
    "supabase_upsert_rows": supabase_upsert_rows,
    "supabase_delete_rows": supabase_delete_rows,
    "BASE_URL": BASE_URL,
    "DB_PATH": DB_PATH,
})


@app.after_request
def auto_sync_after_write(response):
    try:
        is_client_api = (
            request.path in CLIENT_API_PATHS
            or request.path.startswith("/api/invoices/")
            or request.path.startswith("/api/client/orders/")
        )
        if is_client_api:
            origin = norm(request.headers.get("Origin")).rstrip("/")
            if origin and origin in CLIENT_ALLOWED_ORIGINS:
                response.headers["Access-Control-Allow-Origin"] = origin
                response.headers["Vary"] = "Origin"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, PATCH, OPTIONS"
            response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, Idempotency-Key"
        elif request.path.startswith("/api/"):
            response.headers["Access-Control-Allow-Origin"] = "*"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
            response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"

        no_auto_sync_paths = {
            "/api/client_search_log", "/api/client_order_email", "/api/client/profile"
        }
        if response.status_code < 400 and request.method in ("POST", "PUT", "PATCH", "DELETE") and request.path not in no_auto_sync_paths:
            trigger_background_supabase_sync(reason=f"{request.method} {request.path}")
    except Exception:
        pass
    return response


# -------------------------
# COMPANY
# -------------------------

def normalize_company_bank_fields(bank_account, bank_swift=""):
    """Separate a legacy `IBAN BIC` value and store both fields canonically."""
    account_raw = norm(bank_account).upper()
    swift_raw = norm(bank_swift).upper()

    # Starsze instalacje przechowywały czasem IBAN i SWIFT w jednym polu.
    match = re.fullmatch(
        r"\s*(PL)?(\d{26})(?:\s+([A-Z]{6}[A-Z0-9]{2}(?:[A-Z0-9]{3})?))?\s*",
        account_raw,
    )
    if match:
        account_raw = match.group(2)
        if not swift_raw and match.group(3):
            swift_raw = match.group(3)

    # Numer konta zapisujemy bez PL, spacji i myślników. Walidację sumy
    # kontrolnej nadal wykonuje moduł KSeF przed utworzeniem XML.
    compact_account = re.sub(r"[\s-]+", "", account_raw)
    if compact_account.startswith("PL"):
        compact_account = compact_account[2:]
    if re.fullmatch(r"\d{26}", compact_account):
        account_raw = compact_account

    return account_raw, re.sub(r"\s+", "", swift_raw)
# -------------------------
# PRICING
# -------------------------
def parse_eur_pricing_xlsx(file_obj) -> list[tuple[str, str, float, float]]:
    """Read the supplier EU price list without changing products or stock."""
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ValueError("Brak biblioteki openpyxl do odczytu XLSX") from exc

    try:
        wb = load_workbook(file_obj, data_only=True, read_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        raise ValueError("Nie udało się odczytać pliku XLSX") from exc
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not raw_rows:
        raise ValueError("Pusty plik")

    header_index = None
    indexes = None
    for row_index, raw_header in enumerate(raw_rows[:20]):
        headers = [norm(value) for value in (raw_header or ())]
        i_sku = guess_col(headers, ["articel", "artikel", "article", "sku"])
        i_ean = guess_col(headers, ["gtin", "ean"])
        i_price = guess_col(headers, ["preis eur", "price eur", "cena eur"])
        i_uvp = guess_col(headers, ["uvp"])
        if i_sku is not None and i_price is not None:
            header_index = row_index
            indexes = (i_sku, i_ean, i_price, i_uvp)
            break

    if header_index is None or indexes is None:
        raise ValueError("Nie znaleziono kolumn Articel/SKU i PREIS EUR")

    i_sku, i_ean, i_price, i_uvp = indexes
    parsed = []
    seen_skus = set()
    for excel_row_no, row in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        row = tuple(row or ())
        sku = norm(row[i_sku]) if len(row) > i_sku else ""
        if not sku:
            continue
        sku_key = sku.casefold()
        if sku_key in seen_skus:
            raise ValueError(f"Powtórzony SKU w wierszu {excel_row_no}: {sku}")
        price_value = row[i_price] if len(row) > i_price else None
        price_eur = money_float(to_float(price_value, -1))
        if price_eur <= 0:
            raise ValueError(f"Nieprawidłowa cena EUR w wierszu {excel_row_no}: {sku}")
        uvp_value = row[i_uvp] if i_uvp is not None and len(row) > i_uvp else None
        uvp_eur = money_float(to_float(uvp_value, price_eur * 1.45))
        if uvp_eur <= 0:
            uvp_eur = money_float(price_eur * 1.45)
        ean = norm(row[i_ean]) if i_ean is not None and len(row) > i_ean else ""
        parsed.append((sku, ean, price_eur, uvp_eur))
        seen_skus.add(sku_key)

    if not parsed:
        raise ValueError("Plik nie zawiera żadnych pozycji cennika UE")
    return parsed

# -------------------------
# CUSTOMERS
# -------------------------
# -------------------------
# PRODUCTS
# -------------------------
# -------------------------
# STOCK
# -------------------------
# -------------------------
# ORDERS
# -------------------------

def missed_stock_issue_candidates(cur):
    cur.execute("""
      SELECT o.id, o.order_no, o.created_at, o.shipped_at, o.status,
             COALESCE(SUM(oi.qty),0) AS item_qty,
             GROUP_CONCAT(oi.sku || ' × ' || oi.qty, ', ') AS items_label
      FROM orders o
      JOIN order_items oi ON oi.order_id=o.id
      WHERE COALESCE(o.warehouse_issued,0)=0
        AND TRIM(COALESCE(o.shipped_at,''))<>''
        AND LOWER(COALESCE(o.status,'')) IN ('shipped','issued','completed')
      GROUP BY o.id
      ORDER BY o.shipped_at, o.id
    """)
    return [dict(row) for row in cur.fetchall()]

def _packed_package_orders(cur, order):
    packed_at = norm(order.get("packed_at"))
    recipient = _email_key(order.get("customer_email"))
    if not packed_at or not recipient:
        return [order]
    cur.execute(
        """SELECT * FROM orders
           WHERE packed_at=? AND LOWER(TRIM(COALESCE(customer_email,'')))=?
             AND LOWER(COALESCE(status,'')) NOT IN ('cancelled')
           ORDER BY id""",
        (packed_at, recipient),
    )
    rows = [dict(row) for row in cur.fetchall()]
    return rows or [order]


def inpost_label_allowed_for_status(status) -> bool:
    return norm(status).lower() in {"packed", "packed_partial", "shipped", "completed", "issued"}

# -------------------------
# LABEL 30x50 (QR + dane)
# -------------------------
def _client_stock_catalog_rows(profile: dict) -> list[dict]:
    """Return live Supabase stock with the server-selected customer price list."""
    price_list = normalize_client_price_list((profile or {}).get("price_list"))
    currency = price_list_currency(price_list)
    rows = supabase_request(
        "/rest/v1/client_stock_catalog_v",
        method="GET",
        params={
            "select": (
                "product_id,sku,model,name,qty_physical,qty_reserved,"
                "qty_on_stock,net_price,gross_price"
            ),
            "order": "sku.asc",
            "limit": 5000,
        },
        timeout=30,
    ) or []
    if not isinstance(rows, list):
        raise RuntimeError("Nieprawidłowa odpowiedź katalogu magazynowego")

    # Do panelu klienta przekazujemy jedynie zagregowaną liczbę sztuk w
    # aktywnych dostawach oraz identyfikator przypisanego zdjęcia. Nie
    # ujawniamy numerów P/O, dostawców, kosztów ani danych trackingowych.
    assignments = supabase_request(
        "/rest/v1/product_image_assignments",
        method="GET",
        params={"select": "product_id,image_id", "limit": 5000},
        timeout=20,
    ) or []
    image_by_product = {
        to_int(item.get("product_id"), 0): to_int(item.get("image_id"), 0)
        for item in assignments if isinstance(item, dict)
    }
    product_details = supabase_request(
        "/rest/v1/products",
        method="GET",
        params={"select": "id,ean", "archived": "eq.0", "limit": 5000},
        timeout=20,
    ) or []
    ean_by_product = {
        to_int(item.get("id"), 0): norm(item.get("ean"))
        for item in product_details if isinstance(item, dict)
    }

    active_packages = supabase_request(
        "/rest/v1/china_packages",
        method="GET",
        params={
            "select": "id",
            "status": "in.(ordered,shipped,problem)",
            "limit": 5000,
        },
        timeout=20,
    ) or []
    active_package_ids = {
        to_int(item.get("id"), 0) for item in active_packages if isinstance(item, dict)
    }
    incoming_by_product = {}
    if active_package_ids:
        package_filter = ",".join(str(value) for value in sorted(active_package_ids) if value)
        incoming_rows = supabase_request(
            "/rest/v1/china_items",
            method="GET",
            params={
                "select": "product_id,package_id,qty",
                "package_id": f"in.({package_filter})",
                "limit": 10000,
            },
            timeout=20,
        ) or []
        for item in incoming_rows:
            if not isinstance(item, dict):
                continue
            product_id = to_int(item.get("product_id"), 0)
            if product_id:
                incoming_by_product[product_id] = incoming_by_product.get(product_id, 0) + max(0, to_int(item.get("qty"), 0))

    eur_by_sku = {}
    if price_list == "eu_eur":
        eur_rows = supabase_request(
            "/rest/v1/pricing_eur",
            method="GET",
            params={"select": "sku,price_eur,uvp_eur", "order": "sku.asc", "limit": 5000},
            timeout=30,
        ) or []
        if not isinstance(eur_rows, list):
            raise RuntimeError("Nieprawidłowa odpowiedź cennika EUR")
        eur_by_sku = {norm(row.get("sku")).lower(): row for row in eur_rows}

    catalog = []
    for raw in rows:
        row = dict(raw)
        product_id = to_int(row.get("product_id"), 0)
        incoming_qty = incoming_by_product.get(product_id, 0)
        physical_qty = max(0, to_int(row.get("qty_physical"), 0))
        reserved_qty = max(0, to_int(row.get("qty_reserved"), 0))
        # Ta sama wartość, którą magazyn pokazuje w kolumnie
        # „Dostępne w drodze”: z dostawy odejmujemy tę część rezerwacji,
        # której nie pokrywa aktualny stan fizyczny.
        reserved_incoming = min(incoming_qty, max(0, reserved_qty - physical_qty))
        available_incoming = max(0, incoming_qty - reserved_incoming)
        sku_key = norm(row.get("sku")).lower()
        if price_list == "eu_eur":
            price = eur_by_sku.get(sku_key) or {}
            net_price = money_float(price.get("price_eur"))
            # Cennik UE jest cennikiem B2B. Kwota zamówienia brutto jest równa
            # cenie transakcyjnej; UVP pozostaje osobną ceną sugerowaną.
            gross_price = net_price
            retail_price = money_float(price.get("uvp_eur"))
        else:
            net_price = money_float(row.get("net_price"))
            gross_price = money_float(row.get("gross_price"))
            retail_price = money_float(money_dec(net_price) * Decimal("1.45") * Decimal("1.23"))
        row.update({
            "net_price": net_price,
            "gross_price": gross_price,
            "retail_price": retail_price,
            "currency": currency,
            "price_list": price_list,
            "price_available": bool(net_price > 0),
            "image_id": image_by_product.get(product_id, 0),
            "qty_in_delivery": available_incoming,
            "qty_available_in_delivery": available_incoming,
            "ean": ean_by_product.get(product_id, ""),
        })
        catalog.append(row)
    return catalog

def _email_event_already_ok(event_key):
    if not event_key:
        return False
    c = conn()
    try:
        cur = c.cursor()
        cur.execute("SELECT ok FROM email_events WHERE event_key=? LIMIT 1", (event_key,))
        row = cur.fetchone()
        return bool(row and to_int(row["ok"], 0) == 1)
    except Exception:
        return False
    finally:
        c.close()


def _record_email_event(event_key, event_type, ref_id, recipient, result):
    if not event_key:
        return
    ok = 1 if isinstance(result, dict) and result.get("ok") else 0
    try:
        payload = json.dumps(result or {}, ensure_ascii=False)[:6000]
    except Exception:
        payload = json.dumps({"raw": str(result)}, ensure_ascii=False)[:6000]
    c = conn()
    try:
        cur = c.cursor()
        cur.execute("SELECT id FROM email_events WHERE event_key=? LIMIT 1", (event_key,))
        row = cur.fetchone()
        if row:
            cur.execute("""
                UPDATE email_events
                SET event_type=?, ref_id=?, recipient=?, ok=?, result_json=?, created_at=?
                WHERE event_key=?
            """, (event_type, str(ref_id or ""), recipient or "", ok, payload, now_iso(), event_key))
        else:
            cur.execute("""
                INSERT INTO email_events(event_key,event_type,ref_id,recipient,ok,result_json,created_at)
                VALUES(?,?,?,?,?,?,?)
            """, (event_key, event_type, str(ref_id or ""), recipient or "", ok, payload, now_iso()))
        c.commit()
    except Exception:
        pass
    finally:
        c.close()


def _partial_packing_order_ids(order_ids, packing_items) -> set[int]:
    """Return orders whose current packing list does not cover every ordered item."""
    clean_ids = sorted({to_int(value, 0) for value in (order_ids or []) if to_int(value, 0) > 0})
    if not clean_ids or not packing_items:
        return set()
    selected_by_item = {}
    for item in packing_items:
        item_id = to_int(item.get("order_item_id") or item.get("id"), 0)
        if item_id <= 0:
            continue
        selected_by_item[item_id] = selected_by_item.get(item_id, 0) + max(0, to_int(item.get("qty"), 0))
    c = conn()
    try:
        placeholders = ",".join(["?"] * len(clean_ids))
        cur = c.cursor()
        cur.execute(
            f"SELECT id, order_id, qty FROM order_items WHERE order_id IN ({placeholders})",
            tuple(clean_ids),
        )
        rows = [dict(row) for row in cur.fetchall()]
    finally:
        c.close()
    rows_by_order = {}
    for row in rows:
        rows_by_order.setdefault(to_int(row.get("order_id"), 0), []).append(row)
    partial_ids = set()
    for order_id in clean_ids:
        order_rows = rows_by_order.get(order_id, [])
        if not order_rows or any(
            selected_by_item.get(to_int(row.get("id"), 0), 0) < max(0, to_int(row.get("qty"), 0))
            for row in order_rows
        ):
            partial_ids.add(order_id)
    return partial_ids


def mark_orders_packed(order_ids, packing_path: str = "", packing_items=None) -> list[int]:
    """Mark selected orders as being packed without issuing stock again."""
    clean_ids = sorted({to_int(value, 0) for value in (order_ids or []) if to_int(value, 0) > 0})
    if not clean_ids:
        return []
    partial_ids = _partial_packing_order_ids(clean_ids, packing_items)
    c = conn()
    try:
        placeholders = ",".join(["?"] * len(clean_ids))
        cur = c.cursor()
        cur.execute(
            f"SELECT id, LOWER(COALESCE(status,'')) AS status FROM orders WHERE id IN ({placeholders})",
            tuple(clean_ids),
        )
        already_packed_ids = {
            int(row["id"]) for row in cur.fetchall()
            if norm(row["status"]).lower() in {"packed", "packed_partial"}
        }
        packed_at = now_iso()
        for packed_order_id in clean_ids:
            next_status = "packed_partial" if packed_order_id in partial_ids else "packed"
            cur.execute(
                """UPDATE orders SET status=?, packed_at=?
                   WHERE id=? AND LOWER(COALESCE(status,''))
                   NOT IN ('issued','completed','cancelled','shipped')""",
                (next_status, packed_at, packed_order_id),
            )
        c.commit()
        cur.execute(
            f"SELECT id FROM orders WHERE id IN ({placeholders}) AND status IN ('packed','packed_partial')",
            tuple(clean_ids),
        )
        changed_ids = [int(row["id"]) for row in cur.fetchall()]
    finally:
        c.close()
    if changed_ids and supabase_enabled():
        try:
            sync_local_rows_to_supabase("orders", "id", changed_ids)
        except Exception as exc:
            app.logger.warning("Nie udało się zsynchronizować statusu pakowania: %s", exc)
    if changed_ids:
        c = conn()
        try:
            placeholders = ",".join(["?"] * len(changed_ids))
            cur = c.cursor()
            cur.execute(f"SELECT * FROM orders WHERE id IN ({placeholders})", tuple(changed_ids))
            packed_orders = [dict(row) for row in cur.fetchall()]
        finally:
            c.close()
        pending_orders = [
            order for order in packed_orders
            if to_int(order.get("id"), 0) not in already_packed_ids
            if not _email_event_already_ok(f"order_packed:{to_int(order.get('id'), 0)}")
        ]
        orders_by_recipient = {}
        for packed_order in pending_orders:
            recipient = _email_key(packed_order.get("customer_email"))
            orders_by_recipient.setdefault(recipient, []).append(packed_order)
        for recipient, recipient_orders in orders_by_recipient.items():
            try:
                result = _send_orders_packed_email(
                    recipient_orders,
                    packing_path=packing_path,
                )
            except Exception as exc:
                result = {"ok": False, "error": str(exc)}
            for packed_order in recipient_orders:
                packed_id = to_int(packed_order.get("id"), 0)
                _record_email_event(
                    f"order_packed:{packed_id}",
                    "order_packed",
                    packed_id,
                    recipient,
                    result,
                )
            if not result.get("ok"):
                app.logger.warning(
                    "Nie udalo sie wyslac zbiorczej informacji o pakowaniu zamowien %s: %s",
                    ",".join(str(to_int(order.get("id"), 0)) for order in recipient_orders),
                    norm(result.get("error")) or "nieznany blad",
                )
    return changed_ids


def _send_orders_packed_email(orders: list[dict], packing_path: str = "") -> dict:
    if not send_email:
        return {"ok": False, "error": "Modul wysylki e-mail nie jest dostepny"}
    orders = [dict(order) for order in (orders or []) if order]
    if not orders:
        return {"ok": False, "error": "Brak zamowien do wyslania"}
    order = orders[0]
    recipient = _email_key(order.get("customer_email"))
    if not recipient:
        return {"ok": False, "error": "Zamowienie nie ma adresu e-mail klienta"}
    try:
        language = normalize_client_language(_client_profile_for_email(recipient).get("language"))
    except Exception:
        language = "pl"
    order_numbers = [
        canonical_order_no(item.get("id"), item.get("created_at"), item.get("order_no"))
        for item in orders
    ]
    multiple = len(order_numbers) > 1
    messages = {
        "pl": (("Rozpoczęliśmy pakowanie zamówień" if multiple else "Rozpoczęliśmy pakowanie zamówienia"), ("Twoje zamówienia są pakowane razem i będą oczekiwać na odbiór przez kuriera." if multiple else "Twoje zamówienie jest w trakcie pakowania i będzie oczekiwać na odbiór przez kuriera.")),
        "de": (("Wir verpacken Ihre Bestellungen" if multiple else "Wir verpacken Ihre Bestellung"), ("Ihre Bestellungen werden gemeinsam verpackt und warten anschließend auf die Abholung durch den Kurier." if multiple else "Ihre Bestellung wird derzeit verpackt und wartet anschließend auf die Abholung durch den Kurier.")),
        "en": (("We are packing your orders" if multiple else "We are packing your order"), ("Your orders are being packed together and will then wait for courier collection." if multiple else "Your order is being packed and will then wait for courier collection.")),
        "es": (("Estamos preparando tus pedidos" if multiple else "Estamos preparando tu pedido"), ("Tus pedidos se están preparando juntos y después quedarán a la espera de la recogida por el transportista." if multiple else "Tu pedido se está preparando y después quedará a la espera de la recogida por el transportista.")),
        "it": (("Stiamo preparando i tuoi ordini" if multiple else "Stiamo preparando il tuo ordine"), ("I tuoi ordini vengono preparati insieme e saranno poi in attesa del ritiro da parte del corriere." if multiple else "Il tuo ordine è in fase di preparazione e sarà poi in attesa del ritiro da parte del corriere.")),
    }
    subject_text, intro = messages.get(language, messages["pl"])
    attachment_note = {
        "pl": "Lista pakowania znajduje się w załączniku.",
        "de": "Die Packliste finden Sie im Anhang.",
        "en": "The packing list is attached.",
        "es": "La lista de embalaje está adjunta.",
        "it": "La lista di imballaggio è allegata.",
    }.get(language, "Lista pakowania znajduje się w załączniku.")
    footer_text = {
        "pl": "Zespół Niedźwieccy", "de": "Ihr Niedźwieccy-Team",
        "en": "The Niedźwieccy Team", "es": "Equipo Niedźwieccy",
        "it": "Il team Niedźwieccy",
    }.get(language, "Zespół Niedźwieccy")
    safe_orders = "".join(
        f"<li style='margin:4px 0'><b>{html.escape(str(order_no), quote=True)}</b></li>"
        for order_no in order_numbers
    )
    html_body = (
        "<div style='margin:0;padding:28px 14px;background:#f3f6fb;font-family:Arial,sans-serif;color:#10203d'>"
        "<div style='max-width:620px;margin:0 auto;background:#fff;border:1px solid #e2e8f2;border-radius:18px;overflow:hidden'>"
        "<div style='padding:30px 34px 28px'>"
        f"<h1 style='margin:0 0 24px;font-size:26px;line-height:1.25'>{html.escape(subject_text)}</h1>"
        f"<p style='margin:0 0 22px;line-height:1.6'>{html.escape(intro)}</p>"
        "<div style='padding:20px;background:#f7f9fd;border:1px solid #e3e9f3;border-radius:14px;line-height:1.8'>"
        f"<div style='color:#62708c'>Numery zamówień:</div><ul style='margin:4px 0;padding-left:22px'>{safe_orders}</ul>"
        "</div>"
        f"<p style='margin:22px 0 0;line-height:1.6;color:#52617c'>{html.escape(attachment_note)}</p>"
        f"<p style='margin:26px 0 0;line-height:1.6;color:#52617c'><b>{html.escape(footer_text)}</b></p>"
        "</div></div></div>"
    )
    text_body = f"{subject_text}\n{intro}\n" + "\n".join(order_numbers) + f"\n{attachment_note}"
    if packing_path and os.path.exists(packing_path):
        with open(packing_path, "rb") as pdf_file:
            packing_content = pdf_file.read()
        if not packing_content:
            return {"ok": False, "error": "Wygenerowana lista pakowania jest pusta"}
        packing_attachment = {
            "filename": f"{safe_filename(order_numbers[0])}_lista_pakowania.pdf",
            "content": packing_content,
        }
    else:
        packing_attachment = _order_packing_list_email_attachment(order)
    subject_orders = ", ".join(order_numbers)
    return send_email(
        recipient,
        f"{subject_text} – {subject_orders}",
        html_body,
        text_body,
        attachments=[packing_attachment],
    )


def _send_order_packed_email(order: dict) -> dict:
    """Compatibility wrapper for callers that still pack one order."""
    return _send_orders_packed_email([order])


def _order_packing_list_email_attachment(order: dict) -> dict:
    """Return the packed PDF, or safely recreate it from the saved order."""
    order_id = to_int(order.get("id"), 0)
    order_no = canonical_order_no(order_id, order.get("created_at"), order.get("order_no"))
    customer_dir = invoice_dir_for_customer(order.get("customer_name") or "Klient")
    expected_path = packing_list_pdf_path_for_invoice(
        os.path.join(customer_dir, f"{safe_filename(order_no)}.pdf"),
        order_no,
    )
    if not os.path.exists(expected_path):
        c = conn()
        try:
            cur = c.cursor()
            cur.execute("""
              SELECT oi.id, oi.product_id, oi.qty,
                     COALESCE(NULLIF(oi.sku,''), p.sku, '') AS sku,
                     COALESCE(p.model, '') AS model,
                     COALESCE(p.name, '') AS name
              FROM order_items oi
              LEFT JOIN products p ON p.id=oi.product_id
              WHERE oi.order_id=? AND COALESCE(oi.qty,0)>0
              ORDER BY oi.id
            """, (order_id,))
            items = [dict(row) for row in cur.fetchall()]
        finally:
            c.close()
        if not items:
            raise ValueError("zamówienie nie ma pozycji do listy pakowania")
        note = norm(order.get("note"))
        for item in items:
            item["source_order_no"] = order_no
            item["source_order_note"] = note
        meta = {
            "invoice_no": order_no,
            "document_label_key": "order",
            "buyer_name": norm(order.get("customer_name")),
            "buyer_email": norm(order.get("customer_email")),
        }
        expected_path = generate_invoice_packing_list_pdf(order, items, meta)
    with open(expected_path, "rb") as pdf_file:
        content = pdf_file.read()
    if not content:
        raise ValueError("wygenerowana lista pakowania jest pusta")
    return {
        "filename": f"{safe_filename(order_no)}_lista_pakowania.pdf",
        "content": content,
    }


def _send_orders_shipped_email(orders: list[dict], tracking_no: str, carrier: str, packing_attachment: dict) -> dict:
    """Send one shipment message for every order included in one packing batch."""
    orders = [dict(order) for order in (orders or []) if order]
    if not orders:
        return {"ok": False, "error": "Brak zamówień w przesyłce"}
    order = orders[0]
    order_numbers = []
    for packed_order in orders:
        order_no = canonical_order_no(
            packed_order.get("id"),
            packed_order.get("created_at"),
            packed_order.get("order_no"),
        )
        if order_no and order_no not in order_numbers:
            order_numbers.append(order_no)
    if not order_numbers:
        return {"ok": False, "error": "Brak numerów zamówień w przesyłce"}
    if len(order_numbers) == 1:
        return _send_order_shipped_email(order, tracking_no, carrier, packing_attachment)

    if not send_email:
        return {"ok": False, "error": "Moduł wysyłki e-mail nie jest dostępny"}
    recipient = _email_key(order.get("customer_email"))
    if not recipient:
        return {"ok": False, "error": "Zamówienia nie mają adresu e-mail klienta"}
    try:
        language = normalize_client_language(_client_profile_for_email(recipient).get("language"))
    except Exception:
        language = "pl"

    messages = {
        "pl": {
            "subject": "Twoje zamówienia {numbers} są już w drodze",
            "title": "Zamówienia zostały wysłane",
            "greeting": "Dzień dobry,",
            "intro": "Z przyjemnością informujemy, że poniższe zamówienia zostały spakowane razem i przekazane kurierowi w jednej przesyłce.",
            "orders": "Numery zamówień", "carrier": "Przewoźnik", "tracking": "Numer przesyłki",
            "button": "Śledź swoją przesyłkę",
            "attachment": "W załączniku znajduje się wspólna lista pakowania ze szczegółami wysłanych produktów.",
            "thanks": "Dziękujemy za zamówienia i życzymy udanego dnia!", "team": "Zespół Niedźwieccy",
        },
        "de": {
            "subject": "Ihre Bestellungen {numbers} sind unterwegs",
            "title": "Ihre Bestellungen wurden versandt",
            "greeting": "Guten Tag,",
            "intro": "Wir freuen uns, Ihnen mitzuteilen, dass die folgenden Bestellungen gemeinsam verpackt und in einer Sendung an den Paketdienst übergeben wurden.",
            "orders": "Bestellnummern", "carrier": "Paketdienst", "tracking": "Sendungsnummer",
            "button": "Sendung verfolgen",
            "attachment": "Im Anhang finden Sie die gemeinsame Packliste mit den Details der versandten Produkte.",
            "thanks": "Vielen Dank für Ihre Bestellungen. Wir wünschen Ihnen einen schönen Tag!", "team": "Ihr Niedźwieccy-Team",
        },
        "en": {
            "subject": "Your orders {numbers} are on their way",
            "title": "Your orders have been shipped",
            "greeting": "Hello,",
            "intro": "We are pleased to let you know that the following orders were packed together and handed over to the courier in one shipment.",
            "orders": "Order numbers", "carrier": "Courier", "tracking": "Tracking number",
            "button": "Track your shipment",
            "attachment": "The attached combined packing list contains the details of the shipped products.",
            "thanks": "Thank you for your orders and have a great day!", "team": "The Niedźwieccy Team",
        },
        "es": {
            "subject": "Tus pedidos {numbers} ya están en camino",
            "title": "Tus pedidos han sido enviados",
            "greeting": "Buenos días,",
            "intro": "Nos complace informarte de que los siguientes pedidos se embalaron juntos y se entregaron al transportista en un solo envío.",
            "orders": "Números de pedido", "carrier": "Transportista", "tracking": "Número de seguimiento",
            "button": "Seguir el envío",
            "attachment": "En el archivo adjunto encontrarás la lista de embalaje conjunta con los detalles de los productos enviados.",
            "thanks": "¡Gracias por tus pedidos y que tengas un buen día!", "team": "Equipo Niedźwieccy",
        },
        "it": {
            "subject": "I tuoi ordini {numbers} sono in viaggio",
            "title": "I tuoi ordini sono stati spediti",
            "greeting": "Buongiorno,",
            "intro": "Siamo lieti di informarti che i seguenti ordini sono stati imballati insieme e affidati al corriere in un'unica spedizione.",
            "orders": "Numeri ordine", "carrier": "Corriere", "tracking": "Numero di tracciamento",
            "button": "Traccia la spedizione",
            "attachment": "In allegato trovi la lista di imballaggio cumulativa con i dettagli dei prodotti spediti.",
            "thanks": "Grazie per i tuoi ordini e buona giornata!", "team": "Il team Niedźwieccy",
        },
    }
    copy = messages.get(language, messages["pl"])
    if any(norm(item.get("status")).lower() == "partially_shipped" for item in orders):
        partial_copy = {
            "pl": ("Częściowa wysyłka zamówień {numbers}", "Zamówienia zostały wysłane częściowo", "Przekazaliśmy kurierowi część produktów z poniższych zamówień. Pozostałe pozycje wyślemy osobno."),
            "de": ("Teillieferung der Bestellungen {numbers}", "Ihre Bestellungen wurden teilweise versandt", "Ein Teil der Produkte aus den folgenden Bestellungen wurde an den Paketdienst übergeben. Die übrigen Positionen werden separat versandt."),
            "en": ("Partial shipment of orders {numbers}", "Your orders have been partially shipped", "Some products from the following orders have been handed over to the courier. The remaining items will be shipped separately."),
            "es": ("Envío parcial de los pedidos {numbers}", "Tus pedidos se han enviado parcialmente", "Hemos entregado al transportista una parte de los productos de los siguientes pedidos. Los artículos restantes se enviarán por separado."),
            "it": ("Spedizione parziale degli ordini {numbers}", "I tuoi ordini sono stati spediti parzialmente", "Abbiamo affidato al corriere una parte dei prodotti dei seguenti ordini. Gli articoli rimanenti saranno spediti separatamente."),
        }.get(language)
        if partial_copy:
            copy = dict(copy, subject=partial_copy[0], title=partial_copy[1], intro=partial_copy[2])
    numbers_text = ", ".join(order_numbers)
    tracking_url = carrier_tracking_url(carrier, tracking_no)
    carrier_names = {"inpost": "InPost", "dpd": "DPD", "fedex": "FedEx", "dhl": "DHL", "ups": "UPS"}
    carrier_name = carrier_names.get(norm(carrier).lower(), norm(carrier) or "—")
    order_list_html = "".join(
        f"<li style='margin:3px 0'><b>{html.escape(str(number), quote=True)}</b></li>"
        for number in order_numbers
    )
    safe_tracking = html.escape(str(tracking_no), quote=True)
    safe_carrier = html.escape(carrier_name, quote=True)
    safe_tracking_url = html.escape(tracking_url, quote=True)
    html_body = (
        "<div style='margin:0;padding:28px 14px;background:#f3f6fb;font-family:Arial,sans-serif;color:#10203d'>"
        "<div style='max-width:620px;margin:0 auto;background:#fff;border:1px solid #e2e8f2;border-radius:18px;overflow:hidden'>"
        "<div style='padding:30px 34px 20px'>"
        f"<h1 style='margin:0 0 24px;font-size:26px;line-height:1.25'>{html.escape(copy['title'])}</h1>"
        f"<p style='margin:0 0 14px'>{html.escape(copy['greeting'])}</p>"
        f"<p style='margin:0 0 24px;line-height:1.6'>{html.escape(copy['intro'])}</p>"
        "<div style='padding:20px;background:#f7f9fd;border:1px solid #e3e9f3;border-radius:14px;line-height:1.8'>"
        f"<div style='color:#62708c'>{html.escape(copy['orders'])}:</div><ul style='margin:3px 0 10px;padding-left:22px'>{order_list_html}</ul>"
        f"<div><span style='color:#62708c'>{html.escape(copy['carrier'])}:</span> <b>{safe_carrier}</b></div>"
        f"<div><span style='color:#62708c'>{html.escape(copy['tracking'])}:</span> <b>{safe_tracking}</b></div>"
        "</div>"
        f"<p style='margin:24px 0'><a href='{safe_tracking_url}' style='display:inline-block;padding:13px 22px;background:#4f70eb;color:#fff;text-decoration:none;font-weight:bold;border-radius:10px'>{html.escape(copy['button'])}</a></p>"
        f"<p style='margin:0 0 24px;line-height:1.6;color:#52617c'>{html.escape(copy['attachment'])}</p>"
        f"<p style='margin:0;line-height:1.6'>{html.escape(copy['thanks'])}<br><br><b>{html.escape(copy['team'])}</b></p>"
        "</div></div></div>"
    )
    text_orders = "\n".join(f"- {number}" for number in order_numbers)
    text_body = (
        f"{copy['greeting']}\n\n{copy['intro']}\n\n{copy['orders']}:\n{text_orders}\n\n"
        f"{copy['carrier']}: {carrier_name}\n{copy['tracking']}: {tracking_no}\n{tracking_url}\n\n"
        f"{copy['attachment']}\n\n{copy['thanks']}\n{copy['team']}"
    )
    return send_email(
        recipient,
        copy["subject"].format(numbers=numbers_text),
        html_body,
        text_body,
        attachments=[packing_attachment],
    )


def _send_order_shipped_email(order: dict, tracking_no: str, carrier: str, packing_attachment: dict) -> dict:
    if not send_email:
        return {"ok": False, "error": "Moduł wysyłki e-mail nie jest dostępny"}
    recipient = _email_key(order.get("customer_email"))
    if not recipient:
        return {"ok": False, "error": "Zamówienie nie ma adresu e-mail klienta"}
    try:
        language = normalize_client_language(_client_profile_for_email(recipient).get("language"))
    except Exception:
        language = "pl"
    order_no = canonical_order_no(order.get("id"), order.get("created_at"), order.get("order_no"))
    messages = {
        "pl": {
            "subject": "Twoje zamówienie {order_no} jest już w drodze",
            "title": "Zamówienie zostało wysłane",
            "greeting": "Dzień dobry,",
            "intro": "Z przyjemnością informujemy, że Twoje zamówienie zostało przekazane kurierowi.",
            "order": "Numer zamówienia", "carrier": "Przewoźnik", "tracking": "Numer przesyłki",
            "button": "Śledź swoją przesyłkę",
            "attachment": "W załączniku znajduje się lista pakowania zawierająca szczegóły wysłanych produktów.",
            "thanks": "Dziękujemy za zamówienie i życzymy udanego dnia!", "team": "Zespół Niedźwieccy",
        },
        "de": {
            "subject": "Ihre Bestellung {order_no} ist unterwegs",
            "title": "Ihre Bestellung wurde versandt",
            "greeting": "Guten Tag,",
            "intro": "Wir freuen uns, Ihnen mitzuteilen, dass Ihre Bestellung an den Paketdienst übergeben wurde.",
            "order": "Bestellnummer", "carrier": "Paketdienst", "tracking": "Sendungsnummer",
            "button": "Sendung verfolgen",
            "attachment": "Im Anhang finden Sie die Packliste mit den Details zu den versandten Produkten.",
            "thanks": "Vielen Dank für Ihre Bestellung. Wir wünschen Ihnen einen schönen Tag!", "team": "Ihr Niedźwieccy-Team",
        },
        "en": {
            "subject": "Your order {order_no} is on its way",
            "title": "Your order has been shipped",
            "greeting": "Hello,",
            "intro": "We are pleased to let you know that your order has been handed over to the courier.",
            "order": "Order number", "carrier": "Courier", "tracking": "Tracking number",
            "button": "Track your shipment",
            "attachment": "The attached packing list contains the details of the shipped products.",
            "thanks": "Thank you for your order and have a great day!", "team": "The Niedźwieccy Team",
        },
        "es": {
            "subject": "Tu pedido {order_no} ya está en camino",
            "title": "Tu pedido ha sido enviado",
            "greeting": "Buenos días,",
            "intro": "Nos complace informarte de que tu pedido ha sido entregado al transportista.",
            "order": "Número de pedido", "carrier": "Transportista", "tracking": "Número de seguimiento",
            "button": "Seguir el envío",
            "attachment": "En el archivo adjunto encontrarás la lista de embalaje con los detalles de los productos enviados.",
            "thanks": "¡Gracias por tu pedido y que tengas un buen día!", "team": "Equipo Niedźwieccy",
        },
        "it": {
            "subject": "Il tuo ordine {order_no} è in viaggio",
            "title": "Il tuo ordine è stato spedito",
            "greeting": "Buongiorno,",
            "intro": "Siamo lieti di informarti che il tuo ordine è stato affidato al corriere.",
            "order": "Numero ordine", "carrier": "Corriere", "tracking": "Numero di tracciamento",
            "button": "Traccia la spedizione",
            "attachment": "In allegato trovi la lista di imballaggio con i dettagli dei prodotti spediti.",
            "thanks": "Grazie per il tuo ordine e buona giornata!", "team": "Il team Niedźwieccy",
        },
    }
    copy = messages.get(language, messages["pl"])
    if norm(order.get("status")).lower() == "partially_shipped":
        partial_copy = {
            "pl": ("Częściowa wysyłka zamówienia {order_no}", "Zamówienie zostało wysłane częściowo", "Przekazaliśmy kurierowi część produktów z Twojego zamówienia. Pozostałe pozycje wyślemy osobno."),
            "de": ("Teillieferung der Bestellung {order_no}", "Ihre Bestellung wurde teilweise versandt", "Ein Teil der Produkte aus Ihrer Bestellung wurde an den Paketdienst übergeben. Die übrigen Positionen werden separat versandt."),
            "en": ("Partial shipment of order {order_no}", "Your order has been partially shipped", "Some products from your order have been handed over to the courier. The remaining items will be shipped separately."),
            "es": ("Envío parcial del pedido {order_no}", "Tu pedido se ha enviado parcialmente", "Hemos entregado al transportista una parte de los productos de tu pedido. Los artículos restantes se enviarán por separado."),
            "it": ("Spedizione parziale dell'ordine {order_no}", "Il tuo ordine è stato spedito parzialmente", "Abbiamo affidato al corriere una parte dei prodotti del tuo ordine. Gli articoli rimanenti saranno spediti separatamente."),
        }.get(language)
        if partial_copy:
            copy = dict(copy, subject=partial_copy[0], title=partial_copy[1], intro=partial_copy[2])
    tracking_url = carrier_tracking_url(carrier, tracking_no)
    carrier_names = {"inpost": "InPost", "dpd": "DPD", "fedex": "FedEx", "dhl": "DHL", "ups": "UPS"}
    carrier_name = carrier_names.get(norm(carrier).lower(), norm(carrier) or "—")
    safe_order = html.escape(str(order_no), quote=True)
    safe_tracking = html.escape(str(tracking_no), quote=True)
    safe_carrier = html.escape(carrier_name, quote=True)
    safe_tracking_url = html.escape(tracking_url, quote=True)
    subject_text = copy["subject"].format(order_no=order_no)
    html_body = (
        "<div style='margin:0;padding:28px 14px;background:#f3f6fb;font-family:Arial,sans-serif;color:#10203d'>"
        "<div style='max-width:620px;margin:0 auto;background:#ffffff;border:1px solid #e2e8f2;border-radius:18px;overflow:hidden'>"
        "<div style='padding:30px 34px 20px'>"
        f"<h1 style='margin:0 0 24px;font-size:26px;line-height:1.25'>{html.escape(copy['title'])}</h1>"
        f"<p style='margin:0 0 14px'>{html.escape(copy['greeting'])}</p>"
        f"<p style='margin:0 0 24px;line-height:1.6'>{html.escape(copy['intro'])}</p>"
        "<div style='padding:20px;background:#f7f9fd;border:1px solid #e3e9f3;border-radius:14px;line-height:1.8'>"
        f"<div><span style='color:#62708c'>{html.escape(copy['order'])}:</span> <b>{safe_order}</b></div>"
        f"<div><span style='color:#62708c'>{html.escape(copy['carrier'])}:</span> <b>{safe_carrier}</b></div>"
        f"<div><span style='color:#62708c'>{html.escape(copy['tracking'])}:</span> <b>{safe_tracking}</b></div>"
        "</div>"
        f"<p style='margin:24px 0'><a href='{safe_tracking_url}' style='display:inline-block;padding:13px 22px;background:#4f70eb;color:#ffffff;text-decoration:none;font-weight:bold;border-radius:10px'>{html.escape(copy['button'])}</a></p>"
        f"<p style='margin:0 0 24px;line-height:1.6;color:#52617c'>{html.escape(copy['attachment'])}</p>"
        f"<p style='margin:0;line-height:1.6'>{html.escape(copy['thanks'])}<br><br><b>{html.escape(copy['team'])}</b></p>"
        "</div></div></div>"
    )
    text_body = (
        f"{copy['greeting']}\n\n{copy['intro']}\n\n"
        f"{copy['order']}: {order_no}\n{copy['carrier']}: {carrier_name}\n"
        f"{copy['tracking']}: {tracking_no}\n{tracking_url}\n\n"
        f"{copy['attachment']}\n\n{copy['thanks']}\n{copy['team']}"
    )
    return send_email(
        recipient,
        subject_text,
        html_body,
        text_body,
        attachments=[packing_attachment],
    )


def _send_saved_order_confirmation(order_id: int, force: bool = False) -> dict:
    """Send a confirmation using the order saved by the warehouse backend."""
    c = conn()
    try:
        cur = c.cursor()
        cur.execute("SELECT * FROM orders WHERE id=? LIMIT 1", (order_id,))
        row = cur.fetchone()
        if not row:
            return {"ok": False, "error": "Nie znaleziono zamówienia"}
        order = dict(row)
        cur.execute("""
          SELECT
            oi.sku,
            oi.qty,
            COALESCE(p.name, fallback_product.name, '') AS name,
            COALESCE(oi.unit_net_price, price.net_price, 0) AS net_price,
            COALESCE(oi.currency, o.currency, 'PLN') AS currency
          FROM order_items oi
          JOIN orders o ON o.id=oi.order_id
          LEFT JOIN products p ON p.id = oi.product_id
          LEFT JOIN products fallback_product ON fallback_product.sku = oi.sku
          LEFT JOIN pricing price ON TRIM(LOWER(price.model)) = TRIM(LOWER(oi.sku))
          WHERE oi.order_id=?
          ORDER BY oi.id
        """, (order_id,))
        items = [dict(x) for x in cur.fetchall()]
    finally:
        c.close()

    try:
        profile = _client_profile_for_email(order.get("customer_email"))
        order["language"] = profile.get("language", "pl")
        order["currency"] = normalize_order_currency(order.get("currency") or profile.get("currency"))
    except Exception:
        order["language"] = "pl"
        order["currency"] = normalize_order_currency(order.get("currency"))

    try:
        admin_email = norm(email_config_summary().get("admin_email"))
    except Exception:
        admin_email = ""
    recipient = ", ".join([x for x in [norm(order.get("customer_email")), admin_email] if x])
    recipient_hash = hashlib.sha1(recipient.lower().encode("utf-8")).hexdigest()[:12] if recipient else "no-recipient"
    event_key = f"order_confirmation:{order_id}:{recipient_hash}"

    if not force and _email_event_already_ok(event_key):
        return {"ok": True, "duplicate": True, "skipped": True, "to": recipient}
    if not send_order_confirmation:
        result = {"ok": False, "skipped": True, "error": "Brak modułu email_module.py"}
    else:
        try:
            result = send_order_confirmation(order, items, admin_email=admin_email)
        except Exception as exc:
            result = {"ok": False, "error": str(exc)}
    _record_email_event(event_key, "order_confirmation", order_id, recipient, result)
    return result


def _safe_saved_order_confirmation(order_id: int, force: bool = False) -> dict:
    """Email is secondary: it must never turn a saved order into an HTTP 500."""
    try:
        result = (
            _send_saved_order_confirmation(order_id, force=True)
            if force
            else _send_saved_order_confirmation(order_id)
        )
        if not isinstance(result, dict):
            return {"ok": False, "pending_retry": True, "error": "Nieprawidłowa odpowiedź modułu e-mail"}
        if not result.get("ok"):
            result["pending_retry"] = True
        return result
    except Exception as exc:
        app.logger.exception("Zamówienie %s zapisane, ale potwierdzenie e-mail nie powiodło się", order_id)
        return {"ok": False, "pending_retry": True, "error": str(exc)}


def _authenticated_client_user() -> dict | None:
    auth = norm(request.headers.get("Authorization"))
    if not auth.lower().startswith("bearer "):
        return None
    token = auth.split(None, 1)[1].strip()
    if not token:
        return None
    cache_key = hashlib.sha256(token.encode("utf-8")).hexdigest()
    now_monotonic = time.monotonic()
    # Miniatury są pobierane równolegle. Blokada sprawia, że tylko pierwsze
    # żądanie weryfikuje token w Supabase Auth, a pozostałe korzystają przez
    # minutę z wyniku tej samej, już potwierdzonej sesji.
    with _client_auth_cache_lock:
        cached = _client_auth_cache.get(cache_key)
        if cached and cached[0] > now_monotonic:
            return dict(cached[1])
        req = urllib.request.Request(f"{SUPABASE_URL}/auth/v1/user", method="GET")
        api_key = SUPABASE_ANON_KEY or SUPABASE_SERVICE_ROLE_KEY
        if not api_key:
            app.logger.error("Weryfikacja klienta niemożliwa: brak SUPABASE_ANON_KEY")
            return None
        req.add_header("apikey", api_key)
        req.add_header("Authorization", f"Bearer {token}")
        try:
            with urllib.request.urlopen(req, timeout=20) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
            email = norm(payload.get("email")).lower()
            if not payload.get("id") or not email:
                return None
            metadata = payload.get("user_metadata") if isinstance(payload.get("user_metadata"), dict) else {}
            user = {
                "id": str(payload.get("id")),
                "email": email,
                "name": norm(metadata.get("full_name") or metadata.get("name")) or email.split("@")[0],
            }
            if len(_client_auth_cache) >= 500:
                _client_auth_cache.clear()
            _client_auth_cache[cache_key] = (time.monotonic() + CLIENT_AUTH_CACHE_TTL_SEC, user)
            return dict(user)
        except urllib.error.HTTPError as exc:
            app.logger.warning("Supabase odrzucił token klienta: HTTP %s", exc.code)
            return None
        except Exception as exc:
            app.logger.warning("Nie udało się zweryfikować tokenu klienta: %s", type(exc).__name__)
            return None


def _client_profile_for_email(email: str) -> dict:
    email = _email_key(email)
    fallback = {
        "id": None,
        "name": email.split("@")[0] if email else "",
        "address": "",
        "phone": "",
        "email": email,
        "nip": "",
        "language": "pl",
        "price_list": "pln",
        "currency": "PLN",
    }
    if not email or not supabase_enabled():
        return fallback

    params = {
        "select": "id,name,address,phone,email,nip,language,price_list",
        "email": f"ilike.{email}",
        "order": "id.desc",
        "limit": 1,
    }
    try:
        rows = supabase_request("/rest/v1/customers", params=params, timeout=20) or []
    except urllib.error.HTTPError as exc:
        # Bezpieczny fallback podczas wdrażania, zanim kolumna language zostanie dodana.
        if exc.code != 400:
            raise
        fallback_params = dict(params)
        fallback_params["select"] = "id,name,address,phone,email,nip,language"
        rows = supabase_request("/rest/v1/customers", params=fallback_params, timeout=20) or []

    if not isinstance(rows, list) or not rows:
        return fallback
    row = dict(rows[0])
    language = normalize_client_language(row.get("language"))
    price_list = price_list_for_language(language)
    # Reguła jest egzekwowana także dla istniejących kont. Aktualizacja jest
    # wykonywana podczas logowania, więc nie trzeba ręcznie poprawiać klientów.
    stored_price_list = normalize_client_price_list(row.get("price_list"))
    if row.get("id") is not None and stored_price_list != price_list:
        try:
            supabase_update_rows("customers", {"price_list": price_list}, {"id": row["id"]})
            c = conn()
            try:
                c.execute("UPDATE customers SET price_list=? WHERE id=?", (price_list, row["id"]))
                c.commit()
            finally:
                c.close()
        except Exception as exc:
            # Profil nadal zwracamy z poprawnym cennikiem; błąd zapisu jest
            # widoczny w logu Rendera i zostanie ponowiony przy kolejnym odczycie.
            app.logger.warning("Nie udało się automatycznie zapisać cennika klienta id=%s: %s", row.get("id"), exc)
    return {
        "id": row.get("id"),
        "name": norm(row.get("name")) or fallback["name"],
        "address": norm(row.get("address")),
        "phone": norm(row.get("phone")),
        "email": _email_key(row.get("email")) or email,
        "nip": norm(row.get("nip")),
        "language": language,
        "price_list": price_list,
        "currency": price_list_currency(price_list),
    }

def _order_by_idempotency_key(idempotency_key: str) -> dict | None:
    c = conn()
    try:
        cur = c.cursor()
        cur.execute("SELECT id, order_no, customer_email FROM orders WHERE idempotency_key=? LIMIT 1", (idempotency_key,))
        row = cur.fetchone()
        if row:
            return dict(row)
    finally:
        c.close()
    if not supabase_enabled():
        return None
    rows = supabase_request(
        "/rest/v1/orders",
        params={"select": "id,order_no,customer_email", "idempotency_key": f"eq.{idempotency_key}", "limit": 1},
    )
    return dict(rows[0]) if isinstance(rows, list) and rows else None


def _client_order_origin_allowed() -> bool:
    origin = norm(request.headers.get("Origin")).rstrip("/")
    return not origin or origin in CLIENT_ALLOWED_ORIGINS

def _client_order_items_local(c, order: dict) -> list[dict]:
    """Load order lines without assuming that an old SQLite file has new EUR columns.

    Prices saved on an order line always win.  Historical PLN orders fall back to
    the PLN price list, while historical EUR orders fall back to the imported EUR
    price list.  Keeping the optional columns out of SQL also makes this safe on
    Render during a rolling schema upgrade.
    """
    order_id = to_int(order.get("id"), 0)
    if order_id <= 0:
        return []

    cur = c.cursor()
    cur.execute("""
      SELECT
        oi.*,
        COALESCE(p.model, '') AS model,
        COALESCE(p.ean, '') AS ean,
        COALESCE(p.name, '') AS name,
        COALESCE(s.qty, 0) AS stock_qty,
        COALESCE(s.qty, 0) AS stock,
        COALESCE((
          SELECT SUM(ci.qty)
          FROM china_items ci
          JOIN china_packages cp ON cp.id=ci.package_id
          WHERE ci.product_id=oi.product_id
            AND cp.status IN ('ordered', 'shipped', 'problem')
        ), 0) AS in_delivery
      FROM order_items oi
      LEFT JOIN products p ON p.id=oi.product_id
      LEFT JOIN stock s ON s.product_id=oi.product_id
      WHERE oi.order_id=?
      ORDER BY oi.id
    """, (order_id,))
    items = [dict(row) for row in cur.fetchall()]
    if not items:
        return []

    cur.execute("SELECT model, net_price, gross_price FROM pricing")
    pricing = {}
    for row in cur.fetchall():
        key = norm(row["model"]).strip().lower()
        if key:
            pricing[key] = dict(row)

    eur_pricing = {}
    try:
        cur.execute("SELECT sku, price_eur, uvp_eur FROM pricing_eur")
        for row in cur.fetchall():
            key = norm(row["sku"]).strip().lower()
            if key:
                eur_pricing[key] = dict(row)
    except sqlite3.OperationalError:
        # The table is optional only for databases created before EU pricing.
        eur_pricing = {}

    order_currency = normalize_order_currency(order.get("currency"))
    for item in items:
        item_currency = normalize_order_currency(item.get("currency") or order_currency)
        sku_key = norm(item.get("sku")).strip().lower()
        model_key = norm(item.get("model")).strip().lower()

        snapshot_net = item.get("unit_net_price")
        snapshot_gross = item.get("unit_gross_price")
        snapshot_retail = item.get("unit_retail_price")

        if item_currency == "EUR":
            price_row = eur_pricing.get(sku_key) or {}
            fallback_net = money_float(price_row.get("price_eur"))
            fallback_gross = fallback_net
            fallback_retail = money_float(price_row.get("uvp_eur"))
        else:
            price_row = pricing.get(model_key) or pricing.get(sku_key) or {}
            fallback_net = money_float(price_row.get("net_price"))
            fallback_gross = money_float(price_row.get("gross_price"))
            fallback_retail = money_float(
                Decimal(str(fallback_net or 0)) * Decimal("1.45") * Decimal("1.23")
            )

        net_price = money_float(snapshot_net) if snapshot_net is not None else fallback_net
        gross_price = money_float(snapshot_gross) if snapshot_gross is not None else fallback_gross
        retail_price = money_float(snapshot_retail) if snapshot_retail is not None else fallback_retail
        qty = to_int(item.get("qty"), 0)

        item.update({
            "net_price": net_price,
            "gross_price": gross_price,
            "retail_price": retail_price,
            "currency": item_currency,
            "line_value_net": money_float(Decimal(str(net_price or 0)) * qty),
            "line_value_gross": money_float(Decimal(str(gross_price or 0)) * qty),
            "line_value_retail": money_float(Decimal(str(retail_price or 0)) * qty),
        })
    return items

def _api_order_lookup_impl():
    maybe_pull_shared_from_supabase(force=True)
    token = norm(request.args.get("token"))
    if not token:
        return jsonify(ok=False, error="Brak tokenu"), 400

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM orders WHERE order_no=? LIMIT 1", (token,))
    o = cur.fetchone()
    if not o:
        cur.execute("SELECT * FROM orders ORDER BY id DESC")
        all_orders = cur.fetchall()
        for row in all_orders:
            if canonical_order_no(row["id"], row["created_at"], row["order_no"]) == norm(token):
                o = row
                break
    if not o:
        c.close()
        return jsonify(ok=False, error="Nie znaleziono zamĂłwienia"), 404
    order = dict(o)
    if _email_key(order.get("customer_email")) != _email_key(g.client_user["email"]):
        c.close()
        return jsonify(ok=False, error="Brak dostępu"), 403

    items = _client_order_items_local(c, order)

    for it in items:
        it["in_delivery_available"] = int(it.get("in_delivery", 0) or 0)
        it["delivery_used"] = 0
        it["line_shortage"] = 0

    order_id = int(order["id"])
    if norm(order.get("status")).lower() in CURRENT_ORDER_STATUSES:
        status_ph = ",".join(["?"] * len(CURRENT_ORDER_STATUSES))
        cur.execute(f"SELECT id FROM orders WHERE LOWER(COALESCE(status,'')) IN ({status_ph}) AND id<=? ORDER BY id", (*sorted(CURRENT_ORDER_STATUSES), order_id))
        scoped_order_ids = [int(r["id"]) for r in cur.fetchall()]
        if scoped_order_ids:
            sph = ",".join(["?"] * len(scoped_order_ids))
            cur.execute(f"""
              SELECT oi.id, oi.order_id, oi.product_id, oi.qty
              FROM order_items oi
              WHERE oi.order_id IN ({sph})
              ORDER BY oi.order_id, oi.id
            """, tuple(scoped_order_ids))
            seq_items = cur.fetchall()

            product_ids = {int(r["product_id"]) for r in seq_items}
            pool_stock = {}
            pool_delivery = {}
            if product_ids:
                pph = ",".join(["?"] * len(product_ids))
                cur.execute(f"""
                  SELECT p.id AS product_id,
                         COALESCE(s.qty,0) AS stock_qty,
                         COALESCE((
                           SELECT SUM(ci.qty)
                           FROM china_items ci
                           JOIN china_packages cp ON cp.id=ci.package_id
                           WHERE ci.product_id=p.id
                             AND cp.status IN ('ordered', 'shipped', 'problem')
                         ),0) AS in_delivery_qty
                  FROM products p
                  LEFT JOIN stock s ON s.product_id=p.id
                  WHERE p.id IN ({pph})
                """, tuple(product_ids))
                for pr in cur.fetchall():
                    pid = int(pr["product_id"])
                    pool_stock[pid] = int(pr["stock_qty"])
                    pool_delivery[pid] = int(pr["in_delivery_qty"])

            item_alloc = {}
            for sr in seq_items:
                pid = int(sr["product_id"])
                need = int(sr["qty"])

                stock_now = pool_stock.get(pid, 0)
                from_stock = min(stock_now, need)
                pool_stock[pid] = stock_now - from_stock
                need_after_stock = need - from_stock

                delivery_now = pool_delivery.get(pid, 0)
                from_delivery = min(delivery_now, need_after_stock)
                pool_delivery[pid] = delivery_now - from_delivery
                shortage = need_after_stock - from_delivery

                if int(sr["order_id"]) == order_id:
                    item_alloc[int(sr["id"])] = {
                        "in_delivery_available": from_delivery,
                        "delivery_used": from_delivery,
                        "line_shortage": shortage,
                    }

            for it in items:
                al = item_alloc.get(int(it["id"]))
                if al:
                    it.update(al)
    c.close()

    invoiced_by_item = invoiced_qty_by_order_item_ids([int(it["id"]) for it in items])
    for it in items:
        ordered_qty = int(it.get("qty") or 0)
        invoiced_qty = int(invoiced_by_item.get(int(it["id"])) or 0)
        it["ordered_qty"] = ordered_qty
        it["invoiced_qty"] = invoiced_qty
        it["remaining_qty"] = max(0, ordered_qty - invoiced_qty)
        stock_qty = int(it.get("stock_qty") or 0)
        delivery_used = int(it.get("delivery_used") or 0)
        line_shortage = int(it.get("line_shortage") or 0)
        if order.get("status") in ("new", "packed", "confirmed", "in_delivery"):
            it["availability_label"] = "dostępne" if stock_qty >= ordered_qty else "10/20 dni"
        else:
            it["availability_label"] = "dostępne" if stock_qty >= ordered_qty else "10/20 dni"
        if ordered_qty > 0 and invoiced_qty >= ordered_qty:
            it["realization_label"] = "w całości"
        elif invoiced_qty > 0:
            it["realization_label"] = f"częściowo: {invoiced_qty}/{ordered_qty} szt."
        else:
            it["realization_label"] = "0 szt."

    total_net = round(sum(float(it.get("line_value_net") or 0) for it in items), 2)
    total_gross = round(sum(float(it.get("line_value_gross") or 0) for it in items), 2)
    total_retail = round(sum(float(it.get("line_value_retail") or 0) for it in items), 2)
    order_currency = normalize_order_currency(order.get("currency"))
    order_price_list = normalize_client_price_list(order.get("price_list"))

    return jsonify(
        ok=True,
        order={
            "id": order.get("id"),
            "order_no": order.get("order_no"),
            "status": order.get("status"),
            "tracking_no": order.get("tracking_no") or "",
            "carrier": order.get("carrier") or "",
            "packed_at": order.get("packed_at") or "",
            "shipped_at": order.get("shipped_at") or "",
            "created_at": order.get("created_at"),
            "customer_name": order.get("customer_name"),
            "customer_address": order.get("customer_address"),
            "customer_phone": order.get("customer_phone"),
            "customer_email": order.get("customer_email"),
            "note": order.get("note"),
            "qr_data_url": order.get("qr_data_url") or "",
            "warehouse_issued": int(order.get("warehouse_issued") or 0),
            "currency": order_currency,
            "price_list": order_price_list,
            "total_net": total_net,
            "total_gross": total_gross,
            "total_retail": total_retail,
        },
        items=items
    )

def _api_client_order_pdf_impl(order_id: int, retail_prices: bool = False):
    maybe_pull_shared_from_supabase(force=True)
    email = _email_key(g.client_user.get("email"))

    c = conn()
    try:
        cur = c.cursor()
        cur.execute("SELECT * FROM orders WHERE id=? LIMIT 1", (order_id,))
        row = cur.fetchone()
        if not row:
            return jsonify(ok=False, error="Nie znaleziono zamówienia"), 404
        order = dict(row)
        if _email_key(order.get("customer_email")) != email:
            return jsonify(ok=False, error="Brak dostępu"), 403

        items = _client_order_items_local(c, order)
    finally:
        c.close()

    try:
        language = _client_profile_for_email(email).get("language", "pl")
    except Exception:
        language = "pl"
    pdf_buffer, filename = generate_client_order_pdf(
        order, items, language, retail_prices=retail_prices
    )
    response = send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
        max_age=0,
    )
    response.headers["Cache-Control"] = "no-store"
    return response

def load_invoice_with_meta(invoice_id: int):
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT i.*, COALESCE(m.pdf_path,'') AS pdf_path, COALESCE(m.sent_to_client,0) AS sent_to_client,
             COALESCE(m.paid,0) AS paid, COALESCE(m.paid_at,'') AS paid_at,
             COALESCE(m.invoice_items_json,'') AS invoice_items_json,
             COALESCE(o.currency,'PLN') AS order_currency
      FROM invoices i
      LEFT JOIN orders o ON o.id=i.order_id
      LEFT JOIN invoice_meta m ON m.invoice_id = i.id
      WHERE i.id=?
      LIMIT 1
    """, (invoice_id,))
    row = cur.fetchone()
    c.close()
    return dict(row) if row else None

def invoice_meta_payload(invoice_row: dict):
    buyer_address = "\n".join([x for x in [
        invoice_row.get("buyer_street") or "",
        f"{invoice_row.get('buyer_post_code') or ''} {invoice_row.get('buyer_city') or ''}".strip()
    ] if x]).strip()
    ksef_number = invoice_row.get("ksef_number") or ""
    if not ksef_number and invoice_row.get("id"):
        try:
            ksef_number = load_ksef_doc(int(invoice_row.get("id") or 0)).get("ksef_number") or ""
        except Exception:
            ksef_number = ""

    return {
        "invoice_no": invoice_row.get("invoice_no") or "",
        "place": "Kotuszów",
        "issue_date": invoice_row.get("issue_date") or app_now().strftime("%Y-%m-%d"),
        "sell_date": invoice_row.get("sell_date") or app_now().strftime("%Y-%m-%d"),
        "payment_type": invoice_row.get("payment_type") or "przelew",
        "payment_to": invoice_row.get("payment_to") or "",
        "invoice_type": invoice_row.get("invoice_type") or "",
        "currency": invoice_row.get("currency") or invoice_row.get("order_currency") or "PLN",
        "paid": int(invoice_row.get("paid") or 0),
        "paid_at": invoice_row.get("paid_at") or "",
        "buyer_name": invoice_row.get("buyer_name") or "",
        "buyer_tax_no": invoice_row.get("buyer_tax_no") or "",
        "buyer_address": buyer_address,
        "buyer_country": invoice_row.get("buyer_country") or "PL",
        "buyer_email": invoice_row.get("buyer_email") or "",
        "buyer_phone": invoice_row.get("buyer_phone") or "",
        "discount_percent": "0",
        "ksef_number": ksef_number,
    }

def invoice_items_from_saved_json(invoice_id: int):
    meta = load_invoice_meta(invoice_id) or {}
    raw = meta.get("invoice_items_json") or ""
    if raw:
        try:
            data = json.loads(raw)
            if isinstance(data, list) and data:
                return data
        except Exception:
            pass

    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT oi.*, p.model, p.name,
             COALESCE(pr.net_price, 0) AS net_price,
             COALESCE(pr.gross_price, 0) AS gross_price,
             (oi.qty * COALESCE(pr.net_price, 0)) AS line_value_net,
             (oi.qty * COALESCE(pr.gross_price, 0)) AS line_value_gross
      FROM order_items oi
      JOIN products p ON p.id=oi.product_id
      LEFT JOIN pricing pr ON (TRIM(LOWER(pr.model)) = TRIM(LOWER(p.model)) OR TRIM(LOWER(pr.model)) = TRIM(LOWER(p.sku)))
      WHERE oi.order_id=(SELECT order_id FROM invoices WHERE id=?)
      ORDER BY oi.id
    """, (invoice_id,))
    items = [dict(r) for r in cur.fetchall()]
    c.close()
    return items


def load_company_profile() -> dict:
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM company_profile WHERE id=1")
    row = cur.fetchone()
    c.close()
    return dict(row) if row else {}


def ksef_dir() -> str:
    path = os.path.join(DATA_DIR, "ksef")
    os.makedirs(path, exist_ok=True)
    return path


def ksef_xml_path(invoice_id: int, invoice_no: str) -> str:
    return os.path.join(ksef_dir(), f"{int(invoice_id)}_{xml_filename(invoice_no)}")


def ksef_schema_path() -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "fa3_schemat.xsd")


def load_ksef_doc(invoice_id: int) -> dict:
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM ksef_documents WHERE invoice_id=?", (invoice_id,))
    row = cur.fetchone()
    c.close()
    return dict(row) if row else {}


def upsert_ksef_doc(invoice_id: int, status: str, xml_path: str = "", last_error: str = "", ksef_number: str = ""):
    current = load_ksef_doc(invoice_id)
    sent_at = current.get("sent_at", "")
    if status == "sent" and not sent_at:
        sent_at = now_iso()
    c = conn()
    cur = c.cursor()
    cur.execute("""
      INSERT INTO ksef_documents(invoice_id, status, ksef_number, xml_path, last_error, validated_at, sent_at, updated_at)
      VALUES(?,?,?,?,?,?,?,?)
      ON CONFLICT(invoice_id) DO UPDATE SET
        status=excluded.status,
        ksef_number=COALESCE(NULLIF(excluded.ksef_number,''), ksef_documents.ksef_number),
        xml_path=COALESCE(NULLIF(excluded.xml_path,''), ksef_documents.xml_path),
        last_error=excluded.last_error,
        validated_at=excluded.validated_at,
        sent_at=COALESCE(ksef_documents.sent_at, excluded.sent_at),
        updated_at=excluded.updated_at
    """, (
        invoice_id,
        status,
        ksef_number or current.get("ksef_number", ""),
        xml_path or current.get("xml_path", ""),
        last_error or "",
        now_iso(),
        sent_at,
        now_iso(),
    ))
    c.commit()
    c.close()
    try:
        sync_local_rows_to_supabase("ksef_documents", "invoice_id", [invoice_id])
    except Exception:
        pass


def regenerate_invoice_pdf_after_ksef_send(invoice_id: int, ksef_number: str) -> bool:
    inv = load_invoice_with_meta(invoice_id)
    if not inv:
        return False
    items = invoice_items_from_saved_json(invoice_id)
    if not items:
        return False

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM orders WHERE id=?", (inv.get("order_id"),))
    order_row = cur.fetchone()
    c.close()

    meta_payload = invoice_meta_payload(inv)
    meta_payload["ksef_number"] = ksef_number
    pdf_path, total_net, total_gross = generate_order_invoice_pdf(order_row, items, meta_payload)
    packing_pdf_path = generate_invoice_packing_list_pdf(order_row, items, meta_payload, pdf_path)
    stored_pdf_path = upload_invoice_pdfs_to_supabase(invoice_id, inv.get("invoice_no") or f"FV_{invoice_id}", pdf_path, packing_pdf_path)

    current_meta = load_invoice_meta(invoice_id) or {}
    upsert_invoice_meta(
        invoice_id,
        stored_pdf_path,
        current_meta.get("invoice_items_json") or json.dumps(items, ensure_ascii=False),
        sent_to_client=int(current_meta.get("sent_to_client") or 0),
        seen_by_client=int(current_meta.get("seen_by_client") or 0),
        seen_at=current_meta.get("seen_at"),
        payment_reminder=int(current_meta.get("payment_reminder") or 0),
        paid=int(current_meta.get("paid") or 0),
        paid_at=current_meta.get("paid_at"),
    )

    if supabase_enabled():
        try:
            sync_local_rows_to_supabase("invoices", "id", [invoice_id])
        except Exception:
            pass
        try:
            sync_invoice_meta_to_supabase(invoice_id)
        except Exception:
            pass
    return True


def validate_ksef_invoice(invoice, company, items):
    """Select the validator by tax type without persisting any change."""
    invoice_type = resolve_invoice_type(invoice, items)
    payload = dict(invoice)
    payload["invoice_type"] = invoice_type
    if invoice_type == "domestic":
        return ksef_domestic.validate(
            payload, company, items, validator=_legacy_validate_ksef_invoice
        )
    return ksef_foreign.validate(payload, company, items)


def build_ksef_draft_xml(invoice, company, items):
    """Pure FA(3) dispatcher for domestic, WDT and export documents."""
    invoice_type = resolve_invoice_type(invoice, items)
    payload = dict(invoice)
    payload["invoice_type"] = invoice_type
    if invoice_type == "domestic":
        return ksef_domestic.generate(
            payload, company, items, generator=_legacy_build_ksef_draft_xml
        )
    if invoice_type in {"wdt", "export"}:
        return ksef_foreign.generate(payload, company, items)
    raise ValueError("Nieobsługiwany typ faktury")


def build_invoice_ksef_payload(invoice_id: int):
    inv = load_invoice_with_meta(invoice_id)
    if not inv:
        return None, {}, [], ["Nie znaleziono faktury."]
    company = load_company_profile()
    items = invoice_items_from_saved_json(invoice_id)
    inv["currency"] = inv.get("currency") or (items[0].get("currency") if items else None) or inv.get("order_currency") or "PLN"
    inv["invoice_type"] = resolve_invoice_type(inv, items)
    problems = validate_ksef_invoice(inv, company, items)
    return inv, company, items, problems

def _redirect_after_invoice_action(default_endpoint="invoices"):
    target = norm(request.values.get("next")) or request.referrer or url_for(default_endpoint)
    return redirect(target)


def _invoice_json_order_ids(raw_items_json: str) -> set[int]:
    """Odczytuje powiązania wielu zamówień ze starszego zapisu faktury."""
    try:
        items = json.loads(raw_items_json or "[]")
    except Exception:
        return set()
    if not isinstance(items, list):
        return set()
    order_ids = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        try:
            order_id = int(item.get("source_order_id") or item.get("order_id") or 0)
        except (TypeError, ValueError):
            order_id = 0
        if order_id > 0:
            order_ids.add(order_id)
    return order_ids


def _invoice_source_order_ids(cur, invoice_id: int) -> list[int]:
    """Zwraca wszystkie zamówienia rozliczane daną fakturą."""
    cur.execute("""
      SELECT DISTINCT order_id
      FROM invoice_allocations
      WHERE invoice_id=? AND order_id IS NOT NULL
    """, (invoice_id,))
    order_ids = {int(row["order_id"]) for row in cur.fetchall() if row["order_id"]}

    # Starsze faktury mogły powstać przed tabelą invoice_allocations.
    cur.execute("SELECT order_id FROM invoices WHERE id=?", (invoice_id,))
    invoice_row = cur.fetchone()
    if invoice_row and invoice_row["order_id"]:
        order_ids.add(int(invoice_row["order_id"]))

    # Stare faktury zbiorcze przechowywały dodatkowe zamówienia wyłącznie
    # w JSON-ie pozycji; invoices.order_id wskazywało tylko pierwsze z nich.
    cur.execute("SELECT invoice_items_json FROM invoice_meta WHERE invoice_id=?", (invoice_id,))
    meta_row = cur.fetchone()
    if meta_row:
        order_ids.update(_invoice_json_order_ids(meta_row["invoice_items_json"]))
    return sorted(order_ids)


def _order_invoice_ids(cur, order_id: int) -> list[int]:
    """Zwraca faktury rozliczające zamówienie, także rekordy historyczne."""
    cur.execute("""
      SELECT DISTINCT invoice_id
      FROM invoice_allocations
      WHERE order_id=? AND invoice_id IS NOT NULL
    """, (order_id,))
    invoice_ids = {int(row["invoice_id"]) for row in cur.fetchall() if row["invoice_id"]}
    cur.execute("SELECT id FROM invoices WHERE order_id=?", (order_id,))
    invoice_ids.update(int(row["id"]) for row in cur.fetchall())
    cur.execute("SELECT invoice_id, invoice_items_json FROM invoice_meta WHERE TRIM(COALESCE(invoice_items_json,''))<>''")
    for row in cur.fetchall():
        if order_id in _invoice_json_order_ids(row["invoice_items_json"]):
            invoice_ids.add(int(row["invoice_id"]))
    return sorted(invoice_ids)


def _all_order_invoices_paid(cur, order_id: int) -> bool:
    invoice_ids = _order_invoice_ids(cur, order_id)
    if not invoice_ids:
        return False
    placeholders = ",".join(["?"] * len(invoice_ids))
    cur.execute(f"""
      SELECT i.id AS invoice_id, m.invoice_id AS meta_invoice_id,
             COALESCE(m.sent_to_client,0) AS sent_to_client,
             COALESCE(m.paid,0) AS paid
      FROM invoices i
      LEFT JOIN invoice_meta m ON m.invoice_id=i.id
      WHERE i.id IN ({placeholders})
    """, tuple(invoice_ids))
    rows = [dict(row) for row in cur.fetchall()]
    # Niewysłany szkic starej faktury nie jest należnością klienta i nie może
    # blokować zamknięcia zamówienia. Rekordy bez invoice_meta są historycznie
    # widoczne, więc nadal wymagają jawnego oznaczenia płatności.
    payable = [row for row in rows if row["meta_invoice_id"] is None or int(row["sent_to_client"] or 0) == 1 or int(row["paid"] or 0) == 1]
    return bool(payable) and all(int(row["paid"] or 0) == 1 for row in payable)


def _order_fully_invoiced_for_payment(cur, order_id: int) -> bool:
    """Obsługuje również pełne faktury sprzed tabeli invoice_allocations.

    Jeżeli zamówienie ma choć jedną alokację, obowiązuje dokładne sprawdzenie
    ilości pozycji. Brak alokacji jest uznawany za historyczną pełną fakturę
    tylko wtedy, gdy faktura wskazuje zamówienie bezpośrednio przez order_id.
    """
    cur.execute("SELECT 1 FROM invoice_allocations WHERE order_id=? LIMIT 1", (order_id,))
    if cur.fetchone():
        return order_fully_invoiced(cur, order_id)
    cur.execute("SELECT 1 FROM invoices WHERE order_id=? LIMIT 1", (order_id,))
    if cur.fetchone() is not None:
        return True
    cur.execute("SELECT invoice_items_json FROM invoice_meta WHERE TRIM(COALESCE(invoice_items_json,''))<>''")
    return any(order_id in _invoice_json_order_ids(row["invoice_items_json"]) for row in cur.fetchall())


def reconcile_paid_order_statuses():
    """Zamyka tylko w pełni zafakturowane i opłacone zamówienia.

    Samo wcześniejsze wydanie magazynowe nie dowodzi, że wysłano całość.
    Dotyczy to zwłaszcza zamówień realizowanych kilkoma paczkami.
    """
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT invoice_id FROM invoice_meta WHERE COALESCE(paid,0)=1")
    paid_invoice_ids = [int(row["invoice_id"]) for row in cur.fetchall()]
    order_ids = sorted({
        order_id
        for invoice_id in paid_invoice_ids
        for order_id in _invoice_source_order_ids(cur, invoice_id)
    })
    changed_order_ids = []
    for order_id in order_ids:
        cur.execute("SELECT status, warehouse_issued FROM orders WHERE id=?", (order_id,))
        order_row = cur.fetchone()
        if not order_row:
            continue
        current_status = norm(order_row["status"]).lower()
        if current_status in {"completed", "cancelled"}:
            continue
        fully_invoiced = _order_fully_invoiced_for_payment(cur, order_id)
        if fully_invoiced and _all_order_invoices_paid(cur, order_id):
            cur.execute("UPDATE orders SET status='completed' WHERE id=?", (order_id,))
            changed_order_ids.append(order_id)
    c.commit()
    c.close()

    if changed_order_ids and supabase_enabled():
        try:
            sync_local_rows_to_supabase("orders", "id", changed_order_ids)
        except Exception:
            app.logger.exception("Nie udało się zsynchronizować statusów opłaconych zamówień")
    return changed_order_ids


def reconcile_legacy_orders_by_age(days: int = 14):
    """Wyłączona reguła historyczna.

    Wiek zamówienia nie świadczy o jego realizacji. W szczególności zamówień
    wysłanych częściowo nie wolno automatycznie zamykać po 14 dniach.
    """
    return []


def _set_invoice_payment_state(invoice_id: int, *, reminder: int | None = None, paid: int | None = None):
    meta = load_invoice_meta(invoice_id) or {}
    pdf_path = meta.get("pdf_path", "")
    items_json = meta.get("invoice_items_json", "")
    sent_to_client = int(meta.get("sent_to_client") or 0)
    seen_by_client = int(meta.get("seen_by_client") or 0)
    seen_at = meta.get("seen_at")
    current_reminder = int(meta.get("payment_reminder") or 0)
    current_paid = int(meta.get("paid") or 0)
    current_paid_at = meta.get("paid_at")

    next_paid = current_paid if paid is None else int(paid)
    next_reminder = current_reminder if reminder is None else int(reminder)
    next_paid_at = current_paid_at
    if next_paid:
        next_reminder = 0
        state_changed_at = now_iso()
        next_paid_at = state_changed_at
        # Opłacona faktura nie powinna nadal oczekiwać na potwierdzenie
        # klienta. Zachowujemy jedną prawdę w invoice_meta/Supabase.
        if not seen_by_client:
            seen_by_client = 1
            seen_at = state_changed_at
    elif paid == 0:
        next_paid_at = None

    upsert_invoice_meta(
        invoice_id,
        pdf_path,
        items_json,
        sent_to_client=sent_to_client,
        seen_by_client=seen_by_client,
        seen_at=seen_at,
        payment_reminder=next_reminder,
        paid=next_paid,
        paid_at=next_paid_at
    )

    changed_order_ids = []
    c = conn()
    cur = c.cursor()
    for order_id in _invoice_source_order_ids(cur, invoice_id):
        cur.execute("SELECT status, tracking_no, packed_at FROM orders WHERE id=?", (order_id,))
        order_row = cur.fetchone()
        if not order_row:
            continue

        current_status = norm(order_row["status"]).lower()
        fully_invoiced = _order_fully_invoiced_for_payment(cur, order_id)

        # Zakończenie następuje dopiero po pełnym zafakturowaniu zamówienia
        # i opłaceniu wszystkich faktur, które je rozliczają.
        if next_paid and fully_invoiced and _all_order_invoices_paid(cur, order_id):
            if current_status not in {"completed", "cancelled"}:
                cur.execute("UPDATE orders SET status='completed' WHERE id=?", (order_id,))
                changed_order_ids.append(order_id)
        elif paid == 0 and current_status in {"completed", "issued"}:
            if norm(order_row["tracking_no"]):
                next_status = "shipped" if fully_invoiced else "partially_shipped"
            elif fully_invoiced:
                next_status = "packed"
            else:
                next_status = "packed_partial" if order_row["packed_at"] else "confirmed"
            cur.execute("UPDATE orders SET status=? WHERE id=?", (next_status, order_id))
            changed_order_ids.append(order_id)

    c.commit()
    c.close()

    if supabase_enabled():
        try:
            sync_invoice_meta_to_supabase(invoice_id)
        except Exception:
            pass
        if changed_order_ids:
            try:
                sync_local_rows_to_supabase("orders", "id", changed_order_ids)
            except Exception:
                pass

def _delete_invoice_everywhere(invoice_id: int):
    inv = load_invoice_with_meta(invoice_id)
    if not inv:
        abort(404)

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT DISTINCT order_id FROM invoice_allocations WHERE invoice_id=?", (invoice_id,))
    touched_order_ids = [int(r["order_id"]) for r in cur.fetchall()]
    meta_items_raw = inv.get("invoice_items_json") or ""
    if meta_items_raw:
        try:
            meta_items = json.loads(meta_items_raw)
            if isinstance(meta_items, list):
                for it in meta_items:
                    oid = int(it.get("source_order_id") or it.get("order_id") or 0)
                    if oid and oid not in touched_order_ids:
                        touched_order_ids.append(oid)
        except Exception:
            pass
    if int(inv.get("order_id") or 0) and int(inv.get("order_id") or 0) not in touched_order_ids:
        touched_order_ids.append(int(inv.get("order_id") or 0))
    c.close()

    ok_pdf, abs_path = invoice_pdf_exists(inv.get("pdf_path", ""), inv.get("invoice_no", ""))
    try:
        pack_path = packing_list_pdf_path_for_invoice(abs_path if ok_pdf else "", inv.get("invoice_no", ""))
        if pack_path and os.path.exists(pack_path):
            os.remove(pack_path)
        if ok_pdf and abs_path and os.path.exists(abs_path):
            os.remove(abs_path)
    except Exception:
        pass

    c = conn()
    cur = c.cursor()
    cur.execute("DELETE FROM invoice_allocations WHERE invoice_id=?", (invoice_id,))
    cur.execute("DELETE FROM invoice_meta WHERE invoice_id=?", (invoice_id,))
    cur.execute("DELETE FROM ksef_documents WHERE invoice_id=?", (invoice_id,))
    cur.execute("DELETE FROM invoices WHERE id=?", (invoice_id,))
    c.commit()
    c.close()

    changed_order_ids, changed_product_ids = reconcile_orders_after_invoice_change(touched_order_ids)

    if supabase_enabled():
        try:
            supabase_delete_rows("invoice_allocations", {"invoice_id": invoice_id})
        except Exception:
            pass
        try:
            supabase_delete_rows("invoice_meta", {"invoice_id": invoice_id})
        except Exception:
            pass
        try:
            supabase_delete_rows("ksef_documents", {"invoice_id": invoice_id})
        except Exception:
            pass
        try:
            supabase_delete_rows("invoices", {"id": invoice_id})
        except Exception:
            pass
        if changed_order_ids:
            try:
                sync_local_rows_to_supabase("orders", "id", changed_order_ids)
            except Exception:
                pass
        if changed_product_ids:
            try:
                sync_local_rows_to_supabase("stock", "product_id", changed_product_ids)
            except Exception:
                pass

    return inv

def _invoice_email_context(invoice_id: int):
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT
        i.*,
        o.order_no AS order_no,
        o.customer_email AS customer_email,
        o.customer_name AS customer_name,
        COALESCE(cu.language, 'pl') AS language,
        COALESCE(o.currency, 'PLN') AS currency,
        m.pdf_path AS pdf_path,
        m.payment_reminder AS payment_reminder,
        m.paid AS paid,
        m.seen_by_client AS seen_by_client,
        m.seen_at AS seen_at
      FROM invoices i
      LEFT JOIN orders o ON o.id = i.order_id
      LEFT JOIN customers cu ON cu.id = o.customer_id
      LEFT JOIN invoice_meta m ON m.invoice_id = i.id
      WHERE i.id=?
      LIMIT 1
    """, (invoice_id,))
    row = cur.fetchone()
    c.close()
    if not row:
        abort(404)

    invoice = dict(row)
    # E-mail z fakturą nie kieruje klienta do panelu. Faktura jest wysyłana jako załącznik PDF.
    return invoice, ""


def send_automatic_payment_reminders(reference_time=None) -> dict:
    """Wysyła jedno przypomnienie od dnia następującego po terminie płatności.

    Harmonogram uruchamia tę funkcję o 12:00 czasu Europe/Warsaw. Znacznik
    payment_reminder jest zapisywany dopiero po udanej wysyłce, dzięki czemu
    ponowne uruchomienie jest bezpieczne i nie dubluje wiadomości.
    """
    now = reference_time or app_now()
    overdue_before_or_on = (now.date() - timedelta(days=1)).isoformat()
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT i.id
      FROM invoices i
      LEFT JOIN invoice_meta m ON m.invoice_id=i.id
      WHERE DATE(SUBSTR(TRIM(COALESCE(i.payment_to,'')),1,10)) <= DATE(?)
        AND COALESCE(m.paid,0)=0
        AND COALESCE(m.payment_reminder,0)=0
      ORDER BY i.id
    """, (overdue_before_or_on,))
    invoice_ids = [int(row["id"]) for row in cur.fetchall()]
    c.close()

    sent_ids = []
    failed = []
    for invoice_id in invoice_ids:
        try:
            if not send_payment_reminder:
                raise RuntimeError("Moduł wysyłki przypomnień nie jest dostępny")
            invoice_row, pdf_url = _invoice_email_context(invoice_id)
            result = send_payment_reminder(invoice_row, pdf_url=pdf_url)
            if not result.get("ok"):
                raise RuntimeError(norm(result.get("error")) or "Wysyłka nie powiodła się")
            _set_invoice_payment_state(invoice_id, reminder=1, paid=None)
            sent_ids.append(invoice_id)
        except Exception as exc:
            failed.append({"invoice_id": invoice_id, "error": str(exc)[:300]})
            app.logger.exception("Automatyczne przypomnienie nie zostało wysłane dla faktury %s", invoice_id)
    return {
        "ok": not failed,
        "overdue_before_or_on": overdue_before_or_on,
        "eligible": len(invoice_ids),
        "sent": len(sent_ids),
        "sent_ids": sent_ids,
        "failed": failed,
    }


def _send_invoice_to_client(invoice_id: int) -> tuple[int, bool, str]:
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT i.id, i.order_id, i.buyer_email, i.invoice_no, o.customer_email
      FROM invoices i
      LEFT JOIN orders o ON o.id = i.order_id
      WHERE i.id=?
      LIMIT 1
    """, (invoice_id,))
    row = cur.fetchone()
    if not row:
        c.close()
        abort(404)

    buyer_email = _email_key(row["buyer_email"])
    order_email = _email_key(row["customer_email"])
    if not buyer_email and order_email:
        cur.execute("UPDATE invoices SET buyer_email=? WHERE id=?", (order_email, invoice_id))
        c.commit()
        if supabase_enabled():
            try:
                supabase_update_rows("invoices", {"buyer_email": order_email}, {"id": invoice_id})
            except Exception:
                pass
    c.close()

    meta = load_invoice_meta(invoice_id) or {}
    pdf_path = norm(meta.get("pdf_path"))
    stored_pdf_path = pdf_path

    if not parse_supabase_storage_ref(stored_pdf_path):
        local_pdf_path = ""
        if pdf_path:
            candidate = pdf_path if os.path.isabs(pdf_path) else invoice_pdf_abspath(pdf_path)
            if os.path.exists(candidate):
                local_pdf_path = candidate

        if not local_pdf_path:
            fallback = find_invoice_pdf_fallback(row["invoice_no"])
            if fallback:
                local_pdf_path = fallback

        items = invoice_items_from_saved_json(invoice_id)
        if not local_pdf_path and items:
            c = conn()
            cur = c.cursor()
            cur.execute("SELECT * FROM orders WHERE id=?", (row["order_id"],))
            order_row = cur.fetchone()
            c.close()
            if order_row:
                meta_payload = invoice_meta_payload(dict(row))
                local_pdf_path, _total_net, _total_gross = generate_order_invoice_pdf(order_row, items, meta_payload)

        if local_pdf_path:
            packing_pdf_path = ""
            if items:
                pack_candidate = packing_list_pdf_path_for_invoice(local_pdf_path, row["invoice_no"])
                if os.path.exists(pack_candidate):
                    packing_pdf_path = pack_candidate
                else:
                    c = conn()
                    cur = c.cursor()
                    cur.execute("SELECT * FROM orders WHERE id=?", (row["order_id"],))
                    order_row = cur.fetchone()
                    c.close()
                    if order_row:
                        packing_pdf_path = generate_invoice_packing_list_pdf(order_row, items, invoice_meta_payload(dict(row)), local_pdf_path)
            stored_pdf_path = upload_invoice_pdfs_to_supabase(invoice_id, row["invoice_no"], local_pdf_path, packing_pdf_path)

    upsert_invoice_meta(invoice_id, stored_pdf_path, meta.get("invoice_items_json",""), sent_to_client=1, seen_by_client=0, seen_at=None)

    if supabase_enabled():
        try:
            sync_invoice_meta_to_supabase(invoice_id)
        except Exception:
            pass

    email_ok = False
    email_error = ""
    try:
        if not send_invoice_available:
            email_error = "Brak modułu wysyłki e-mail."
        else:
            invoice_row, pdf_url = _invoice_email_context(invoice_id)
            pdf_attachment = None
            try:
                attachment_bytes = b""
                attachment_name = safe_filename(row["invoice_no"] or f"invoice_{invoice_id}") + ".pdf"
                if parse_supabase_storage_ref(stored_pdf_path):
                    attachment_bytes, downloaded_name = supabase_storage_download_bytes(stored_pdf_path)
                    if downloaded_name:
                        attachment_name = downloaded_name
                else:
                    candidate = stored_pdf_path if os.path.isabs(stored_pdf_path) else invoice_pdf_abspath(stored_pdf_path)
                    if candidate and os.path.exists(candidate):
                        with open(candidate, "rb") as fh:
                            attachment_bytes = fh.read()
                if attachment_bytes:
                    pdf_attachment = {"filename": attachment_name, "content": attachment_bytes}
            except Exception:
                app.logger.exception("Nie udało się przygotować załącznika PDF dla faktury %s", invoice_id)
            if not pdf_attachment:
                raise RuntimeError("Nie udało się przygotować pliku PDF faktury do wysyłki.")
            result = send_invoice_available(invoice_row, pdf_url=pdf_url, pdf_attachment=pdf_attachment) or {}
            email_ok = bool(result.get("ok"))
            if not email_ok:
                email_error = norm(result.get("error")) or "Usługa pocztowa odrzuciła wiadomość."
    except Exception as exc:
        email_error = str(exc) or type(exc).__name__
        app.logger.exception("Nie udało się wysłać e-maila z fakturą %s", invoice_id)

    if not email_ok:
        app.logger.error("Nie wysłano e-maila z fakturą %s: %s", invoice_id, email_error)

    return int(row["order_id"] or 0), email_ok, email_error

# -------------------------
# CHINA (prosty start)
# -------------------------
# =========================
# RUN
# =========================

# Domain routes are registered after infrastructure and shared helpers exist.
from routes import admin as routes_admin, china as routes_china, customers as routes_customers
from routes import inventory as routes_inventory, invoices as routes_invoices
from routes import orders as routes_orders, shipping as routes_shipping
# Legacy code replaces the original /searches handler with client_searches_v2.
# Temporarily release the endpoint so its URL rule can be registered in the
# customer module, then restore the same v2 handler as before the extraction.
app.view_functions.pop("client_searches", None)
for _routes_module in (routes_admin, routes_customers, routes_orders, routes_inventory, routes_shipping, routes_invoices, routes_china):
    globals().update(_routes_module.register_routes(globals()))
if "client_searches_v2" in globals():
    app.view_functions["client_searches"] = client_searches_v2

_DOMAIN_ROUTE_MODULES = (routes_admin, routes_customers, routes_orders, routes_inventory, routes_shipping, routes_invoices, routes_china)

@app.before_request
def _refresh_domain_route_context():
    # Moduły otrzymują pełny kontekst raz podczas rejestracji. Ponowne kopiowanie
    # kilku tysięcy nazw do siedmiu słowników przy każdym requestcie nie jest
    # potrzebne w produkcji. Zachowujemy je wyłącznie dla izolowanych testów,
    # które celowo podmieniają zależności po imporcie aplikacji.
    if os.environ.get("PYTEST_CURRENT_TEST"):
        context = globals()
        for module in _DOMAIN_ROUTE_MODULES:
            module.__dict__.update(context)

if __name__ == "__main__":
    # debug=True moĹĽesz zostawiÄ‡ na czas budowy
    app.run(host="0.0.0.0", port=5000, debug=True)
