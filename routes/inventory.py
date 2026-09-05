"""Mechanically extracted Flask routes; business logic is unchanged."""

def register_routes(context):
    globals().update(context)


    @app.get("/pricing")
    def pricing():
        maybe_pull_shared_from_supabase()
        q = norm(request.args.get("q"))
        eur_imported = max(0, int(to_float(request.args.get("eur_imported"), 0)))
        eur_import_error = norm(request.args.get("eur_import_error"))
        eur_local_saved = max(0, int(to_float(request.args.get("eur_local_saved"), 0)))
        c = conn()
        cur = c.cursor()
        if q:
            like = f"%{q}%"
            cur.execute("SELECT * FROM pricing WHERE model LIKE ? ORDER BY model LIMIT 2000", (like,))
        else:
            cur.execute("SELECT * FROM pricing ORDER BY model LIMIT 2000")
        rows = cur.fetchall()
        if q:
            like = f"%{q}%"
            cur.execute(
                "SELECT * FROM pricing_eur WHERE sku LIKE ? OR ean LIKE ? ORDER BY sku LIMIT 2000",
                (like, like),
            )
        else:
            cur.execute("SELECT * FROM pricing_eur ORDER BY sku LIMIT 2000")
        eur_rows = cur.fetchall()
        c.close()

        tpl = r"""
        {% extends "base.html" %}
        {% block content %}
          {% if eur_imported %}
            <div class="card" style="border-color:#9ad9c4;background:#f0fff9;">
              <b>Cennik UE zapisany w Supabase: {{ eur_imported }} pozycji.</b>
            </div>
          {% endif %}
          {% if eur_import_error %}
            <div class="card" style="border-color:#f3b8b8;background:#fff4f4;">
              <b>Cennik zapisano lokalnie ({{ eur_local_saved }} pozycji), ale Supabase odrzucił synchronizację.</b>
              <div class="muted" style="margin-top:8px;word-break:break-word;">{{ eur_import_error }}</div>
            </div>
          {% endif %}
          <div class="card">
            <h1>Cennik</h1>
            <div class="muted">Import pliku cen (kolumny: model, netto, brutto). ObsĹ‚uga CSV i XLSX (jeĹ›li dostÄ™pny openpyxl).</div>
          </div>

          <div class="card">
            <h2>Import cennika</h2>
            <form method="post" action="{{ url_for('pricing_import') }}" enctype="multipart/form-data" class="row">
              <div>
                <input type="file" name="file" accept=".csv,.xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/csv" required>
              </div>
              <div class="flex" style="align-items:flex-end;">
                <button class="btn primary" type="submit">Importuj cennik</button>
              </div>
            </form>
          </div>

          <div class="card">
            <h2>Cennik UE — EUR</h2>
            <div class="muted" style="margin-bottom:12px;">
              Import arkusza Preisliste.xlsx. Wymagane kolumny: Articel (SKU), PREIS EUR i UVP; GTIN/EAN jest opcjonalny.
              Import nie zmienia polskiego cennika ani stanów magazynowych.
            </div>
            <form method="post" action="{{ url_for('pricing_eur_import') }}" enctype="multipart/form-data" class="row">
              <div>
                <input type="file" name="file" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" required>
              </div>
              <div class="flex" style="align-items:flex-end;">
                <button class="btn primary" type="submit">Importuj cennik UE</button>
              </div>
            </form>
            <div class="muted small" style="margin-top:10px;">Aktualnie zapisanych pozycji: {{ eur_rows|length }}</div>
          </div>

          <div class="card">
            <form method="get" class="grid3" style="margin-bottom:10px;">
              <input name="q" value="{{ q }}" placeholder="Szukaj modelu">
              <button class="btn primary" type="submit">Szukaj</button>
              <a class="btn" href="{{ url_for('pricing') }}">WyczyĹ›Ä‡</a>
            </form>
            <h2>Pozycje cennika</h2>
            <table>
              <thead><tr><th>Model</th><th>Netto</th><th>Brutto</th></tr></thead>
              <tbody>
                {% for r in rows %}
                  <tr>
                    <td><b>{{ r['model'] }}</b></td>
                    <td>{{ "%.2f"|format(r['net_price']) }}</td>
                    <td>{{ "%.2f"|format(r['gross_price']) }}</td>
                  </tr>
                {% endfor %}
                {% if not rows %}
                  <tr><td colspan="3" class="muted">Brak pozycji cennika.</td></tr>
                {% endif %}
              </tbody>
            </table>
          </div>

          <div class="card">
            <h2>Pozycje cennika UE</h2>
            <table>
              <thead><tr><th>SKU</th><th>EAN</th><th>Cena EUR</th><th>UVP EUR</th></tr></thead>
              <tbody>
                {% for r in eur_rows %}
                  <tr>
                    <td><b>{{ r['sku'] }}</b></td>
                    <td>{{ r['ean'] or '-' }}</td>
                    <td>{{ "%.2f"|format(r['price_eur']) }} EUR</td>
                    <td>{{ "%.2f"|format(r['uvp_eur']) }} EUR</td>
                  </tr>
                {% endfor %}
                {% if not eur_rows %}
                  <tr><td colspan="4" class="muted">Cennik UE nie został jeszcze zaimportowany.</td></tr>
                {% endif %}
              </tbody>
            </table>
          </div>
        {% endblock %}
        """
        return render_template_string(
            tpl,
            title="Cennik",
            base_url=BASE_URL,
            db_path=DB_PATH,
            rows=rows,
            eur_rows=eur_rows,
            q=q,
            eur_imported=eur_imported,
            eur_import_error=eur_import_error,
            eur_local_saved=eur_local_saved,
        )



    @app.post("/pricing/import")
    def pricing_import():
        f = request.files.get("file")
        if not f:
            return "Brak pliku", 400

        filename = norm(f.filename).lower()
        parsed_rows = []

        if filename.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except Exception:
                return "Brak biblioteki openpyxl do odczytu XLSX. UĹĽyj CSV albo doinstaluj openpyxl.", 400

            wb = load_workbook(f, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return "Pusty plik", 400
            headers = [norm(x) for x in rows[0]]
            data = rows[1:]
            i_model = guess_col(headers, ["model"])
            i_sku = guess_col(headers, ["sku", "symbol", "index", "indeks", "kod", "code"])
            i_name = guess_col(headers, ["nazwa", "name", "produkt", "product"])
            i_ean = guess_col(headers, ["ean", "gtin"])
            i_net = guess_col(headers, ["netto", "net", "cena netto"])
            i_gross = guess_col(headers, ["brutto", "gross", "cena brutto"])
            if i_model is None or i_net is None or i_gross is None:
                return "Plik musi mieÄ‡ kolumny: model, netto, brutto", 400
            for r in data:
                if not r:
                    continue
                model = norm(r[i_model]) if len(r) > i_model else ""
                if not model:
                    continue
                sku = norm(r[i_sku]) if i_sku is not None and len(r) > i_sku else model
                name = norm(r[i_name]) if i_name is not None and len(r) > i_name else ""
                ean = norm(r[i_ean]) if i_ean is not None and len(r) > i_ean else ""
                net = to_float(r[i_net] if len(r) > i_net else "", 0.0)
                gross = to_float(r[i_gross] if len(r) > i_gross else "", 0.0)
                parsed_rows.append((sku, model, name, ean, net, gross))

        else:
            raw = f.read()
            try:
                text = raw.decode("utf-8-sig")
            except Exception:
                text = raw.decode("latin2", errors="replace")
            sample = text[:5000]
            delim = ";" if sample.count(";") >= sample.count(",") else ","
            rdr = csv.reader(io.StringIO(text), delimiter=delim)
            rows = list(rdr)
            if not rows:
                return "Pusty plik", 400
            headers = rows[0]
            data = rows[1:]
            i_model = guess_col(headers, ["model"])
            i_sku = guess_col(headers, ["sku", "symbol", "index", "indeks", "kod", "code"])
            i_name = guess_col(headers, ["nazwa", "name", "produkt", "product"])
            i_ean = guess_col(headers, ["ean", "gtin"])
            i_net = guess_col(headers, ["netto", "net", "cena netto"])
            i_gross = guess_col(headers, ["brutto", "gross", "cena brutto"])
            if i_model is None or i_net is None or i_gross is None:
                return "Plik musi mieÄ‡ kolumny: model, netto, brutto", 400
            for r in data:
                if not r:
                    continue
                model = norm(r[i_model]) if len(r) > i_model else ""
                if not model:
                    continue
                sku = norm(r[i_sku]) if i_sku is not None and len(r) > i_sku else model
                name = norm(r[i_name]) if i_name is not None and len(r) > i_name else ""
                ean = norm(r[i_ean]) if i_ean is not None and len(r) > i_ean else ""
                net = to_float(r[i_net] if len(r) > i_net else "", 0.0)
                gross = to_float(r[i_gross] if len(r) > i_gross else "", 0.0)
                parsed_rows.append((sku, model, name, ean, net, gross))

        c = conn()
        cur = c.cursor()
        changed_product_ids = []
        for sku, model, name, ean, net, gross in parsed_rows:
            cur.execute("""
              INSERT INTO pricing(model, net_price, gross_price, created_at)
              VALUES(?,?,?,?)
              ON CONFLICT(model) DO UPDATE SET
                net_price=excluded.net_price,
                gross_price=excluded.gross_price,
                created_at=excluded.created_at
            """, (model, net, gross, now_iso()))
            if sku:
                cur.execute("SELECT id FROM products WHERE sku=? LIMIT 1", (sku,))
                existing = cur.fetchone()
                if existing:
                    cur.execute("""
                      UPDATE products
                      SET model=COALESCE(NULLIF(?, ''), model),
                          ean=COALESCE(NULLIF(?, ''), ean),
                          name=COALESCE(NULLIF(?, ''), name),
                          archived=0
                      WHERE sku=?
                    """, (model, ean, name, sku))
                    pid = int(existing["id"])
                else:
                    cur.execute(
                        "INSERT INTO products(sku, model, ean, name, created_at) VALUES (?,?,?,?,?)",
                        (sku, model, ean, name, now_iso())
                    )
                    pid = int(cur.lastrowid)
                changed_product_ids.append(pid)
                cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
        c.commit()
        c.close()
        if supabase_enabled():
            try:
                sync_local_table_to_supabase("pricing", "model")
            except Exception:
                pass
            try:
                sync_local_rows_to_supabase("products", "id", changed_product_ids)
            except Exception:
                pass
            try:
                sync_local_rows_to_supabase("stock", "product_id", changed_product_ids)
            except Exception:
                pass
        return redirect(url_for("pricing"))




    @app.post("/pricing/eur/import")
    def pricing_eur_import():
        uploaded = request.files.get("file")
        if not uploaded or not norm(uploaded.filename).lower().endswith(".xlsx"):
            return "Wybierz plik XLSX z cennikiem UE", 400
        try:
            parsed_rows = parse_eur_pricing_xlsx(uploaded)
        except ValueError as exc:
            return str(exc), 400

        timestamp = now_iso()
        c = conn()
        try:
            cur = c.cursor()
            for sku, ean, price_eur, uvp_eur in parsed_rows:
                cur.execute(
                    """
                    INSERT INTO pricing_eur(sku, ean, price_eur, uvp_eur, created_at, updated_at)
                    VALUES(?,?,?,?,?,?)
                    ON CONFLICT(sku) DO UPDATE SET
                      ean=excluded.ean,
                      price_eur=excluded.price_eur,
                      uvp_eur=excluded.uvp_eur,
                      updated_at=excluded.updated_at
                    """,
                    (sku, ean, price_eur, uvp_eur, timestamp, timestamp),
                )
            c.commit()
        finally:
            c.close()

        if supabase_enabled():
            try:
                sync_local_table_to_supabase("pricing_eur", "sku")
            except Exception as exc:
                error_detail = norm(str(exc))[:600] or type(exc).__name__
                app.logger.exception("Nie udało się zsynchronizować cennika UE z Supabase")
                return redirect(url_for(
                    "pricing",
                    eur_local_saved=len(parsed_rows),
                    eur_import_error=error_detail,
                ))
        return redirect(url_for("pricing", eur_imported=len(parsed_rows)))




    @app.get("/products")
    def products():
        maybe_pull_shared_from_supabase()
        q = norm(request.args.get("q"))
        c = conn()
        cur = c.cursor()
        if q:
            like = f"%{q}%"
            cur.execute("""
              SELECT p.*, COALESCE(s.qty,0) AS stock
              FROM products p
              LEFT JOIN stock s ON s.product_id=p.id
              WHERE COALESCE(p.archived,0)=0
                AND (p.sku LIKE ? OR p.model LIKE ? OR p.ean LIKE ? OR p.name LIKE ?)
              ORDER BY p.sku
              LIMIT 1000
            """, (like, like, like, like))
        else:
            cur.execute("""
              SELECT p.*, COALESCE(s.qty,0) AS stock
              FROM products p
              LEFT JOIN stock s ON s.product_id=p.id
              WHERE COALESCE(p.archived,0)=0
              ORDER BY p.sku
              LIMIT 1000
            """)
        rows = cur.fetchall()
        c.close()

        tpl = r"""
        {% extends "base.html" %}
        {% block content %}
          <div class="card">
            <div class="flex">
              <h1 style="margin:0;">Produkty</h1>
              <div class="right"></div>
            </div>
            {% if request.args.get('product_deleted') %}<div class="notice" style="margin-top:10px;color:#067a2d;">Usunięto produkt {{ request.args.get('product_deleted') }}.</div>{% endif %}
            {% if request.args.get('product_error') %}<div class="notice" style="margin-top:10px;color:#b00020;">{{ request.args.get('product_error') }}</div>{% endif %}
            <form method="get" class="grid3" style="margin-top:10px;">
              <input name="q" value="{{ q }}" placeholder="Szukaj: SKU / model / EAN / nazwa">
              <button class="btn primary" type="submit">Szukaj</button>
              <a class="btn" href="{{ url_for('products') }}">WyczyĹ›Ä‡</a>
            </form>
          </div>

          <div class="card">
            <h2>Import CSV (478 pozycji)</h2>
            <div class="muted">Wybierz plik CSV z Excela. Minimalnie: kolumna SKU (unikalna). PozostaĹ‚e: model, ean, name/nazwa.</div>
            <form method="post" action="{{ url_for('products_import') }}" enctype="multipart/form-data" class="row" style="margin-top:10px;">
              <div>
                <input type="file" name="file" accept=".csv,text/csv" required>
                <div class="muted small" style="margin-top:6px;">Kodowanie: najlepiej UTF-8. Separator zwykle â€ž;â€ť lub â€ž,â€ť â€“ program sam sprĂłbuje.</div>
              </div>
              <div class="flex" style="align-items:flex-end;">
                <button class="btn primary" type="submit">Importuj</button>
              </div>
            </form>
          </div>

          <div class="card">
            <h2>Lista (max 1000)</h2>
            <table>
              <thead>
                <tr>
                  <th>SKU</th>
                  <th>Model</th>
                  <th>EAN</th>
                  <th>Nazwa</th>
                  <th>Stan</th>
                  <th>Akcje</th>
                </tr>
              </thead>
              <tbody>
                {% for r in rows %}
                <tr>
                  <td><b>{{ r["sku"] }}</b></td>
                  <td>{{ r["model"] or "" }}</td>
                  <td>{{ r["ean"] or "" }}</td>
                  <td>{{ r["name"] or "" }}</td>
                  <td><span class="badge">{{ r["stock"] }}</span></td>
                  <td>
                    <form method="post" action="{{ url_for('product_delete', product_id=r['id']) }}" onsubmit="return confirm('Usunąć wybrany produkt? Tej operacji nie można cofnąć.');">
                      <button class="btn danger" type="submit">Usuń</button>
                    </form>
                  </td>
                </tr>
                {% endfor %}
                {% if not rows %}
                  <tr><td colspan="5" class="muted">Brak produktĂłw. ZrĂłb import CSV.</td></tr>
                {% endif %}
              </tbody>
            </table>
          </div>
        {% endblock %}
        """
        return render_template_string(tpl, title="Produkty", base_url=BASE_URL, db_path=DB_PATH, rows=rows, q=q)



    @app.post("/products/<int:product_id>/delete")
    def product_delete(product_id):
        c = conn()
        cur = c.cursor()
        cur.execute("""
          SELECT p.id, p.sku, COALESCE(s.qty,0) AS stock
          FROM products p
          LEFT JOIN stock s ON s.product_id=p.id
          WHERE p.id=?
        """, (product_id,))
        product = cur.fetchone()
        if not product:
            c.close()
            return redirect(url_for("products", product_error="Nie znaleziono produktu."))

        # Usuwanie z katalogu jest archiwizacją produktu. Historyczne dokumenty
        # nadal wskazują ten sam stabilny identyfikator, ale produkt nie jest już
        # dostępny w magazynie, panelu klienta ani formularzach nowych dokumentów.
        active_references = []
        cur.execute("""
          SELECT COUNT(*) AS n
          FROM order_items oi
          JOIN orders o ON o.id=oi.order_id
          WHERE oi.product_id=?
            AND LOWER(COALESCE(o.status,'')) IN
                ('new','pending','unconfirmed','confirmed','packed','in_delivery','shipped')
            AND COALESCE(o.warehouse_issued,0)=0
        """, (product_id,))
        if int(cur.fetchone()["n"] or 0) > 0:
            active_references.append("aktywnych zamówieniach")
        cur.execute("""
          SELECT COUNT(*) AS n
          FROM china_items ci
          JOIN china_packages cp ON cp.id=ci.package_id
          WHERE ci.product_id=?
            AND LOWER(COALESCE(cp.status,'')) IN ('planned','ordered','shipped')
        """, (product_id,))
        if int(cur.fetchone()["n"] or 0) > 0:
            active_references.append("aktywnych dostawach P/O")
        if int(product["stock"] or 0) != 0:
            active_references.append("niezerowym stanie magazynowym")

        sku = product["sku"]
        if active_references:
            c.close()
            return redirect(url_for(
                "products",
                q=sku,
                product_error="Nie można usunąć produktu, ponieważ jest używany w " + ", ".join(active_references) + ".",
            ))

        try:
            if supabase_enabled():
                # Kolejność ma znaczenie: najpierw ukrycie produktu, następnie dane
                # aktywnego katalogu. Stare order_items i invoice_allocations zostają.
                supabase_update_rows("products", {"archived": True}, {"id": product_id})
                supabase_delete_rows("stock", {"product_id": product_id})
                try:
                    supabase_delete_rows("pricing_eur", {"sku": sku})
                except Exception:
                    app.logger.warning("Nie udało się usunąć ceny EUR dla SKU %s", sku, exc_info=True)

            cur.execute("UPDATE products SET archived=1 WHERE id=?", (product_id,))
            cur.execute("DELETE FROM stock WHERE product_id=?", (product_id,))
            cur.execute("DELETE FROM pricing_eur WHERE sku=?", (sku,))
            c.commit()
        except Exception:
            c.rollback()
            c.close()
            app.logger.exception("Nie udało się zarchiwizować produktu %s", product_id)
            return redirect(url_for(
                "products",
                q=sku,
                product_error="Nie udało się usunąć produktu z aktywnego magazynu. Najpierw uruchom migrację Supabase.",
            ))
        c.close()
        return redirect(url_for("products", product_deleted=sku))

        references = []
        for table, label in (
            ("order_items", "zamówieniach"),
            ("china_items", "dostawach P/O"),
            ("invoice_allocations", "fakturach"),
        ):
            cur.execute(f"SELECT COUNT(*) AS n FROM {table} WHERE product_id=?", (product_id,))
            if int(cur.fetchone()["n"] or 0) > 0:
                references.append(label)
        if int(product["stock"] or 0) != 0:
            references.append("stanie magazynowym")

        sku = product["sku"]
        if references:
            c.close()
            return redirect(url_for(
                "products",
                q=sku,
                product_error="Nie można usunąć produktu, ponieważ jest używany w " + ", ".join(references) + ".",
            ))

        try:
            if supabase_enabled():
                supabase_delete_rows("stock", {"product_id": product_id})
                try:
                    supabase_delete_rows("products", {"id": product_id})
                except Exception:
                    cur.execute("SELECT * FROM stock WHERE product_id=?", (product_id,))
                    stock_row = cur.fetchone()
                    if stock_row:
                        supabase_upsert_rows("stock", [dict(stock_row)], "product_id")
                    raise
                try:
                    supabase_delete_rows("pricing_eur", {"sku": sku})
                except Exception:
                    # Brak osobnej ceny EUR nie może cofnąć poprawnego usunięcia
                    # produktu. Osierocona cena nie jest widoczna bez produktu.
                    app.logger.warning("Nie udało się usunąć ceny EUR dla SKU %s", sku, exc_info=True)

            cur.execute("DELETE FROM pricing_eur WHERE sku=?", (sku,))
            cur.execute("DELETE FROM stock WHERE product_id=?", (product_id,))
            cur.execute("DELETE FROM products WHERE id=?", (product_id,))
            c.commit()
        except Exception:
            c.rollback()
            c.close()
            app.logger.exception("Nie udało się bezpiecznie usunąć produktu %s", product_id)
            return redirect(url_for("products", q=sku, product_error="Nie udało się usunąć produktu. Sprawdź synchronizację magazynu."))
        c.close()
        return redirect(url_for("products", product_deleted=sku))



    @app.post("/products/import")
    def products_import():
        f = request.files.get("file")
        if not f:
            return "Brak pliku", 400

        raw = f.read()
        # SprĂłbuj UTF-8, jak nie pĂłjdzie to latin2
        try:
            text = raw.decode("utf-8-sig")
        except:
            text = raw.decode("latin2", errors="replace")

        # SprĂłbuj wykryÄ‡ delimiter
        sample = text[:5000]
        delim = ";" if sample.count(";") >= sample.count(",") else ","

        rdr = csv.reader(io.StringIO(text), delimiter=delim)
        rows = list(rdr)
        if not rows:
            return "Pusty CSV", 400

        headers = rows[0]
        data = rows[1:]

        i_sku = guess_col(headers, ["sku", "symbol", "index", "indeks", "kod", "code"])
        i_model = guess_col(headers, ["model", "model_uchwytu", "nazwa_modelu"])
        i_ean = guess_col(headers, ["ean", "gtin"])
        i_name = guess_col(headers, ["name", "nazwa", "produkt", "product"])

        if i_sku is None:
            return "CSV musi mieÄ‡ kolumnÄ™ SKU / Symbol / Indeks", 400

        c = conn()
        cur = c.cursor()
        added = 0
        updated = 0

        for row in data:
            if not row or len(row) <= i_sku:
                continue
            sku = norm(row[i_sku])
            if not sku:
                continue
            model = norm(row[i_model]) if i_model is not None and len(row) > i_model else ""
            ean = norm(row[i_ean]) if i_ean is not None and len(row) > i_ean else ""
            name = norm(row[i_name]) if i_name is not None and len(row) > i_name else ""

            cur.execute("SELECT id FROM products WHERE sku=?", (sku,))
            exists = cur.fetchone()
            if exists:
                cur.execute("UPDATE products SET model=?, ean=?, name=?, archived=0 WHERE sku=?", (model, ean, name, sku))
                updated += 1
                pid = exists["id"]
            else:
                cur.execute(
                    "INSERT INTO products(sku, model, ean, name, created_at) VALUES (?,?,?,?,?)",
                    (sku, model, ean, name, now_iso())
                )
                pid = cur.lastrowid
                added += 1

            cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))

        c.commit()
        changed_ids = [int(r["id"]) for r in cur.execute(
            "SELECT id FROM products WHERE COALESCE(archived,0)=0"
        ).fetchall()]
        c.close()

        if supabase_enabled() and changed_ids:
            try:
                sync_local_rows_to_supabase("products", "id", changed_ids)
                sync_local_rows_to_supabase("stock", "product_id", changed_ids)
            except Exception:
                app.logger.warning("Nie udało się zsynchronizować importu produktów", exc_info=True)

        return redirect(url_for("products", q=""))




    @app.get("/stock")
    def stock():
        # Przy zwykłym użyciu synchronizacja działa w tle. Jedynie po zimnym
        # starcie z pustą lokalną bazą istniejący mechanizm wykonuje jednorazowy
        # bootstrap przed renderem, aby nie pokazywać fałszywych zer.
        maybe_pull_shared_from_supabase()
        q = norm(request.args.get("q"))
        active_filter = norm(request.args.get("filter")) or "all"
        page = max(1, to_int(request.args.get("page"), 1))
        per_page = to_int(request.args.get("per_page"), 25)
        if per_page not in (25, 50, 100):
            per_page = 25
        rows = build_replenishment_analysis(conn, today=app_now().date(), horizon_days=60)
        all_rows = rows
        for row in rows:
            available = to_int(row.get("available_qty"), 0)
            incoming = to_int(row.get("incoming_qty"), 0)
            reserved = to_int(row.get("reserved_qty"), 0)
            stock_qty = to_int(row.get("stock_qty"), 0)
            if reserved > stock_qty + incoming:
                row["status_label"], row["status_class"] = "Problem", "inv-red"
            elif available <= 0 and incoming > 0:
                row["status_label"], row["status_class"] = "Tylko w drodze", "inv-blue"
            elif available <= 0:
                row["status_label"], row["status_class"] = "Brak", "inv-red"
            elif available <= 5:
                row["status_label"], row["status_class"] = "Niski stan", "inv-orange"
            elif reserved > 0:
                row["status_label"], row["status_class"] = "Zarezerwowany", "inv-orange"
            else:
                row["status_label"], row["status_class"] = "OK", "inv-green"

        counts = {
            "missing": sum(1 for r in all_rows if to_int(r.get("available_qty"), 0) == 0),
            "low": sum(1 for r in all_rows if 0 < to_int(r.get("available_qty"), 0) <= 5),
            "incoming": sum(1 for r in all_rows if to_int(r.get("incoming_qty"), 0) > 0),
            "reserved": sum(1 for r in all_rows if to_int(r.get("reserved_qty"), 0) > 0),
            "problem": sum(1 for r in all_rows if to_int(r.get("reserved_qty"), 0) > to_int(r.get("stock_qty"), 0) + to_int(r.get("incoming_qty"), 0)),
        }
        if q:
            query = q.casefold()
            rows = [
                row for row in rows
                if any(query in norm(row.get(field)).casefold() for field in ("sku", "model", "ean", "name"))
            ]
        predicates = {
            "missing": lambda r: to_int(r.get("available_qty"), 0) == 0,
            "low": lambda r: 0 < to_int(r.get("available_qty"), 0) <= 5,
            "incoming": lambda r: to_int(r.get("incoming_qty"), 0) > 0,
            "reserved": lambda r: to_int(r.get("reserved_qty"), 0) > 0,
            "unreserved": lambda r: to_int(r.get("reserved_qty"), 0) == 0,
            "problem": lambda r: to_int(r.get("reserved_qty"), 0) > to_int(r.get("stock_qty"), 0) + to_int(r.get("incoming_qty"), 0),
        }
        if active_filter in predicates:
            rows = [row for row in rows if predicates[active_filter](row)]
        rows = sorted(rows, key=lambda row: norm(row.get("sku")).casefold())

        c = conn()
        image_by_product = {int(r["product_id"]): int(r["image_id"]) for r in c.execute(
            "SELECT product_id, image_id FROM product_image_assignments"
        ).fetchall()}
        c.close()
        for row in rows:
            row["image_id"] = image_by_product.get(int(row["id"]))

        total = len(rows)
        pages = max(1, (total + per_page - 1) // per_page)
        page = min(page, pages)
        start = (page - 1) * per_page
        rows = rows[start:start + per_page]
        def page_url(number):
            return url_for("stock", q=q, filter=active_filter, per_page=per_page, page=number)
        page_numbers = sorted(set(n for n in (1, page-1, page, page+1, pages) if 1 <= n <= pages))
        kpis = [
            {"label":"Dostępne od ręki", "value":f"{sum(to_int(r.get('available_qty'),0) for r in all_rows)} szt.", "note":"Fizycznie dostępne", "url":url_for("stock")},
            {"label":"W drodze (Chiny)", "value":f"{sum(to_int(r.get('incoming_qty'),0) for r in all_rows)} szt.", "note":"W aktywnych dostawach", "url":url_for("stock", filter="incoming")},
            {"label":"Zarezerwowane", "value":f"{sum(to_int(r.get('reserved_qty'),0) for r in all_rows)} szt.", "note":"W aktywnych zamówieniach", "url":url_for("stock", filter="reserved")},
            {"label":"Niski stan", "value":f"{counts['low']} SKU", "note":"1–5 dostępnych", "url":url_for("stock", filter="low")},
            {"label":"Braki", "value":f"{counts['missing']} SKU", "note":"0 dostępnych", "url":url_for("stock", filter="missing")},
            {"label":"Wszystkie produkty", "value":len(all_rows), "note":"Aktywne SKU", "url":url_for("stock")},
        ]
        alerts = [
            {"value":counts["missing"], "label":"brak dostępnych sztuk", "url":url_for("stock", filter="missing")},
            {"value":counts["low"], "label":"niski stan", "url":url_for("stock", filter="low")},
            {"value":counts["problem"], "label":"rezerwacje większe niż zapas", "url":url_for("stock", filter="problem")},
            {"value":sum(1 for r in all_rows if to_int(r.get("available_qty"),0)==0 and to_int(r.get("incoming_qty"),0)>0), "label":"dostępne tylko w drodze", "url":url_for("stock", filter="incoming")},
        ]
        return render_template(
            "stock.html",
            title="Magazyn",
            base_url=BASE_URL,
            db_path=DB_PATH,
            rows=rows,
            q=q,
            active_filter=active_filter,
            filters=(("missing","Braki"),("low","Niski stan"),("incoming","W drodze"),("reserved","Z rezerwacją"),("unreserved","Bez rezerwacji")),
            kpis=kpis, alerts=alerts, page=page, pages=pages, per_page=per_page,
            total=total, first_item=(start+1 if total else 0), last_item=min(start+per_page,total),
            page_numbers=page_numbers, page_url=page_url,
        )



    @app.post("/api/stock_delta")
    def api_stock_delta():
        data = request.get_json(force=True, silent=True) or {}
        sku = norm(data.get("sku"))
        delta_raw = norm(data.get("delta"))

        if not sku:
            return jsonify(ok=False, error="Brak SKU"), 400

        delta = to_int(delta_raw, None)
        if delta is None:
            # sprĂłbuj +10 / -3
            try:
                delta = int(delta_raw)
            except:
                return jsonify(ok=False, error="NieprawidĹ‚owa zmiana (np. +10 lub -3)"), 400

        c = conn()
        cur = c.cursor()
        cur.execute("SELECT id FROM products WHERE sku=? AND COALESCE(archived,0)=0", (sku,))
        p = cur.fetchone()
        if not p:
            c.close()
            return jsonify(ok=False, error="Nie ma takiego SKU"), 404
        pid = p["id"]
        cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
        cur.execute("UPDATE stock SET qty = qty + ? WHERE product_id=?", (delta, pid))
        cur.execute("SELECT qty FROM stock WHERE product_id=?", (pid,))
        new_qty = cur.fetchone()["qty"]
        c.commit()
        c.close()
        return jsonify(ok=True, new_qty=new_qty)


    @app.get("/api/stock/autocomplete")
    def stock_autocomplete():
        q = norm(request.args.get("q"))
        if len(q) < 2:
            return jsonify(ok=True, results=[])
        like = f"%{q}%"
        c = conn()
        rows = c.execute("""
          SELECT p.id,p.sku,p.model,p.name,p.ean,COALESCE(s.qty,0) stock
          FROM products p LEFT JOIN stock s ON s.product_id=p.id
          WHERE COALESCE(p.archived,0)=0
            AND (p.sku LIKE ? OR p.model LIKE ? OR p.name LIKE ? OR p.ean LIKE ?)
          ORDER BY CASE WHEN p.sku LIKE ? THEN 0 ELSE 1 END,p.sku LIMIT 15
        """, (like,like,like,like,f"{q}%")).fetchall()
        ids = [int(r["id"]) for r in rows]
        reservations = {}
        if ids:
            marks = ",".join("?" for _ in ids)
            active = ("new","pending","unconfirmed","confirmed","packed","packed_partial","in_delivery","shipped","partially_shipped")
            status_marks = ",".join("?" for _ in active)
            reserved_rows = c.execute(f"""SELECT oi.product_id,
              COALESCE(SUM(MAX(0,oi.qty-COALESCE(a.allocated_qty,0))),0) reserved
              FROM order_items oi JOIN orders o ON o.id=oi.order_id
              LEFT JOIN (SELECT order_item_id,SUM(qty) allocated_qty FROM invoice_allocations GROUP BY order_item_id) a ON a.order_item_id=oi.id
              WHERE oi.product_id IN ({marks}) AND COALESCE(o.warehouse_issued,0)=0
                AND lower(COALESCE(o.status,'')) IN ({status_marks}) GROUP BY oi.product_id""", (*ids,*active)).fetchall()
            reservations = {int(x["product_id"]):int(x["reserved"] or 0) for x in reserved_rows}
        c.close()
        return jsonify(ok=True, results=[{
            "id":int(r["id"]), "sku":r["sku"], "model":r["model"], "name":r["name"],
            "ean":r["ean"], "stock":int(r["stock"] or 0),
            "reserved":reservations.get(int(r["id"]),0),
            "available":max(0,int(r["stock"] or 0)-reservations.get(int(r["id"]),0)),
        } for r in rows])


    @app.get("/api/stock/products/<int:product_id>")
    def stock_product_details(product_id):
        analysis = build_replenishment_analysis(conn, today=app_now().date(), horizon_days=60)
        product = next((r for r in analysis if int(r["id"]) == product_id), None)
        if not product:
            return jsonify(ok=False, error="Nie ma takiego produktu"), 404
        c = conn()
        image_row = c.execute(
            "SELECT image_id FROM product_image_assignments WHERE product_id=?",
            (product_id,),
        ).fetchone()
        incoming = [dict(r) for r in c.execute("""
          SELECT cp.package_no,cp.status,ci.qty,cp.created_at
          FROM china_items ci JOIN china_packages cp ON cp.id=ci.package_id
          WHERE ci.product_id=? AND lower(COALESCE(cp.status,'')) IN ('ordered','shipped','problem')
          ORDER BY cp.created_at DESC LIMIT 20
        """, (product_id,)).fetchall()]
        adjustments = [dict(r) for r in c.execute("""
          SELECT old_qty,new_qty,delta,mode,created_at FROM stock_adjustments
          WHERE product_id=? ORDER BY id DESC LIMIT 10
        """, (product_id,)).fetchall()]
        c.close()
        keys = ("id","sku","model","name","ean","stock_qty","reserved_qty","available_qty","incoming_qty","reserved_incoming","available_incoming","status_label")
        product_payload = {k:product.get(k) for k in keys}
        product_payload["image_url"] = url_for(
            "inventory_image", image_id=int(image_row["image_id"])
        ) if image_row else ""
        return jsonify(ok=True, product=product_payload, incoming=incoming, adjustments=adjustments)


    @app.post("/api/stock/correction")
    def stock_correction():
        data = request.get_json(silent=True) or {}
        product_id = to_int(data.get("product_id"), 0)
        mode = norm(data.get("mode"))
        quantity = to_int(data.get("quantity"), None)
        if not product_id or mode not in ("delta", "set") or quantity is None:
            return jsonify(ok=False, error="Niepełne dane korekty"), 400
        c = conn()
        try:
            c.execute("BEGIN IMMEDIATE")
            product = c.execute("SELECT sku FROM products WHERE id=? AND COALESCE(archived,0)=0", (product_id,)).fetchone()
            if not product:
                c.rollback(); return jsonify(ok=False, error="Nie ma takiego produktu"), 404
            c.execute("INSERT OR IGNORE INTO stock(product_id,qty) VALUES(?,0)", (product_id,))
            old_qty = int(c.execute("SELECT qty FROM stock WHERE product_id=?", (product_id,)).fetchone()["qty"])
            new_qty = old_qty + quantity if mode == "delta" else quantity
            if new_qty < 0:
                c.rollback(); return jsonify(ok=False, error="Stan nie może być ujemny"), 400
            c.execute("UPDATE stock SET qty=? WHERE product_id=?", (new_qty,product_id))
            c.execute("INSERT INTO stock_adjustments(product_id,old_qty,new_qty,delta,mode,created_at) VALUES(?,?,?,?,?,?)",
                      (product_id,old_qty,new_qty,new_qty-old_qty,mode,now_iso()))
            c.commit()
        finally:
            c.close()
        if supabase_enabled():
            try:
                sync_local_rows_to_supabase("stock", "product_id", [product_id])
            except Exception:
                app.logger.warning("Korekta zapisana lokalnie, synchronizacja Supabase nieudana", exc_info=True)
        return jsonify(ok=True, sku=product["sku"], old_qty=old_qty, new_qty=new_qty)


    def _sanitize_svg(raw):
        import xml.etree.ElementTree as ET
        if len(raw) > 1024 * 1024:
            raise ValueError("Plik jest większy niż 1 MB")
        if b"<!DOCTYPE" in raw.upper() or b"<!ENTITY" in raw.upper():
            raise ValueError("Niedozwolona deklaracja SVG")
        try:
            root = ET.fromstring(raw)
        except ET.ParseError:
            raise ValueError("Nieprawidłowy plik SVG")
        if root.tag.split("}")[-1].lower() != "svg":
            raise ValueError("Plik nie jest SVG")
        forbidden = {"script","foreignobject","iframe","object","embed","audio","video"}
        for parent in list(root.iter()):
            for child in list(parent):
                if child.tag.split("}")[-1].lower() in forbidden:
                    parent.remove(child)
            for key in list(parent.attrib):
                short = key.split("}")[-1].lower()
                value = str(parent.attrib.get(key) or "").strip().lower()
                if short.startswith("on") or short in ("href","src") and (value.startswith("javascript:") or value.startswith("data:text/html")):
                    del parent.attrib[key]
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)


    def _prepare_product_image(upload):
        """Waliduje rzeczywistą zawartość i usuwa metadane z rastrów."""
        filename = norm(upload.filename)
        extension = os.path.splitext(filename.lower())[1]
        if extension not in (".svg", ".png", ".jpg", ".jpeg"):
            raise ValueError("Dozwolone formaty: SVG, PNG, JPG i JPEG")
        max_size = 5 * 1024 * 1024
        raw = upload.read(max_size + 1)
        if len(raw) > max_size:
            raise ValueError("Plik jest większy niż 5 MB")
        if extension == ".svg":
            return _sanitize_svg(raw), ".svg"
        try:
            from PIL import Image
            source = Image.open(io.BytesIO(raw))
            source.verify()
            source = Image.open(io.BytesIO(raw))
            if source.format not in ("PNG", "JPEG"):
                raise ValueError("Plik nie jest prawidłowym PNG lub JPG")
            if source.width * source.height > 25_000_000:
                raise ValueError("Obraz ma zbyt dużą rozdzielczość")
            output = io.BytesIO()
            if source.format == "JPEG":
                if source.mode not in ("RGB", "L"):
                    source = source.convert("RGB")
                source.save(output, format="JPEG", quality=88, optimize=True)
                return output.getvalue(), ".jpg"
            source.save(output, format="PNG", optimize=True)
            return output.getvalue(), ".png"
        except ValueError:
            raise
        except Exception:
            raise ValueError("Uszkodzony lub nieprawidłowy plik graficzny")


    @app.route("/api/stock/images", methods=["GET","POST"])
    def stock_images():
        c = conn()
        if request.method == "GET":
            rows = c.execute("""SELECT i.id,i.filename,COUNT(a.product_id) assignments,
              GROUP_CONCAT(p.sku, ', ') assigned_skus
              FROM product_images i LEFT JOIN product_image_assignments a ON a.image_id=i.id
              LEFT JOIN products p ON p.id=a.product_id
              GROUP BY i.id ORDER BY i.id DESC""").fetchall()
            c.close()
            return jsonify(ok=True, images=[{"id":int(r["id"]),"filename":r["filename"],"assignments":int(r["assignments"]),"assigned_skus":r["assigned_skus"] or "","url":url_for("inventory_image",image_id=r["id"])} for r in rows])
        upload = request.files.get("image")
        if not upload:
            c.close(); return jsonify(ok=False,error="Wybierz plik SVG, PNG albo JPG"),400
        try:
            cleaned, stored_extension = _prepare_product_image(upload)
        except ValueError as exc:
            c.close(); return jsonify(ok=False,error=str(exc)),400
        image_dir = os.path.join(os.path.dirname(DB_PATH), "product_images")
        os.makedirs(image_dir, exist_ok=True)
        stored_name = hashlib.sha256(cleaned).hexdigest() + stored_extension
        stored_path = os.path.join(image_dir, stored_name)
        if not os.path.exists(stored_path):
            with open(stored_path, "wb") as handle:
                handle.write(cleaned)
        if supabase_enabled():
            try:
                stored_path = supabase_storage_upload_file(
                    stored_path,
                    "product-images/" + stored_name,
                    content_type={".svg":"image/svg+xml",".png":"image/png",".jpg":"image/jpeg"}[stored_extension],
                )
            except Exception as exc:
                c.close()
                app.logger.exception("Nie udało się trwale zapisać zdjęcia produktu")
                return jsonify(ok=False,error="Nie udało się zapisać zdjęcia w Supabase Storage: " + str(exc)),502
        c.execute("INSERT OR IGNORE INTO product_images(stored_path,filename,created_at) VALUES(?,?,?)", (stored_path,os.path.basename(norm(upload.filename)),now_iso()))
        image_id = int(c.execute("SELECT id FROM product_images WHERE stored_path=?", (stored_path,)).fetchone()["id"])
        c.commit(); c.close()
        if supabase_enabled():
            try:
                sync_local_rows_to_supabase("product_images", "id", [image_id])
            except Exception as exc:
                return jsonify(ok=False,error="Plik zapisano, ale nie zapisano jego danych w Supabase. Uruchom migrację zdjęć: " + str(exc)),502
        return jsonify(ok=True,image_id=image_id)


    @app.get("/stock/images/<int:image_id>")
    def inventory_image(image_id):
        c = conn(); row = c.execute("SELECT stored_path FROM product_images WHERE id=?",(image_id,)).fetchone(); c.close()
        if not row:
            abort(404)
        storage_ref = parse_supabase_storage_ref(row["stored_path"])
        extension = os.path.splitext(storage_ref[1] if storage_ref else row["stored_path"])[1].lower()
        mimetype = {".svg":"image/svg+xml", ".png":"image/png", ".jpg":"image/jpeg", ".jpeg":"image/jpeg"}.get(extension)
        if not mimetype:
            abort(404)
        if storage_ref:
            try:
                image_bytes, filename = supabase_storage_download_bytes(row["stored_path"])
            except Exception:
                abort(404)
            response = send_file(io.BytesIO(image_bytes), download_name=filename, mimetype=mimetype, conditional=True, max_age=604800)
        else:
            if not os.path.isfile(row["stored_path"]):
                abort(404)
            response = send_file(row["stored_path"], mimetype=mimetype, conditional=True, max_age=604800)
        if extension == ".svg":
            response.headers["Content-Security-Policy"] = "default-src 'none'; style-src 'unsafe-inline'; sandbox"
        response.headers["X-Content-Type-Options"] = "nosniff"
        return response


    @app.get("/api/client/product-images/<int:image_id>")
    def client_product_image(image_id):
        # Endpoint przechodzi przez autoryzację klienta w security_gate.
        # Metadane czytamy z Supabase, aby działał także po zimnym starcie
        # Render, zanim lokalna kopia tabel zdjęć zostanie uzupełniona.
        rows = supabase_request(
            "/rest/v1/product_images",
            method="GET",
            params={"select": "stored_path", "id": f"eq.{image_id}", "limit": 1},
            timeout=20,
        ) or []
        if not rows:
            abort(404)
        stored_path = norm(rows[0].get("stored_path"))
        storage_ref = parse_supabase_storage_ref(stored_path)
        extension = os.path.splitext(storage_ref[1] if storage_ref else stored_path)[1].lower()
        mimetype = {".svg":"image/svg+xml", ".png":"image/png", ".jpg":"image/jpeg", ".jpeg":"image/jpeg"}.get(extension)
        if not mimetype or not storage_ref:
            abort(404)
        try:
            image_bytes, filename = supabase_storage_download_bytes(stored_path)
        except Exception:
            abort(404)
        response = send_file(io.BytesIO(image_bytes), download_name=filename, mimetype=mimetype, conditional=True, max_age=604800)
        if extension == ".svg":
            response.headers["Content-Security-Policy"] = "default-src 'none'; style-src 'unsafe-inline'; sandbox"
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["Cache-Control"] = "private, max-age=604800"
        return response


    @app.get("/api/stock/images/<int:image_id>/products")
    def stock_image_products(image_id):
        q = norm(request.args.get("q")); like=f"%{q}%"
        c=conn(); rows=c.execute("""SELECT p.id,p.sku,p.model,p.name,p.ean,
          CASE WHEN a.image_id=? THEN 1 ELSE 0 END assigned
          FROM products p LEFT JOIN product_image_assignments a ON a.product_id=p.id
          WHERE COALESCE(p.archived,0)=0 AND (?='' OR p.sku LIKE ? OR p.model LIKE ? OR p.name LIKE ? OR p.ean LIKE ?)
          ORDER BY assigned DESC,p.sku LIMIT 100""",(image_id,q,like,like,like,like)).fetchall(); c.close()
        return jsonify(ok=True,results=[dict(r) for r in rows])


    @app.post("/api/stock/images/<int:image_id>/assign")
    def stock_image_assign(image_id):
        data=request.get_json(silent=True) or {}
        # Zapisujemy wyłącznie różnicę względem stanu wyświetlonego użytkownikowi.
        # Dzięki temu wyszukanie jednego nowego SKU nie usuwa wcześniejszych przypisań.
        if "assign_ids" in data or "unassign_ids" in data:
            ids=sorted(set(to_int(x,0) for x in data.get("assign_ids",[]) if to_int(x,0)>0))[:500]
            removed=sorted(set(to_int(x,0) for x in data.get("unassign_ids",[]) if to_int(x,0)>0))[:500]
        else:
            # Zgodność ze starszą wersją panelu.
            ids=sorted(set(to_int(x,0) for x in data.get("product_ids",[]) if to_int(x,0)>0))[:500]
            visible=sorted(set(to_int(x,0) for x in data.get("visible_ids",[]) if to_int(x,0)>0))[:500]
            removed=[pid for pid in visible if pid not in ids]
        c=conn()
        if not c.execute("SELECT 1 FROM product_images WHERE id=?",(image_id,)).fetchone():
            c.close(); return jsonify(ok=False,error="Nie ma takiego zdjęcia"),404
        if removed:
            marks=",".join("?" for _ in removed)
            c.execute(f"DELETE FROM product_image_assignments WHERE image_id=? AND product_id IN ({marks})",(image_id,*removed))
        for pid in ids:
            c.execute("INSERT INTO product_image_assignments(product_id,image_id,created_at) VALUES(?,?,?) ON CONFLICT(product_id) DO UPDATE SET image_id=excluded.image_id,created_at=excluded.created_at",(pid,image_id,now_iso()))
        c.commit(); c.close()
        if supabase_enabled():
            try:
                for pid in removed:
                    supabase_delete_rows("product_image_assignments", {"product_id":pid, "image_id":image_id})
                sync_local_rows_to_supabase("product_image_assignments", "product_id", ids)
            except Exception as exc:
                return jsonify(ok=False,error="Przypisania zapisano lokalnie, ale synchronizacja Supabase nie powiodła się: " + str(exc)),502
        c=conn(); assigned=int(c.execute("SELECT COUNT(*) n FROM product_image_assignments WHERE image_id=?",(image_id,)).fetchone()["n"]); c.close()
        return jsonify(ok=True,assigned=assigned,changed=len(ids)+len(removed))



    @app.get("/api/product/<int:product_id>")
    def api_product(product_id):
        c = conn()
        cur = c.cursor()
        cur.execute("""
          SELECT p.*, COALESCE(s.qty,0) AS stock
          FROM products p
          LEFT JOIN stock s ON s.product_id=p.id
          WHERE p.id=? AND COALESCE(p.archived,0)=0
        """, (product_id,))
        r = cur.fetchone()
        c.close()
        if not r:
            return jsonify(ok=False), 404
        return jsonify(ok=True, id=r["id"], sku=r["sku"], model=r["model"], ean=r["ean"], name=r["name"], stock=r["stock"])



    exported = {'pricing': pricing, 'pricing_import': pricing_import, 'pricing_eur_import': pricing_eur_import, 'products': products, 'product_delete': product_delete, 'products_import': products_import, 'stock': stock, 'api_stock_delta': api_stock_delta, 'api_product': api_product}
    globals().update(exported)
    return exported
